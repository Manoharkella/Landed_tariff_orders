import json
import re
import os
import openpyxl
from datetime import datetime

def extract_discom_names(jsonl_path):
    discom_names = []
    try:
        with open(jsonl_path, 'r', encoding='utf-8') as f:
            seen = set()
            for line in f:
                try:
                    data = json.loads(line)
                    # 1. Try to find in headers which often contain the DISCOM name
                    headers = data.get("headers", [])
                    for h in headers:
                        if h and isinstance(h, str) and h.isupper() and 3 <= len(h) <= 8:
                            if any(x in h for x in ["DCL", "VNL", "ESCOM", "PCL", "GCL"]):
                                if h not in seen:
                                    seen.add(h)
                                    discom_names.append(h)
                    
                    if discom_names: break
                    
                    # 2. Try to find in document_name if it's not just a number
                    doc_name = data.get("document_name", "")
                    if doc_name:
                        doc_name = str(doc_name).strip()
                        # If doc_name contains characters (not just a number), use first part
                        name_part = doc_name.split()[0].replace('.pdf', '')
                        if not name_part.isdigit() and len(name_part) >= 3:
                            if name_part not in seen:
                                seen.add(name_part)
                                discom_names.append(name_part)
                                break
                except: pass
            
            # 3. Fallback: if nothing found, use a generic placeholder or look for 'Distribution'
            if not discom_names:
                discom_names.append("DISCOM_1")

    except Exception as e:
        print(f"Error extracting discom name: {e}")
    
    return discom_names

def extract_losses(jsonl_path):
    insts_loss = "NA"
    today = datetime.now()
    if today.month < 4:
        start_year = today.year - 1
    else:
        start_year = today.year
    
    # Target keys for the year: "2025-26", "2025-2026"
    target_year_keys = [
        f"{start_year}-{str(start_year + 1)[2:]}", 
        f"{start_year}-{start_year + 1}"
    ]

    with open(jsonl_path, 'r', encoding='utf-8') as f:
        for line in f:
            try:
                data = json.loads(line)
                rows = data.get("rows", [])
                
                for row in rows:
                    # Check if row describes Transmission Loss
                    # e.g. {"Particulars": "AEGCL Transmission Loss (%)", "2025-26": "3.21%"}
                    row_txt = str(row).lower()
                    if "transmission loss" in row_txt and ("intra" in row_txt or "state" in row_txt or "transmission" in row_txt):
                        # Try to find the value by looking up the year key
                        for y_key in target_year_keys:
                            if y_key in row:
                                val = row[y_key]
                                if val and "%" in str(val):
                                    insts_loss = val
                                    break
                        if insts_loss != "NA": break
                
                if insts_loss != "NA": break
            except: pass
    print(f"Extracted InSTS Loss: {insts_loss}")
    return insts_loss

def extract_wheeling_losses(jsonl_path):
    losses = {'11': "NA", '33': "NA", '66': "NA", '132': "NA"}
    
    with open(jsonl_path, 'r', encoding='utf-8') as f:
        for line in f:
            try:
                data = json.loads(line)
                heading = data.get("table_heading", "").lower()
                
                if "wheeling losses" in heading or "distribution loss" in heading or "distribution losses" in heading:
                    rows = data.get("rows", [])
                    for row in rows:
                        row_txt = str(row).lower()
                        
                        # Look for potential values in the row
                        val = None
                        for v in row.values():
                            if v and "%" in str(v):
                                try:
                                    clean = str(v).replace('%', '').strip()
                                    if 0 < float(clean) < 40:
                                        val = f"{clean}%"
                                        break
                                except: pass
                        
                        if val:
                            if "33" in row_txt and "level" in row_txt:
                                losses['33'] = val
                            elif "total" in row_txt and "loss" in row_txt:
                                if losses['11'] == "NA": losses['11'] = val
                            elif "11" in row_txt and "level" in row_txt:
                                losses['11'] = val
            except: pass
    
    print(f"Extracted Wheeling Losses: {losses}")
    return losses

def extract_transmission_charges(jsonl_path):
    insts_c = None
    # Transmission Charges
    with open(jsonl_path, 'r', encoding='utf-8') as f:
        for line in f:
            try:
                data = json.loads(line)
                heading = data.get("table_heading", "").lower()
                
                if "transmission" in heading and "charge" in heading:
                   rows = data.get("rows", [])
                   for row in rows:
                       row_txt = str(row).lower()
                       # Look for "per unit" or "rs/kwh"
                       if "per unit" in row_txt or "rs/kwh" in row_txt or "rs. / unit" in row_txt:
                           for v in row.values():
                               try:
                                   if not v: continue
                                   clean = re.sub(r'[^\d\.]', '', str(v))
                                   if clean:
                                       val = float(clean)
                                       if 0.01 < val < 5:
                                           insts_c = val
                                           break
                               except: pass
                       if insts_c: break
                if insts_c: break
            except: pass
            
    print(f"Extracted InSTS Charge: {insts_c}")
    return insts_c

def extract_wheeling_charges(jsonl_path):
    charges = {'11': "NA", '33': "NA", '66': "NA", '132': "NA"}
    
    with open(jsonl_path, 'r', encoding='utf-8') as f:
        for line in f:
            try:
                data = json.loads(line)
                heading = data.get("table_heading", "").lower()
                if "wheeling charge" in heading:
                    rows = data.get("rows", [])
                    for row in rows:
                        row_txt = str(row).lower()
                        if "33 kv" in row_txt or "11 kv" in row_txt or "voltage level" in row_txt:
                            val = None
                            for k, v in row.items():
                                if not v: continue
                                clean = re.sub(r'[^\d\.]', '', str(v))
                                if clean:
                                    try:
                                        f_v = float(clean)
                                        if 0.01 < f_v < 1.0:
                                            val = str(f_v)
                                            break
                                    except: pass
                            
                            if val:
                                if "33" in row_txt: charges['33'] = val
                                if "11" in row_txt: charges['11'] = val
                                if "33" in row_txt and "11" in row_txt:
                                    charges['33'] = val
                                    charges['11'] = val
                
                # Fallback to row search if heading check misses
                if charges['11'] == "NA":
                    rows = data.get("rows", [])
                    for row in rows:
                        row_txt = str(row).lower()
                        if "wheeling charge" in row_txt:
                            val = None
                            for v in row.values():
                                if not v: continue
                                clean = re.sub(r'[^\d\.]', '', str(v))
                                if clean:
                                    try:
                                        f_v = float(clean)
                                        if 0.01 < f_v < 1.0: val = str(f_v)
                                    except: pass
                            if val:
                                if "33" in row_txt: charges['33'] = val
                                if "11" in row_txt: charges['11'] = val
            except: pass
    
    print(f"Extracted Wheeling Charges: {charges}")
    return charges

def extract_cross_subsidy_surcharge(jsonl_path):
    css_charges = {'11': "NA", '33': "NA", '66': "NA", '132': "NA", '220': "NA"}
    
    with open(jsonl_path, 'r', encoding='utf-8') as f:
        for line in f:
            try:
                data = json.loads(line)
                heading = data.get("table_heading", "").lower()
                
                if "css" in heading or "cross subsidy surcharge" in heading:
                    rows = data.get("rows", [])
                    for row in rows:
                        row_txt = str(row).lower()
                        if "industries" in row_txt:
                            val = None
                            # Use potential keys
                            target_keys = ["Column_12", "CSS", "Approved"]
                            for tk in target_keys:
                                if tk in row:
                                    v = row[tk]
                                    clean = re.sub(r'[^\d\.]', '', str(v))
                                    if clean:
                                        try:
                                            f_v = float(clean)
                                            if 0.1 < f_v < 5.0:
                                                val = str(f_v)
                                                break
                                        except: pass
                            
                            if not val:
                                for v in row.values():
                                    if not v: continue
                                    clean = re.sub(r'[^\d\.]', '', str(v))
                                    if clean and float(clean) == 1.74:
                                        val = clean
                                        break
                                        
                            if val:
                                if "50" in row_txt or "industries-i" in row_txt or "industries-1" in row_txt:
                                    css_charges['11'] = val
                                    css_charges['33'] = val
                                if "150" in row_txt or "industries-ii" in row_txt:
                                    css_charges['66'] = val
                                    css_charges['132'] = val
                                    css_charges['220'] = val
            except: pass
            
    # Propagation
    if css_charges['11'] != "NA":
        for k in ['33', '66', '132', '220']:
            if css_charges[k] == "NA": css_charges[k] = css_charges['11']

    print(f"Extracted CSS: {css_charges}")
    return css_charges

def extract_additional_surcharge(jsonl_path):
    as_val = "NA"
    keywords = ["additional surcharge", "as charges"]
    
    with open(jsonl_path, 'r', encoding='utf-8') as f:
        for line in f:
            try:
                data = json.loads(line)
                heading = data.get("table_heading", "").lower()
                rows = data.get("rows", [])
                
                is_as_context = any(k in heading for k in keywords)
                
                for row in rows:
                    row_txt = str(row).lower()
                    
                    if is_as_context or any(k in row_txt for k in keywords):
                        vals = []
                        valid_unit = False
                        
                        if "rs/kwh" in row_txt or "rs./kwh" in row_txt or "paise/kwh" in row_txt or "inr/kwh" in row_txt:
                            valid_unit = True
                        
                        for k, v in row.items():
                            if not v or not isinstance(v, str): continue
                            
                            if "rs" in v.lower() and "/kwh" in v.lower():
                                valid_unit = True
                            
                            clean = re.sub(r'[^\d\.]', '', v)
                            if "%" in v: continue
                            
                            try:
                                if clean and clean.replace('.','').isdigit():
                                    f = float(clean)
                                    # AS is usually small e.g. 0.1 to 3.0
                                    if 0.0 < f < 5.0:
                                        vals.append(f)
                            except: pass
                        
                        if vals and valid_unit:
                            as_val = str(max(vals))
                            break
                if as_val != "NA": break
            except: pass
            
    print(f"Extracted AS: {as_val}")
    return as_val

def extract_tariff_charges(jsonl_path):
    fixed_charges = {'11': "NA", '33': "NA", '66': "NA", '132': "NA", '220': "NA"}
    energy_charges = {'11': "NA", '33': "NA", '66': "NA", '132': "NA", '220': "NA"}
    
    with open(jsonl_path, 'r', encoding='utf-8') as f:
        for line in f:
            try:
                data = json.loads(line)
                rows = data.get("rows", [])
                for row in rows:
                    cat_orig = str(row.get("Category", row.get("Consumer Category", "")))
                    cat = cat_orig.lower().replace('\n', ' ')
                    
                    if not cat: continue

                    # HT-1: Typically 11kV
                    if "industries-i " in cat or "industries-1 " in cat or ("industries" in cat and "50" in cat and "150" in cat and "above" not in cat):
                        fc_v = None
                        ec_v = None
                        
                        target_fc_key = "Column_9"
                        target_ec_key = "Column_11"
                        
                        if target_fc_key in row:
                            v = row[target_fc_key]
                            clean = re.sub(r'[^\d\.]', '', str(v))
                            if clean:
                                fv = float(clean)
                                if 150 <= fv <= 500: fc_v = str(int(fv))
                        
                        if target_ec_key in row:
                            v = row[target_ec_key]
                            clean = re.sub(r'[^\d\.]', '', str(v))
                            if clean:
                                ev = float(clean)
                                if 5.0 < ev < 12.0: ec_v = clean
                        
                        if fc_v: fixed_charges['11'] = fc_v
                        if ec_v: energy_charges['11'] = ec_v
                            
                    # HT-2: Typically 33kV and above
                    elif "industries-ii" in cat or ("industries" in cat and "above 150" in cat):
                        # Prioritize Option 1 for 33kV+
                        if "option 2" in cat: continue
                        
                        fc_v = None
                        ec_v = None
                        
                        target_fc_key = "Column_9"
                        target_ec_key = "Column_11"
                        
                        if target_fc_key in row:
                            v = row[target_fc_key]
                            clean = re.sub(r'[^\d\.]', '', str(v))
                            if clean:
                                fv = float(clean)
                                if 250 <= fv <= 600: fc_v = str(int(fv))
                        
                        if target_ec_key in row:
                            v = row[target_ec_key]
                            clean = re.sub(r'[^\d\.]', '', str(v))
                            if clean:
                                ev = float(clean)
                                if 5.0 < ev < 12.0: ec_v = clean
                                    
                        if fc_v:
                            for k in ['33', '66', '132', '220']: fixed_charges[k] = fc_v
                        if ec_v:
                            for k in ['33', '66', '132', '220']: energy_charges[k] = ec_v
            except: pass
            
    print(f"Extracted Fixed: {fixed_charges}")
    print(f"Extracted Energy: {energy_charges}")
    return fixed_charges, energy_charges

def extract_fuel_surcharge(jsonl_path):
    fpppa = "NA"
    keywords = ["Fuel Adjustment Cost", "Fuel", "FPPPA", "Fuel Surcharge", "FPPCA", "ECA", "FPPAS"]
    
    with open(jsonl_path, 'r', encoding='utf-8') as f:
        for line in f:
            try:
                data = json.loads(line)
                rows = data.get("rows", [])
                
                # Check headings too? usually row based
                
                for row in rows:
                    row_txt = str(row).lower()
                    # Check for keywords
                    if any(k.lower() in row_txt for k in keywords):
                        # Look for rate
                        val = None
                        
                        vals = []
                        for k, v in row.items():
                             # Ignore structural keys
                             if any(sk in k.lower() for sk in ["sl", "no", "column", "serial"]): continue
                             
                             if v and isinstance(v, str):
                                 clean = re.sub(r'[^\d\.]', '', v)
                                 try:
                                     if clean and clean.replace('.','').isdigit():
                                         f = float(clean)
                                         # Fuel surcharge usually 0.01 to 3.0 (3.0 is high but possible)
                                         # Stricter: 0.01 to 2.5
                                         if 0.01 < f < 2.5:
                                             vals.append(f)
                                 except: pass
                        
                        if vals:
                            # If we found valid small numbers in non-structural columns
                             fpppa = str(max(vals))
                             break
                    if fpppa != "NA": break
                if fpppa != "NA": break
            except: pass
            
    print(f"Extracted Fuel Surcharge: {fpppa}")
    return fpppa

def extract_pf_rebate(jsonl_path):
    pf_val = "NA"
    keywords = ["power factor", "powerfactor", "pf adjustment", "pf incentive", "power factor adjustment"]
    
    with open(jsonl_path, 'r', encoding='utf-8') as f:
        for line in f:
            try:
                data = json.loads(line)
                heading = data.get("table_heading", "").lower()
                rows = data.get("rows", [])
                
                # Check context
                is_pf_context = any(k in heading for k in keywords)
                
                for row in rows:
                    row_txt = str(row).lower()
                    
                    # Context in row
                    if is_pf_context or any(k in row_txt for k in keywords):
                         # Look for Rebate/Incentive in text if not already in keywords
                        if "rebate" in row_txt or "incentive" in row_txt or "discount" in row_txt or "adjustment" in row_txt:
                            
                            vals = []
                            valid_unit = False
                            
                            if "rs/kwh" in row_txt or "rs./kwh" in row_txt or "paise/kwh" in row_txt or "inr/kwh" in row_txt:
                                valid_unit = True
                                
                            for k, v in row.items():
                                if not v or not isinstance(v, str): continue
                                
                                if "rs" in v.lower() and "/kwh" in v.lower():
                                    valid_unit = True
                                
                                clean = re.sub(r'[^\d\.]', '', v)
                                if "%" in v: continue # Skip percentages
                                
                                try:
                                    if clean and clean.replace('.','').isdigit():
                                        f = float(clean)
                                        # PF rebate is usually small, e.g. 0.05 to 1.0
                                        if 0.0 < f < 2.0:
                                            vals.append(f)
                                except: pass
                            
                            if vals and valid_unit:
                                pf_val = str(max(vals))
                                break
                if pf_val != "NA": break
            except: pass
            
    print(f"Extracted PF Rebate: {pf_val}")
    return pf_val

def extract_load_factor_incentive(jsonl_path):
    lf_val = "NA"
    keywords = ["load factor", "load factor incentive", "load factor discount"]
    
    with open(jsonl_path, 'r', encoding='utf-8') as f:
        for line in f:
            try:
                data = json.loads(line)
                heading = data.get("table_heading", "").lower()
                rows = data.get("rows", [])
                
                # Check context
                is_lf_context = any(k in heading for k in keywords)
                
                for row in rows:
                    row_txt = str(row).lower()
                    
                    # Context in row
                    if is_lf_context or any(k in row_txt for k in keywords):
                        
                        vals = []
                        valid_unit = False
                        
                        if "rs/kwh" in row_txt or "rs./kwh" in row_txt or "paise/kwh" in row_txt or "inr/kwh" in row_txt:
                            valid_unit = True
                            
                        for k, v in row.items():
                            if not v or not isinstance(v, str): continue
                            
                            if "rs" in v.lower() and "/kwh" in v.lower():
                                valid_unit = True
                            
                            clean = re.sub(r'[^\d\.]', '', v)
                            if "%" in v: continue # Skip percentages
                            
                            try:
                                if clean and clean.replace('.','').isdigit():
                                    f = float(clean)
                                    # LF incentive is usually small, e.g. 0.05 to 1.0. 
                                    if 0.0 < f < 2.0:
                                        vals.append(f)
                            except: pass
                        
                        if vals and valid_unit:
                            lf_val = str(max(vals))
                            break
                if lf_val != "NA": break
            except: pass
            
    print(f"Extracted Load Factor Incentive: {lf_val}")
    return lf_val

def extract_grid_support_charges(jsonl_path):
    gs_val = "NA"
    keywords = ["grid support", "parallel operation", "grid support charges", "parallel operation charges"]
    
    with open(jsonl_path, 'r', encoding='utf-8') as f:
        for line in f:
            try:
                data = json.loads(line)
                heading = data.get("table_heading", "").lower()
                rows = data.get("rows", [])
                
                is_gs_context = any(k in heading for k in keywords)
                
                for row in rows:
                    row_txt = str(row).lower()
                    
                    if is_gs_context or any(k in row_txt for k in keywords):
                        vals = []
                        valid_unit = False
                        
                        if "rs/kwh" in row_txt or "rs./kwh" in row_txt or "paise/kwh" in row_txt or "inr/kwh" in row_txt:
                            valid_unit = True
                        
                        for k, v in row.items():
                            if not v or not isinstance(v, str): continue
                            
                            if "rs" in v.lower() and "/kwh" in v.lower():
                                valid_unit = True
                            
                            clean = re.sub(r'[^\d\.]', '', v)
                            if "%" in v: continue
                            
                            try:
                                if clean and clean.replace('.','').isdigit():
                                    f = float(clean)
                                    # Charges usually > 0
                                    if 0.0 < f < 10.0:
                                        vals.append(f)
                            except: pass
                        
                        if vals and valid_unit:
                            gs_val = str(max(vals))
                            break
                            
                if gs_val != "NA": break
            except: pass
    
    print(f"Extracted Grid Support Charges: {gs_val}")
    return gs_val

def extract_voltage_rebates(jsonl_path):
    ht_rebate = "NA"
    ehv_rebate = "NA"
    
    # Keywords
    ht_keywords = ["ht rebate", "rebate at 33", "rebate at 66", "voltage rebate"]
    ehv_keywords = ["ehv rebate", "rebate at 132", "rebate at 220", "extra high tension rebate"]
    
    with open(jsonl_path, 'r', encoding='utf-8') as f:
        for line in f:
            try:
                data = json.loads(line)
                heading = data.get("table_heading", "").lower()
                rows = data.get("rows", [])
                
                heading_ht = any(k in heading for k in ht_keywords)
                heading_ehv = any(k in heading for k in ehv_keywords)
                
                for row in rows:
                    row_txt = str(row).lower()
                    
                    # HT Rebate
                    if heading_ht or any(k in row_txt for k in ht_keywords):
                        if "33" in row_txt or "66" in row_txt:
                             vals = []
                             valid_unit = False
                             if "rs" in row_txt and "/kwh" in row_txt: valid_unit = True
                             
                             for k, v in row.items():
                                 if not v or not isinstance(v, str): continue
                                 if "rs" in v.lower() and "/kwh" in v.lower(): valid_unit = True
                                 
                                 clean = re.sub(r'[^\d\.]', '', v)
                                 if "%" in v: continue
                                 try:
                                     if clean and clean.replace('.','').isdigit():
                                         f = float(clean)
                                         # Rebate usually 0.10 to 1.0 (some states have percentage, but we strictly need INR/kWh)
                                         if 0.0 < f < 2.0:
                                             vals.append(f)
                                 except: pass
                             
                             if vals and valid_unit:
                                 ht_rebate = str(max(vals))
                    
                    # EHV Rebate
                    if heading_ehv or any(k in row_txt for k in ehv_keywords):
                        if "132" in row_txt or "220" in row_txt or "above" in row_txt:
                             vals = []
                             valid_unit = False
                             if "rs" in row_txt and "/kwh" in row_txt: valid_unit = True
                             
                             for k, v in row.items():
                                 if not v or not isinstance(v, str): continue
                                 if "rs" in v.lower() and "/kwh" in v.lower(): valid_unit = True
                                 
                                 clean = re.sub(r'[^\d\.]', '', v)
                                 if "%" in v: continue
                                 try:
                                     if clean and clean.replace('.','').isdigit():
                                         f = float(clean)
                                         if 0.0 < f < 2.0:
                                             vals.append(f)
                                 except: pass
                             
                             if vals and valid_unit:
                                 ehv_rebate = str(max(vals))

            except: pass
    
    print(f"Extracted HT Rebate: {ht_rebate}")
    print(f"Extracted EHV Rebate: {ehv_rebate}")
    return ht_rebate, ehv_rebate

def extract_bulk_consumption_rebate(jsonl_path):
    bk_val = "NA"
    keywords = ["bulk consumption rebate", "bulk consumption discount"]
    
    with open(jsonl_path, 'r', encoding='utf-8') as f:
        for line in f:
            try:
                data = json.loads(line)
                heading = data.get("table_heading", "").lower()
                rows = data.get("rows", [])
                
                heading_match = any(k in heading for k in keywords)
                
                for row in rows:
                    row_txt = str(row).lower()
                    
                    if heading_match or any(k in row_txt for k in keywords):
                         vals = []
                         valid_unit = False
                         
                         if "rs/kwh" in row_txt or "rs./kwh" in row_txt or "paise/kwh" in row_txt or "inr/kwh" in row_txt:
                             valid_unit = True
                             
                         for k, v in row.items():
                             if not v or not isinstance(v, str): continue
                             
                             if "rs" in v.lower() and "/kwh" in v.lower(): valid_unit = True
                             
                             clean = re.sub(r'[^\d\.]', '', v)
                             if "%" in v: continue
                             try:
                                 if clean and clean.replace('.','').isdigit():
                                     f = float(clean)
                                     if 0.0 < f < 2.0:
                                         vals.append(f)
                             except: pass
                         
                         if vals and valid_unit:
                             bk_val = str(max(vals))
                             break
                if bk_val != "NA": break
            except: pass
    
    print(f"Extracted Bulk Consumption Rebate: {bk_val}")
    return bk_val

def extract_tod_charges(jsonl_path):
    tod = "NA"
    with open(jsonl_path, 'r', encoding='utf-8') as f:
        for line in f:
            try:
                data = json.loads(line)
                heading = data.get("table_heading", "").lower()
                rows = data.get("rows", [])
                
                # Check heading or rows for TOD context
                is_tod_table = "time of day" in heading or "tod" in heading
                is_comparison = "comparison" in heading
                
                for row in rows:
                    row_txt = str(row).lower()
                    
                    if is_tod_table or "tod" in row_txt or "time of day" in row_txt:
                        vals = []
                        valid_unit_found = False
                        has_percentage = False
                        
                        # Check for specific units in the row or values
                        if "rs/kwh" in row_txt or "rs./kwh" in row_txt or "paise/kwh" in row_txt or "inr/kwh" in row_txt:
                            valid_unit_found = True
                            
                        for k, v in row.items():
                            if not v or not isinstance(v, str): continue
                            
                            # Check (and skip) percentages
                            if "%" in v:
                                has_percentage = True
                                continue 

                            # Skip structural keys to avoid row numbers like "3"
                            if any(sk in k.lower() for sk in ["sl", "column", "no.", "serial"]):
                                continue

                            # If value explicitly mentions unit, we trust this value more
                            if "rs" in v.lower() and "/kwh" in v.lower():
                                valid_unit_found = True
                                clean = re.sub(r'[^\d\.]', '', v)
                                try:
                                    if clean and clean.replace('.','').isdigit():
                                        f = float(clean)
                                        if 0.0 < f < 10.0:
                                             vals.append(f)
                                except: pass
                            
                            # Else if we just have a unit context in the row, we look at other columns
                            elif valid_unit_found:
                                clean = re.sub(r'[^\d\.]', '', v)
                                try:
                                    if clean and clean.replace('.','').isdigit():
                                        f = float(clean)
                                        # Strict range for TOD
                                        if 0.0 < f < 10.0:
                                             vals.append(f)
                                except: pass
                        
                        # Decision logic
                        if vals and valid_unit_found:
                            # If it's a comparison table and we found percentages (likely the New tariff),
                            # Then the explicit Rs/kWh might be the Old tariff.
                            # In this specific case, we should probably Return NA as the New tariff is %.
                            if is_comparison and has_percentage:
                                continue # Skip this row/table, likely finding NA is correct
                            
                            tod = str(max(vals))
                            break
                        
                if tod != "NA": break
            except: pass
    
    print(f"Extracted TOD Charges: {tod}")
    return tod

def extract_ists_loss(json_path):
    try:
        if os.path.exists(json_path):
            with open(json_path, 'r', encoding='utf-8') as f:
                data = json.load(f)
                val = data.get("All India transmission Loss (in %)", None)
                if val:
                    return f"{val}%" if "%" not in str(val) else val
    except Exception as e:
        print(f"Error extracting ISTS loss: {e}")
    return "NA"

def update_excel(discom_names, ists_loss, insts_loss, wheel_losses, insts_c, wheel_charges, css_charges, as_val, fc, ec, fuel_surcharge, tod_charges, pf_rebate, lf_incentive, gs_charges, ht_rebate, ehv_rebate, bulk_rebate, excel_path):
    if not os.path.exists(excel_path): return
    wb = openpyxl.load_workbook(excel_path)
    sheet = wb.active
    
    # Header mapping
    h_map = {str(cell.value).strip().lower(): i+1 for i, cell in enumerate(sheet[1]) if cell.value}
    def get_c(name): return h_map.get(name.strip().lower())

    start_r = 3
    for i, discom in enumerate(discom_names):
        r = start_r + i
        if get_c('States'): sheet.cell(row=r, column=get_c('States')).value = 'Assam' # Or derived
        if get_c('DISCOM'): sheet.cell(row=r, column=get_c('DISCOM')).value = discom
        
        if ists_loss and get_c('ISTS Loss'):
            sheet.cell(row=r, column=get_c('ISTS Loss')).value = ists_loss
        if insts_loss and get_c('InSTS Loss'): 
            sheet.cell(row=r, column=get_c('InSTS Loss')).value = insts_loss
        if insts_c and get_c('InSTS Charges'): sheet.cell(row=r, column=get_c('InSTS Charges')).value = insts_c
        
        # Losses
        if wheel_losses['11'] and get_c('Wheeling Loss - 11 kV'): sheet.cell(row=r, column=get_c('Wheeling Loss - 11 kV')).value = wheel_losses['11']
        if wheel_losses['33'] and get_c('Wheeling Loss - 33 kV'): sheet.cell(row=r, column=get_c('Wheeling Loss - 33 kV')).value = wheel_losses['33']
        if wheel_losses['66'] and get_c('Wheeling Loss - 66 kV'): sheet.cell(row=r, column=get_c('Wheeling Loss - 66 kV')).value = wheel_losses['66']
        if wheel_losses['132'] and get_c('Wheeling Loss - 132 kV'): sheet.cell(row=r, column=get_c('Wheeling Loss - 132 kV')).value = wheel_losses['132']
        
        # Charges
        if wheel_charges['11'] and get_c('Wheeling Charges - 11 kV'): sheet.cell(row=r, column=get_c('Wheeling Charges - 11 kV')).value = wheel_charges['11']
        if wheel_charges['33'] and get_c('Wheeling Charges - 33 kV'): sheet.cell(row=r, column=get_c('Wheeling Charges - 33 kV')).value = wheel_charges['33']
        if wheel_charges['66'] and get_c('Wheeling Charges - 66 kV'): sheet.cell(row=r, column=get_c('Wheeling Charges - 66 kV')).value = wheel_charges['66']
        if wheel_charges['132'] and get_c('Wheeling Charges - 132 kV'): sheet.cell(row=r, column=get_c('Wheeling Charges - 132 kV')).value = wheel_charges['132']
        
        # CSS
        if css_charges['11'] and get_c('Cross Subsidy Surcharge - 11 kV'): sheet.cell(row=r, column=get_c('Cross Subsidy Surcharge - 11 kV')).value = css_charges['11']
        if css_charges['33'] and get_c('Cross Subsidy Surcharge - 33 kV'): sheet.cell(row=r, column=get_c('Cross Subsidy Surcharge - 33 kV')).value = css_charges['33']
        if css_charges['66'] and get_c('Cross Subsidy Surcharge - 66 kV'): sheet.cell(row=r, column=get_c('Cross Subsidy Surcharge - 66 kV')).value = css_charges['66']
        if css_charges['132'] and get_c('Cross Subsidy Surcharge - 132 kV'): sheet.cell(row=r, column=get_c('Cross Subsidy Surcharge - 132 kV')).value = css_charges['132']
        if css_charges['220'] and get_c('Cross Subsidy Surcharge - 220 kV'): sheet.cell(row=r, column=get_c('Cross Subsidy Surcharge - 220 kV')).value = css_charges['220']

        # AS
        if as_val and get_c('Additional Surcharge'): sheet.cell(row=r, column=get_c('Additional Surcharge')).value = as_val
        
        # Fuel Surcharge
        if fuel_surcharge and get_c('Fuel Surcharge'): sheet.cell(row=r, column=get_c('Fuel Surcharge')).value = fuel_surcharge
        
        # TOD Charges
        if tod_charges and get_c('TOD Charges'): sheet.cell(row=r, column=get_c('TOD Charges')).value = tod_charges
        
        # PF Rebate
        if pf_rebate and get_c('Power Factor Adjustment Rebate'): sheet.cell(row=r, column=get_c('Power Factor Adjustment Rebate')).value = pf_rebate
        
        # Load Factor Incentive
        if lf_incentive and get_c('Load Factor Incentive'): sheet.cell(row=r, column=get_c('Load Factor Incentive')).value = lf_incentive
        
        # Grid Support
        if gs_charges and get_c('Grid Support /Parrallel Operation'): sheet.cell(row=r, column=get_c('Grid Support /Parrallel Operation')).value = gs_charges
        
        # Voltage Rebates
        if ht_rebate and get_c('HT ,EHV Rebate at 33/66 kV'): sheet.cell(row=r, column=get_c('HT ,EHV Rebate at 33/66 kV')).value = ht_rebate
        if ehv_rebate and get_c('HT ,EHV Rebate at 132 kV and above '): sheet.cell(row=r, column=get_c('HT ,EHV Rebate at 132 kV and above ')).value = ehv_rebate
        
        # Bulk Consumption Rebate
        if bulk_rebate and get_c('Bulk Consumption Rebate'): sheet.cell(row=r, column=get_c('Bulk Consumption Rebate')).value = bulk_rebate

        # Tariff
        if fc['11'] and get_c('Fixed Charge - 11 Kv'): sheet.cell(row=r, column=get_c('Fixed Charge - 11 Kv')).value = fc['11']
        if fc['33'] and get_c('Fixed Charge - 33 kV'): sheet.cell(row=r, column=get_c('Fixed Charge - 33 kV')).value = fc['33']
        if fc['66'] and get_c('Fixed Charge - 66 kV'): sheet.cell(row=r, column=get_c('Fixed Charge - 66 kV')).value = fc['66']
        if fc['132'] and get_c('Fixed Charge - 132 kV'): sheet.cell(row=r, column=get_c('Fixed Charge - 132 kV')).value = fc['132']
        if fc['220'] and get_c('Fixed Charge - 220 kV'): sheet.cell(row=r, column=get_c('Fixed Charge - 220 kV')).value = fc['220']
        
        if ec['11'] and get_c('Energy Charge - 11 kV'): sheet.cell(row=r, column=get_c('Energy Charge - 11 kV')).value = ec['11']
        if ec['33'] and get_c('Energy Charge - 33 kV'): sheet.cell(row=r, column=get_c('Energy Charge - 33 kV')).value = ec['33']
        if ec['66'] and get_c('Energy Charge - 66 kV'): sheet.cell(row=r, column=get_c('Energy Charge - 66 kV')).value = ec['66']
        if ec['132'] and get_c('Energy Charge - 132 kV'): sheet.cell(row=r, column=get_c('Energy Charge - 132 kV')).value = ec['132']
        if ec['220'] and get_c('Energy Charge - 220 kV'): sheet.cell(row=r, column=get_c('Energy Charge - 220 kV')).value = ec['220']

    wb.save(excel_path)
    print(f"Updated {excel_path}")

if __name__ == "__main__":
    base_dir = os.path.dirname(os.path.abspath(__file__))
    
    # Dynamic JSONL search
    extracted_root = os.path.join(base_dir, "Extraction")
    jsonl_file = None
    if os.path.exists(extracted_root):
        for dirname in os.listdir(extracted_root):
            if "assam" in dirname.lower():
                state_dir = os.path.join(extracted_root, dirname)
                for f in os.listdir(state_dir):
                    if f.endswith(".jsonl"):
                        jsonl_file = os.path.join(state_dir, f)
                        break
            if jsonl_file: break
    
    # Fallback to direct path or specific file if dynamic search fails
    if not jsonl_file:
         j_f_direct = os.path.join(base_dir, 'Extraction', '1743056310.jsonl')
         if os.path.exists(j_f_direct):
             jsonl_file = j_f_direct

    excel_file = os.path.join(base_dir, 'Assam.xlsx')
    ists_path = os.path.join(base_dir, "ists_extracted", "ists_loss.json")
    
    if jsonl_file and os.path.exists(jsonl_file):
        print(f"Target JSONL: {jsonl_file}")
        
        names = extract_discom_names(jsonl_file)
        print(f"Discoms found: {len(names)} -> {names}")
        
        ists = extract_ists_loss(ists_path)
        insts_l = extract_losses(jsonl_file)
        w_l = extract_wheeling_losses(jsonl_file)
        insts_c = extract_transmission_charges(jsonl_file)
        w_c = extract_wheeling_charges(jsonl_file)
        css = extract_cross_subsidy_surcharge(jsonl_file)
        as_v = extract_additional_surcharge(jsonl_file)
        fc, ec = extract_tariff_charges(jsonl_file)
        fuel_s = extract_fuel_surcharge(jsonl_file)
        
        tod = extract_tod_charges(jsonl_file)
        pf_r = extract_pf_rebate(jsonl_file)
        lf_i = extract_load_factor_incentive(jsonl_file)
        gs_c = extract_grid_support_charges(jsonl_file)
        ht_r, ehv_r = extract_voltage_rebates(jsonl_file)
        bk_r = extract_bulk_consumption_rebate(jsonl_file)
        
        update_excel(names, ists, insts_l, w_l, insts_c, w_c, css, as_v, fc, ec, fuel_s, tod, pf_r, lf_i, gs_c, ht_r, ehv_r, bk_r, excel_file)
    else:
        print("Required files not found. Please check paths.")

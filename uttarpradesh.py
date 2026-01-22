import json
import os
from pathlib import Path
import pandas as pd
import re
from openpyxl import load_workbook
try:
    from database.database_utils import save_tariff_row
    DB_SUCCESS = True
except ImportError:
    DB_SUCCESS = False
from datetime import datetime

def find_value_in_jsonl(jsonl_path, table_keywords, row_keywords, value_constraint=lambda x: True):
    if not jsonl_path or not os.path.exists(jsonl_path): return "NA"
    with open(jsonl_path, 'r', encoding='utf-8') as f:
        for line in f:
            try:
                data = json.loads(line)
                h = data.get("table_heading", "").lower()
                heading_match = all(k.lower() in h for k in table_keywords)
                if heading_match:
                    for row in data.get("rows", []):
                        row_txt = str(row).lower()
                        if all(k.lower() in row_txt for k in row_keywords):
                            for v in list(row.values())[::-1]:
                                if not v: continue
                                try:
                                    clean = re.sub(r'[^\d\.]', '', str(v))
                                    if clean:
                                        f_v = float(clean)
                                        if value_constraint(f_v):
                                            return str(v).strip()
                                except: pass
            except: pass
    return "NA"

def extract_transmission_charges_from_dir(input_dir):
    if not input_dir: return "NA"
    for jsonl_path in input_dir.glob("**/*.jsonl"):
        val = find_value_in_jsonl(str(jsonl_path), ["transmission", "charge"], ["rs/kwh"], lambda x: 0.1 <= x <= 2.0)
        if val != "NA": return val
    return "NA"

def get_float_val(val_str):
    if not val_str: return None
    clean = str(val_str).replace(",", "").replace("%", "").strip()
    try: return float(clean)
    except: return None

def extract_ists_loss(json_path):
    try:
        if os.path.exists(json_path):
            with open(json_path, 'r', encoding='utf-8') as f:
                data = json.load(f)
                val = data.get("All India transmission Loss (in %)", None)
                if val:
                    return f"{val}%" if "%" not in str(val) else val
    except Exception as e:
        print(f"Error extracting ISTS loss: {e}")
    return "NA"

def extract_discom_names(input_dir):
    discom_names = set()
    table_keywords = ["discom", "distribution companies"] 
    ignore_keywords = [
        "particular", "s.no", "total", "year", "unit", "column", "approved", 
        "sr. no", "sr.no", "source", "rajasthan", "description", "remark",
        "proposed", "actual", "cost", "energy", "charge", "status", "report",
        "station", "plant", "capacity", "share", "state", "say", "tariff",
        "fixed", "variable", "commission", "month", "date", "category", "type",
        "consumer", "submission", "power", "name of", "domestic", "no", 
        "address", "discom", "name", "wheeling business", "supply business",
        "consolidated", "grand total", "phase", "particulars",
        "audited", "petition", "true-up", "uppcl", "claimed", "true up",
        "derivation", "s.n.", "load", "sales", "growth", "voltage", "level"
    ]
    
    if input_dir.exists():
        for jsonl_path in input_dir.glob("**/*.jsonl"):
            with open(jsonl_path, 'r', encoding='utf-8') as f:
                for line in f:
                    try:
                        data = json.loads(line)
                        heading = data.get("table_heading", data.get("heading", "")).lower()
                        
                        if any(k in heading for k in table_keywords):
                            # Strategy 1: Check Headers
                            for h in data.get("headers", []):
                                if h and isinstance(h, str):
                                    h_clean = h.strip()
                                    lower_h = h_clean.lower()
                                    if any(k in lower_h for k in ignore_keywords): continue
                                    if re.match(r'^s[r\.]*\s*n?[o\.]*$', lower_h): continue
                                    if re.match(r'^column[_\s]?\d*$', lower_h): continue
                                    if len(h_clean) < 25 and sum(c.isdigit() for c in h_clean) <= 2:
                                        discom_names.add(h_clean.upper())
                            
                            # Strategy 2: Check Content of "Discom" columns
                            headers = data.get("headers", [])
                            discom_keys = [h for h in headers if h and "discom" in str(h).lower()]
                            if discom_keys:
                                for row in data.get("rows", []):
                                    for k in discom_keys:
                                        val = row.get(k)
                                        if val and isinstance(val, str):
                                            val_clean = val.strip()
                                            lower_v = val_clean.lower()
                                            if len(val_clean) < 3: continue
                                            if any(k in lower_v for k in ignore_keywords): continue
                                            if any(c.isdigit() for c in val_clean) and len(val_clean) > 5: continue
                                            if len(val_clean) < 20: 
                                                discom_names.add(val_clean.upper())
                    except: pass
    return sorted(list(discom_names))

def extract_discoms():
    base_path = Path(__file__).resolve().parent
    
    # Dynamic Search for Extraction folder
    extraction_root = base_path / "Extraction"
    input_dir = None
    
    # Try multiple variations for folder name
    candidates = ["UttarPradesh", "Uttar Pradesh", "uttarpradesh", "uttar_pradesh"]
    if extraction_root.exists():
        # First check explicit candidates
        for c in candidates:
            p = extraction_root / c
            if p.exists() and p.is_dir():
                input_dir = p
                break
        
        # If not found, look for fuzzy match
        if not input_dir:
            for d in extraction_root.iterdir():
                if d.is_dir() and "uttar" in d.name.lower():
                    input_dir = d
                    break
    
    if not input_dir:
        input_dir = base_path / "table_extraction" # Fallback
        
    print(f"Using Input Directory: {input_dir}")

    # excel_path
    excel_path = base_path / "uttarpradesh.xlsx"
    if not excel_path.exists():
        # Try finding mostly likely match
        xls = [f for f in base_path.glob("*.xlsx") if "uttar" in f.name.lower() and "~$" not in f.name]
        if xls: excel_path = xls[0]
    
    # ISTS Path
    ists_loss_path = base_path / "ists_extracted" / "ists_loss.json"
    ists_loss_val = extract_ists_loss(str(ists_loss_path))

    # Storage
    insts_loss = {}
    retail_sales = {}
    insts_direct = {} # Direct Rate (Rs/kWh)
    insts_cr = {} # Rs. Cr from Cost tables
    
    charge_11kv, charge_33kv, charge_66kv, charge_132kv = {}, {}, {}, {}
    loss_11kv, loss_33kv, loss_66kv, loss_132kv = {}, {}, {}, {}
    a_loss_11kv, a_loss_33kv, a_loss_66kv, a_loss_132kv = {}, {}, {}, {}
    css_11kv, css_33kv, css_66kv, css_132kv, css_220kv = {}, {}, {}, {}, {}
    fixed_11kv, fixed_33kv, fixed_66kv, fixed_132kv, fixed_220kv = {}, {}, {}, {}, {}
    energy_11kv, energy_33kv, energy_66kv, energy_132kv, energy_220kv = {}, {}, {}, {}, {}
    grid_support = {}
    additional_surcharge = {}
    fuel_surcharge = {}
    pf_rebate = {}
    lf_incentive = {}
    ehv_rebate_33_66 = {}
    ehv_rebate_132_above = {}
    bulk_rebate = {}
    
    # FY strings
    now = datetime.now()
    fy_s = now.year if now.month >= 4 else now.year - 1
    f_curr = f"{fy_s}-{str(fy_s+1)[2:]}"
    f_curr_f = f"{fy_s}-{fy_s+1}"
    
    def gvv(d, m): return m.get(d) or m.get("DEFAULT")
    
    def find_val(row, d_norm):
        for k, v in row.items():
            if str(k).upper().strip() == d_norm:
                if v: return v
                keys = list(row.keys())
                idx = keys.index(k)
                if idx > 0:
                    prev = row.get(keys[idx-1])
                    if prev and any(c.isdigit() for c in str(prev)): return prev
        return None

    def get_val_numeric(row, kws=None):
        def is_clean(s):
            if "FY" in s.upper(): return False
            if sum(c.isalpha() for c in s) > 4: return False
            return True

        # Priority 1: Key match
        for k, v in row.items():
            if kws and any(x in str(k).lower() for x in kws):
                if v and any(c.isdigit() for c in str(v)):
                    s = str(v).strip()
                    if not re.fullmatch(r"T\d+", s) and is_clean(s): return s
        # Priority 2: Approved/Final
        for k, v in row.items():
            k_low = str(k).lower()
            if any(x in k_low for x in ["approved", "revised", "s**", "final"]):
                if v and any(c.isdigit() for c in str(v)):
                    s = str(v).strip()
                    if not re.fullmatch(r"T\d+", s) and is_clean(s): return s
        # Priority 3: Any number (prioritize % or .)
        best_v = None
        for k, v in row.items():
            if any(x in str(k).lower() for x in ["particular", "sn", "no", "category", "uom", "reference", "derivation"]): continue
            s = str(v).strip() if v else ""
            if any(c.isdigit() for c in s):
                if not is_clean(s): continue
                if re.fullmatch(r"20\d{2}(-\d{2})?", s): continue
                if re.fullmatch(r"T\d+", s): continue
                if "%" in s or "." in s: return s
                if not best_v: best_v = s
        return best_v

    # Load Discoms
    known = extract_discom_names(input_dir)
    print(f"Extracted Discoms: {known}")
    
    # Process
    if input_dir.exists():
        for jf in input_dir.glob("**/*.jsonl"):
            with open(jf, "r", encoding="utf-8") as f:
                for line in f:
                    try: table = json.loads(line)
                    except: continue
                    heading = table.get("heading", table.get("table_heading", "")).upper()
                    rows = table.get("rows", [])
                    full_text = heading + " " + " ".join(table.get("headers", [])) + " " + (str(rows[0]) if rows else "")
                    full_text = full_text.upper()

                    if f_curr not in full_text and f_curr_f not in full_text and any(x in full_text for x in ["2023", "2024", "2022"]):
                        continue # Skip wrong year

                    # 1. Energy Balance / Sales
                    if "ENERGY BALANCE" in full_text:
                        for row in rows:
                            txt = str(row).upper()
                            if "RETAIL SALES" in txt:
                                for d in known:
                                    v = find_val(row, d); 
                                    if v: retail_sales[d] = get_float_val(v)
                                # Fallback if names are in Column
                                if not retail_sales:
                                    for k, v in row.items():
                                        if any(d in k.upper() for d in known) and v:
                                            for d in known:
                                                if d in k.upper(): retail_sales[d] = get_float_val(v)
                            if "INTRA-STATE TRANS" in txt and "LOSS" in txt:
                                v = get_val_numeric(row)
                                if v:
                                    s_v = str(v).strip()
                                    if "%" not in s_v: s_v += "%"
                                    
                                    # Check if specific Discom row
                                    d_match = next((d for d in known if d in txt), None)
                                    if d_match: insts_loss[d_match] = s_v
                                    else: insts_loss['DEFAULT'] = s_v

                    # 2. Transmission Charges (InSTS / STU)
                    if any(k in full_text for k in ["INTRA-STATE TRANSMISSION SYSTEM CHARGES", "STU CHARGES", "STU TRANSMISSION CHARGES", "TRANSMISSION CHARGES", "INTRA-STATE TRANSMISSION CHARGE"]):
                        # Avoid Inter-State if generic "Transmission Charges" caused match, unless "Intra" or "STU" is explicit
                        if "INTER" in full_text and not any(x in full_text for x in ["INTRA", "STU", "UPPTCL"]): 
                            pass # Skip likely ISTS table
                        else:
                            is_rate = any(u in full_text for u in ["RS./KWH", "RS. / KWH", "INR/KWH", "PAISE/KWH"])
                            is_cost = "RS. CRORE" in full_text or "RS. CR" in full_text
                            
                            for row in rows:
                                txt = str(row).upper()
                                d_match = next((d for d in known if d in txt), None)
                                
                                if is_rate:
                                    v = get_val_numeric(row, [f_curr, "approved"])
                                    if v:
                                        if d_match: insts_direct[d_match] = v
                                        else: insts_direct['DEFAULT'] = v
                                
                                elif is_cost: # Fallback to calculation flow
                                    if d_match:
                                        v = get_val_numeric(row, ["approved", "net"])
                                        if v: insts_cr[d_match] = get_float_val(v)

                    # 3. Transmission Loss Table (Override if found deeper)
                    if "TRANSMISSION LOSSES" in full_text:
                        for row in rows:
                            txt = str(row).upper()
                            v = get_val_numeric(row)
                            if v:
                                if "INTRA" in txt: 
                                    s_v = str(v).strip()
                                    if "%" not in s_v: s_v += "%"
                                    d_match = next((d for d in known if d in txt), None)
                                    if d_match: insts_loss[d_match] = s_v
                                    else: insts_loss['DEFAULT'] = s_v

                    # 5. Distribution Loss / Wheeling Loss
                    if any(k in full_text for k in ["WHEELING LOSS", "DISCOM LOSS", "DISTRIBUTION LOSS", "VOLTAGE WISE LOSS", "DIFFERENT VOLTAGE LEVELS"]):
                        # SKIP PETITIONER TABLES FOR ACCURACY
                        if "petitioner" in heading: pass 
                        else:
                            for row in rows:
                                txt = str(row).upper()
                                
                                # Priority: Explicit "Loss Levels" column
                                v = None
                                for k, val in row.items():
                                    if "LOSS LEVELS" in str(k).upper():
                                        v = val
                                        break
                                
                                # Fallback: Numeric Search
                                if not v:
                                    v = get_val_numeric(row, [f_curr, "approved"]) # Pass dynamic year validation

                                if v:
                                    s_v = str(v).strip()
                                    if "%" not in s_v: s_v += "%"
                                    
                                    d_match = next((d for d in known if d in txt), None)
                                    
                                    if ("132" in txt or "66" in txt or "ABOVE 33" in txt or "EHT" in txt) and "KV" in txt:
                                        if d_match: 
                                            loss_132kv[d_match] = s_v; loss_66kv[d_match] = s_v
                                        else: 
                                            loss_132kv['DEFAULT'] = s_v; loss_66kv['DEFAULT'] = s_v
                                    elif "33" in txt and "KV" in txt:
                                        if d_match: loss_33kv[d_match] = s_v
                                        else: loss_33kv['DEFAULT'] = s_v
                                    elif "11" in txt and "KV" in txt and "BELOW" not in txt: # Exclude Below 11kV
                                        if d_match: loss_11kv[d_match] = s_v
                                        else: loss_11kv['DEFAULT'] = s_v

                    # 5b. Approved Loss Table (Accurate Values Priority)
                    if "approved" in heading and ("loss" in heading or "distribution" in heading):
                         if "rpo" in heading: pass # Skip RPO tables (Table 4-19) as they are not Wheeling Losses
                         else:
                             # Map Discoms to their value columns (Value is at Index - 1 or Index + 1? Dynamic Check)
                             # Strategy: Look for Discom Headers.
                             headers = data.get("headers", [])
                         d_map = {}
                         for d in known:
                             idx = next((i for i, h in enumerate(headers) if h and d in str(h).upper()), -1)
                             if idx >= 0:
                                 d_map[d] = headers[idx] 
                         
                         if d_map:
                             for row in rows:
                                 v_key = next((k for k in row.keys() if any(x in str(k).upper() for x in ["VOLTAGE", "LEVEL", "SYSTEM"])), None)
                                 if v_key:
                                     v_txt = str(row[v_key]).upper()
                                     target = None
                                     if "11" in v_txt and "KV" in v_txt: target = a_loss_11kv
                                     elif "33" in v_txt and "KV" in v_txt: target = a_loss_33kv
                                     elif ("66" in v_txt or "ABOVE 33" in v_txt) and "KV" in v_txt: target = a_loss_66kv
                                     elif "132" in v_txt and "KV" in v_txt: target = a_loss_132kv
                                     elif "220" in v_txt and "KV" in v_txt: target = a_loss_132kv
                                     
                                     if target is not None:
                                         for d, col_k in d_map.items():
                                             val = row.get(col_k)
                                             if val:
                                                 s_v = str(val).strip() 
                                                 # Filter out non-numeric
                                                 if not any(c.isdigit() for c in s_v): continue
                                                 if "%" not in s_v: s_v += "%"
                                                 target[d] = s_v

                    # 6. Wheeling Charges
                    if any(k in full_text for k in ["WHEELING CHARGE", "WHEELING CHARGES", "DISCOM CHARGES", "DISTRIBUTION CHARGES", "VOLTAGE WISE CHARGES"]):
                         is_rate = any(u in full_text for u in ["RS./KWH", "RS. / KWH", "INR/KWH", "PAISE/KWH"])
                         if is_rate:
                            for row in rows:
                                txt = str(row).upper()
                                # Dynamic Year Check
                                v = get_val_numeric(row, [f_curr, "approved"])
                                
                                if v:
                                    d_match = next((d for d in known if d in txt), None)
                                    
                                    if "132" in txt or "66" in txt or "ABOVE 33" in txt or "EHT" in txt:
                                        if d_match: charge_132kv[d_match] = v; charge_66kv[d_match] = v
                                        else: charge_132kv['DEFAULT'] = v; charge_66kv['DEFAULT'] = v
                                    elif "33" in txt:
                                        if d_match: charge_33kv[d_match] = v
                                        else: charge_33kv['DEFAULT'] = v
                                    elif "11" in txt:
                                        if d_match: charge_11kv[d_match] = v
                                        else: charge_11kv['DEFAULT'] = v

                    # 7. CSS - Only from TABLE 10-14
                    if "TABLE 10-14" in heading and any(k in full_text for k in ["CSS", "CROSS SUBSIDY"]):
                         is_rate = any(u in full_text for u in ["RS./KWH", "RS. /KWH", "RS /KWH", "INR/KWH", "PAISE/KWH", "/KWH"])
                         if is_rate:
                            for row in rows:
                                txt = str(row).upper()
                                # Find the Approved column specifically
                                v = None
                                for k, val in row.items():
                                    # Check if this is the Approved column
                                    if k == 'Approved' or 'APPROVED' in str(k).upper() or 'D (LOWER' in str(k).upper():
                                        if val and val != '-' and val != 'None':
                                            v = str(val).strip()
                                            break
                                
                                if v:
                                    d_match = next((d for d in known if d in txt), None)
                                    hv2 = "HV-2" in txt or "INDUSTRIAL" in txt
                                    
                                    # Get category field for more precise matching
                                    category = str(row.get('Category', row.get('Categories', ''))).upper()
                                    
                                    d_key = d_match if d_match else 'DEFAULT'
                                    if hv2:
                                        if ("SUPPLY AT 11" in category or "AT 11 KV" in category) and "ABOVE" not in category: 
                                            css_11kv[d_key] = v
                                        elif "ABOVE 11" in category and ("66" in category or "UP TO 66" in category):
                                            css_33kv[d_key] = v
                                            css_66kv[d_key] = v
                                        elif "ABOVE 66" in category and ("132" in category or "UP TO 132" in category):
                                            css_132kv[d_key] = v
                                        elif ("ABOVE 132 KV" in category or ("ABOVE 132" in category and "KV" in category)) and "66" not in category:
                                            css_220kv[d_key] = v
                                    else:
                                        # Standard fallback: only if not already set by HV-2
                                        target = None
                                        if "220" in txt and "KV" in txt: target = css_220kv
                                        elif "132" in txt and "KV" in txt: target = css_132kv
                                        elif ("66" in txt or "ABOVE 33" in txt) and "KV" in txt: target = css_66kv
                                        elif "33" in txt and "KV" in txt: target = css_33kv
                                        elif "11" in txt and "KV" in txt and "BELOW" not in txt: target = css_11kv

                                        if target is not None and d_key not in target:
                                            target[d_key] = v

                    # 8. Additional Surcharge
                    if any(k in full_text for k in ["AS CHARGES", "ADDITIONAL SURCHARGE"]):
                         is_rate = any(u in full_text for u in ["RS./KWH", "RS. /KWH", "RS /KWH", "INR/KWH", "PAISE/KWH", "/KWH"])
                         if is_rate:
                            for row in rows:
                                txt = str(row).upper()
                                # Find the value with year validation
                                v = None
                                for k, val in row.items():
                                    k_upper = str(k).upper()
                                    if any(x in k_upper for x in ["APPROVED", "SURCHARGE", "AS"]) or k == 'Approved':
                                        if val and val != '-' and val != 'None':
                                            v = str(val).strip()
                                            break
                                
                                if v:
                                    d_match = next((d for d in known if d in txt), None)
                                    d_key = d_match if d_match else 'DEFAULT'
                                    additional_surcharge[d_key] = v

                    # 9. Fixed Charges
                    if any(k in full_text for k in ["FIXED CHARGES", "FIXED CHARGE", "DEMAND CHARGES", "DEMAND CHARGE", "RATE SCHEDULE", "TARIFF SCHEDULE", "URBAN SCHEDULE"]):
                         for row in rows:
                             txt = str(row).upper()
                             # Only process if row mentions FIXED or DEMAND
                             if "FIXED" in txt or "DEMAND" in txt:
                                 # 1. Specialized Urban Schedule extraction
                                 if "URBAN SCHEDULE" in heading:
                                     # Standard columns for this table: Column_1=11, Column_2=33/66, Column_3=132, Column_4=220
                                     for c_key, target_dicts in [
                                         ('Column_1', [fixed_11kv]),
                                         ('Column_2', [fixed_33kv, fixed_66kv]),
                                         ('Column_3', [fixed_132kv]),
                                         ('Column_4', [fixed_220kv])
                                     ]:
                                         val = row.get(c_key)
                                         if val and 'Rs.' in str(val):
                                             # Clean value: "Rs. 300.00 / kVA / month" -> "300.00"
                                             clean_val = str(val).split('/')[0].replace('Rs.', '').replace('\n', '').strip()
                                             for d in target_dicts: d['DEFAULT'] = clean_val
                                     continue # Skip generic extraction for this row

                                 # 2. Generic extraction
                                 v = None
                                 for k, val in row.items():
                                     k_upper = str(k).upper()
                                     if any(x in k_upper for x in ["APPROVED", "FIXED", "DEMAND", "CHARGE", f_curr.upper()]) or k == 'Approved':
                                         if val and val != '-' and val != 'None':
                                             v = str(val).strip()
                                             break
                                 
                                 # Fallback to get_val_numeric if no specific column found
                                 if not v:
                                     v = get_val_numeric(row, ["fixed", "demand", "approved"])
                                 
                                 if v:
                                     d_match = next((d for d in known if d in txt), None)
                                     d_key = d_match if d_match else 'DEFAULT'
                                     
                                     # Match voltage levels
                                     if "220 KV" in txt or "220KV" in txt:
                                         fixed_220kv[d_key] = v
                                     elif "132 KV" in txt or "132KV" in txt:
                                         fixed_132kv[d_key] = v
                                     elif "66 KV" in txt or "66KV" in txt:
                                         fixed_66kv[d_key] = v
                                     elif "33 KV" in txt or "33KV" in txt:
                                         fixed_33kv[d_key] = v
                                     elif "11 KV" in txt or "11KV" in txt:
                                         fixed_11kv[d_key] = v

                    # 10. Energy Charges
                    if any(k in full_text for k in ["ENERGY CHARGES", "ENERGY CHARGE", "VARIABLE CHARGES", "VARIABLE CHARGE", "URBAN SCHEDULE"]):
                         for row in rows:
                             txt = str(row).upper()
                             # Only process if row mentions ENERGY or VARIABLE
                             if "ENERGY" in txt or "VARIABLE" in txt:
                                 # 1. Specialized Urban Schedule extraction
                                 if "URBAN SCHEDULE" in heading:
                                     # Standard columns for this table: Column_1=11, Column_2=33/66, Column_3=132, Column_4=220
                                     for c_key, target_dicts in [
                                         ('Column_1', [energy_11kv]),
                                         ('Column_2', [energy_33kv, energy_66kv]),
                                         ('Column_3', [energy_132kv]),
                                         ('Column_4', [energy_220kv])
                                     ]:
                                         val = row.get(c_key)
                                         if val and 'Rs.' in str(val):
                                             # Clean value: "Rs. 7.10 / kVAh" -> "7.10"
                                             clean_val = str(val).split('/')[0].replace('Rs.', '').replace('\n', '').strip()
                                             for d in target_dicts: d['DEFAULT'] = clean_val
                                     continue # Skip generic extraction for this row

                                 # 2. Generic extraction
                                 v = None
                                 for k, val in row.items():
                                     k_upper = str(k).upper()
                                     if any(x in k_upper for x in ["APPROVED", "ENERGY", "VARIABLE", "CHARGE", f_curr.upper()]) or k == 'Approved':
                                         if val and val != '-' and val != 'None':
                                             v = str(val).strip()
                                             break
                                 
                                 # Fallback to get_val_numeric
                                 if not v:
                                     v = get_val_numeric(row, ["energy", "variable", "approved"])
                                 
                                 if v:
                                     d_match = next((d for d in known if d in txt), None)
                                     d_key = d_match if d_match else 'DEFAULT'
                                     
                                     # Match voltage levels (prioritize HV-2/Industrial)
                                     is_hv_industrial = "HV-2" in txt or "INDUSTRIAL" in txt
                                     if "220 KV" in txt or "220KV" in txt:
                                         if is_hv_industrial or d_key not in energy_220kv:
                                             energy_220kv[d_key] = v
                                     elif "132 KV" in txt or "132KV" in txt:
                                         if is_hv_industrial or d_key not in energy_132kv:
                                             energy_132kv[d_key] = v
                                     elif "66 KV" in txt or "66KV" in txt:
                                         if is_hv_industrial or d_key not in energy_66kv:
                                             energy_66kv[d_key] = v
                                     elif "33 KV" in txt or "33KV" in txt:
                                         if is_hv_industrial or d_key not in energy_33kv:
                                             energy_33kv[d_key] = v
                                     elif "11 KV" in txt or "11KV" in txt:
                                         if is_hv_industrial or d_key not in energy_11kv:
                                             energy_11kv[d_key] = v

                    # 11. Fuel Surcharge
                    fs_kws = ["FUEL ADJUSTMENT COST", "FUEL", "FPPPA", "FUEL SURCHARGE", "FPPCA", "ECA", "FPPAS"]
                    if any(k in full_text for k in fs_kws):
                         is_rate = any(u in full_text for u in ["RS./KWH", "RS. /KWH", "RS /KWH", "INR/KWH", "PAISE/KWH", "/KWH"])
                         if is_rate:
                            for row in rows:
                                txt = str(row).upper()
                                # Extract value with year validation
                                v = None
                                for k, val in row.items():
                                    k_upper = str(k).upper()
                                    # Check for specific keywords in column names
                                    if any(x in k_upper for x in fs_kws + ["APPROVED", f_curr.upper()]) or k == 'Approved':
                                        if val and val != '-' and val != 'None':
                                            v = str(val).strip()
                                            break
                                
                                if v:
                                    d_match = next((d for d in known if d in txt), None)
                                    d_key = d_match if d_match else 'DEFAULT'
                                    fuel_surcharge[d_key] = v

                    # 12. Power Factor Adjustment Rebate
                    pf_kws = ["POWER FACTOR ADJUSTMENT REBATE", "POWER FACTOR ADJUSTMENT DISCOUNT", "POWER FACTOR ADJUSTMENT"]
                    if any(k in full_text for k in pf_kws):
                         is_rate = any(u in full_text for u in ["RS./KWH", "RS. /KWH", "RS /KWH", "INR/KWH", "PAISE/KWH", "/KWH"])
                         if is_rate:
                            for row in rows:
                                txt = str(row).upper()
                                v = None
                                for k, val in row.items():
                                    k_upper = str(k).upper()
                                    if any(x in k_upper for x in pf_kws + ["APPROVED", f_curr.upper()]) or k == 'Approved':
                                        if val and val != '-' and val != 'None':
                                            v = str(val).strip()
                                            break
                                
                                if v:
                                    d_match = next((d for d in known if d in txt), None)
                                    d_key = d_match if d_match else 'DEFAULT'
                                    pf_rebate[d_key] = v

                    # 13. Load Factor Incentive
                    lf_kws = ["LOAD FACTOR INCENTIVE", "LOAD FACTOR DISCOUNT", "LOAD FACTOR"]
                    if any(k in full_text for k in lf_kws):
                         is_rate = any(u in full_text for u in ["RS./KWH", "RS. /KWH", "RS /KWH", "INR/KWH", "PAISE/KWH", "/KWH"])
                         if is_rate:
                            for row in rows:
                                txt = str(row).upper()
                                v = None
                                for k, val in row.items():
                                    k_upper = str(k).upper()
                                    if any(x in k_upper for x in lf_kws + ["APPROVED", f_curr.upper()]) or k == 'Approved':
                                        if val and val != '-' and val != 'None':
                                            v = str(val).strip()
                                            break
                                
                                if v:
                                    d_match = next((d for d in known if d in txt), None)
                                    d_key = d_match if d_match else 'DEFAULT'
                                    lf_incentive[d_key] = v

                    # 14. Grid Support / Parallel Operation
                    gs_kws = ["GRID SUPPORT", "PARALLEL OPERATION"]
                    if any(k in full_text for k in gs_kws):
                         is_rate = any(u in full_text for u in ["RS./KWH", "RS. /KWH", "RS /KWH", "INR/KWH", "PAISE/KWH", "/KWH"])
                         if is_rate:
                            for row in rows:
                                txt = str(row).upper()
                                v = None
                                for k, val in row.items():
                                    k_upper = str(k).upper()
                                    if any(x in k_upper for x in gs_kws + ["APPROVED", f_curr.upper()]) or k == 'Approved':
                                        if val and val != '-' and val != 'None':
                                            v = str(val).strip()
                                            break
                                
                                if v:
                                    d_match = next((d for d in known if d in txt), None)
                                    d_key = d_match if d_match else 'DEFAULT'
                                    grid_support[d_key] = v

                    # 15. HT, EHV Rebates
                    ehv_kws = ["HT REBATE", "EHV REBATE", "HT DISCOUNT", "EHV DISCOUNT"]
                    if any(k in full_text for k in ehv_kws):
                         is_rate = any(u in full_text for u in ["RS./KWH", "RS. /KWH", "RS /KWH", "INR/KWH", "PAISE/KWH", "/KWH"])
                         if is_rate:
                            for row in rows:
                                txt = str(row).upper()
                                v = None
                                for k, val in row.items():
                                    k_upper = str(k).upper()
                                    if any(x in k_upper for x in ehv_kws + ["APPROVED", f_curr.upper()]) or k == 'Approved':
                                        if val and val != '-' and val != 'None':
                                            v = str(val).strip()
                                            break
                                
                                if v:
                                    d_match = next((d for d in known if d in txt), None)
                                    d_key = d_match if d_match else 'DEFAULT'
                                    if any(x in txt for x in ["132", "220"]):
                                        ehv_rebate_132_above[d_key] = v
                                    elif any(x in txt for x in ["33", "66"]):
                                        ehv_rebate_33_66[d_key] = v

                    # 16. Bulk Consumption Rebate
                    bulk_kws = ["BULK CONSUMPTION REBATE", "BULK CONSUMPTION DISCOUNT", "BULK CONSUMPTION"]
                    if any(k in full_text for k in bulk_kws):
                         is_rate = any(u in full_text for u in ["RS./KWH", "RS. /KWH", "RS /KWH", "INR/KWH", "PAISE/KWH", "/KWH"])
                         if is_rate:
                            for row in rows:
                                txt = str(row).upper()
                                v = None
                                for k, val in row.items():
                                    k_upper = str(k).upper()
                                    if any(x in k_upper for x in bulk_kws + ["APPROVED", f_curr.upper()]) or k == 'Approved':
                                        if val and val != '-' and val != 'None':
                                            v = str(val).strip()
                                            break
                                
                                if v:
                                    d_match = next((d for d in known if d in txt), None)
                                    d_key = d_match if d_match else 'DEFAULT'
                                    bulk_rebate[d_key] = v

    insts_final, insts_final = {}, {}
    
    # Merge Approved losses (Priority)
    loss_11kv.update(a_loss_11kv)
    loss_33kv.update(a_loss_33kv)
    loss_66kv.update(a_loss_66kv)
    loss_132kv.update(a_loss_132kv)

    # InSTS Processing
    for d in known:
        rate_d = gvv(d, insts_direct)
        if rate_d:
             insts_final[d] = str(rate_d)
        else:
            cr = insts_cr.get(d)
            sl = retail_sales.get(d)
            if cr and sl: 
                rate = (cr * 10) / sl
                insts_final[d] = f"{rate:.4f}"
            else:
                insts_final[d] = "NA"

    # Excel Update
    try:
        wb = load_workbook(excel_path)
        ws = wb.active
        # Find headers again to be sure
        hr = 1
        for r in range(1, 10):
            if any("DISCOM" in str(c.value).upper() for c in ws[r] if c.value): hr = r; break
        
        cmap = {}
        for c in range(1, ws.max_column + 1):
            val = ws.cell(hr, c).value
            if val:
                h = str(val).upper().strip()
                if h not in cmap: cmap[h] = c
        
        def gc(kws):
            for h, i in cmap.items():
                if all(k.upper() in h for k in kws): return i
            return None

        ts = {
            "il": gc(["ISTS", "LOSS"]), "nl": gc(["INSTS", "LOSS"]),
            "nc": gc(["INSTS", "CHARGE"]),
            "wc": [gc(["WHEELING", "CHARGE", "11"]), gc(["WHEELING", "CHARGE", "33"]), gc(["WHEELING", "CHARGE", "66"]), gc(["WHEELING", "CHARGE", "132"])],
            "wl": [gc(["WHEELING", "LOSS", "11"]), gc(["WHEELING", "LOSS", "33"]), gc(["WHEELING", "LOSS", "66"]), gc(["WHEELING", "LOSS", "132"])],
            "css": [gc(["CROSS", "SUBSIDY", "11"]), gc(["CROSS", "SUBSIDY", "33"]), gc(["CROSS", "SUBSIDY", "66"]), gc(["CROSS", "SUBSIDY", "132"]), gc(["CROSS", "SUBSIDY", "220"])],
            "as": gc(["ADDITIONAL", "SURCHARGE"]),
            "f": [gc(["FIXED", "11"]), gc(["FIXED", "33"]), gc(["FIXED", "66"]), gc(["FIXED", "132"]), gc(["FIXED", "220"])],
            "e": [gc(["ENERGY", "11"]), gc(["ENERGY", "33"]), gc(["ENERGY", "66"]), gc(["ENERGY", "132"]), gc(["ENERGY", "220"])],
            "fs": gc(["FUEL", "SURCHARGE"]),
            "pf": gc(["POWER", "FACTOR", "REBATE"]),
            "lf": gc(["LOAD", "FACTOR", "INCENTIVE"]),
            "g": gc(["GRID", "SUPPORT"]) or gc(["PARALLEL", "OPERATION"]),
            "er33": gc(["HT", "EHV", "REBATE", "33"]),
            "er132": gc(["HT", "EHV", "REBATE", "132"]),
            "br": gc(["BULK", "CONSUMPTION", "REBATE"]),
            "d": cmap.get("DISCOM"), "s": cmap.get("STATE") or cmap.get("STATES")
        }

        if ws.max_row >= 3: ws.delete_rows(3, ws.max_row - 2)

        def wv(r, c, v):
            # Always update if column exists, even if value is None
            if c: 
                ws.cell(row=r, column=c).value = str(v) if v else "NA"

        cr = 3
        # Final Global Extractions
        insts_charges_val = extract_transmission_charges_from_dir(input_dir)
        
        # Global Patterns for specific fields
        patterns = {
            'pf_rebate': (['power factor'], ['rebate', 'incentive']),
            'lf_incentive': (['load factor'], ['incentive', 'rebate']),
            'grid_support': (['grid support', 'parallel operation'], ['charge']),
            'bulk_rebate': (['bulk', 'consumption'], ['rebate']),
        }
        
        extra_field_vals = {}
        for field, (tk, rk) in patterns.items():
            found = "NA"
            for jsonl in input_dir.glob("**/*.jsonl"):
                found = find_value_in_jsonl(str(jsonl), tk, rk)
                if found != "NA": break
            extra_field_vals[field] = found

        for d in known:
            wv(cr, ts["d"], d); wv(cr, ts["s"], "UTTAR PRADESH")
            wv(cr, ts["il"], ists_loss_val) 
            wv(cr, ts["nl"], gvv(d, insts_loss))
            wv(cr, ts["nc"], insts_final.get(d))
            for i, col in enumerate(ts["wc"]): wv(cr, col, gvv(d, [charge_11kv, charge_33kv, charge_66kv, charge_132kv][i]))
            for i, col in enumerate(ts["wl"]): wv(cr, col, gvv(d, [loss_11kv, loss_33kv, loss_66kv, loss_132kv][i]))
            for i, col in enumerate(ts["css"]): wv(cr, col, gvv(d, [css_11kv, css_33kv, css_66kv, css_132kv, css_220kv][i]))
            wv(cr, ts["as"], gvv(d, additional_surcharge))
            wv(cr, ts["fs"], gvv(d, fuel_surcharge))
            wv(cr, ts["pf"], gvv(d, pf_rebate))
            wv(cr, ts["lf"], gvv(d, lf_incentive))
            wv(cr, ts["g"], gvv(d, grid_support))
            wv(cr, ts["er33"], gvv(d, ehv_rebate_33_66))
            wv(cr, ts["er132"], gvv(d, ehv_rebate_132_above))
            wv(cr, ts["br"], gvv(d, bulk_rebate))
            for i, col in enumerate(ts["f"]): wv(cr, col, gvv(d, [fixed_11kv, fixed_33kv, fixed_66kv, fixed_132kv, fixed_220kv][i]))
            for i, col in enumerate(ts["e"]): wv(cr, col, gvv(d, [energy_11kv, energy_33kv, energy_66kv, energy_132kv, energy_220kv][i]))
            
            if DB_SUCCESS:
                db_data = {
                    'financial_year': "FY2025-26",
                    'state': 'Uttar Pradesh',
                    'discom': d,
                    'ists_loss': str(ists_loss_val) if ists_loss_val else "NA",
                    'insts_loss': str(gvv(d, insts_loss)) if gvv(d, insts_loss) else "NA",
                    'wheeling_loss_11kv': gvv(d, loss_11kv),
                    'wheeling_loss_33kv': gvv(d, loss_33kv),
                    'wheeling_loss_66kv': gvv(d, loss_66kv),
                    'wheeling_loss_132kv': gvv(d, loss_132kv),
                    'ists_charges': "NA",
                    'insts_charges': str(insts_charges_val) if insts_charges_val else "NA",
                    'wheeling_charges_11kv': gvv(d, charge_11kv),
                    'wheeling_charges_33kv': gvv(d, charge_33kv),
                    'wheeling_charges_66kv': gvv(d, charge_66kv),
                    'wheeling_charges_132kv': gvv(d, charge_132kv),
                    'css_charges_11kv': gvv(d, css_11kv),
                    'css_charges_33kv': gvv(d, css_33kv),
                    'css_charges_66kv': gvv(d, css_66kv),
                    'css_charges_132kv': gvv(d, css_132kv),
                    'css_charges_220kv': gvv(d, css_220kv),
                    'additional_surcharge': gvv(d, additional_surcharge),
                    'electricity_duty': "NA",
                    'tax_on_sale': "NA",
                    'fixed_charge_11kv': gvv(d, fixed_11kv),
                    'fixed_charge_33kv': gvv(d, fixed_33kv),
                    'fixed_charge_66kv': gvv(d, fixed_66kv),
                    'fixed_charge_132kv': gvv(d, fixed_132kv),
                    'fixed_charge_220kv': gvv(d, fixed_220kv),
                    'energy_charge_11kv': gvv(d, energy_11kv),
                    'energy_charge_33kv': gvv(d, energy_33kv),
                    'energy_charge_66kv': gvv(d, energy_66kv),
                    'energy_charge_132kv': gvv(d, energy_132kv),
                    'energy_charge_220kv': gvv(d, energy_220kv),
                    'fuel_surcharge': gvv(d, fuel_surcharge),
                    'tod_charges': "NA",
                    'pf_rebate': gvv(d, pf_rebate),
                    'lf_incentive': gvv(d, lf_incentive),
                    'grid_support_parallel_op_charges': gvv(d, grid_support),
                    'ht_ehv_rebate_33_66kv': gvv(d, ehv_rebate_33_66),
                    'ht_ehv_rebate_132_above': gvv(d, ehv_rebate_132_above),
                    'bulk_rebate': gvv(d, bulk_rebate)
                }
                # Sanitize data: convert all values to string and replace None with NA
                clean_db_data = {k: (str(v) if v is not None else "NA") for k, v in db_data.items()}
                save_tariff_row(clean_db_data)
            cr += 1
        wb.save(excel_path)
        print(f"Update Success: {excel_path}")
    except Exception as e: print(f"Update Error: {e}")

if __name__ == "__main__": extract_discoms()

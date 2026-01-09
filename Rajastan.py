import json
import re
import os
import openpyxl

def extract_discom_names(jsonl_path):
    discom_names = set()
    table_keywords = ["discom", "distribution companies"] 
    ignore_keywords = [
        "particular", "s.no", "total", "year", "unit", "column", "approved", 
        "sr. no", "sr.no", "source", "rajasthan", "description", "remark",
        "proposed", "actual", "cost", "energy", "charge", "status", "report",
        "station", "plant", "capacity", "share", "state", "say", "tariff",
        "fixed", "variable", "commission", "month", "date", "category", "type",
        "consumer", "submission", "fy", "scheme", "power", "name of", "domestic", "no"
    ]
    
    with open(jsonl_path, 'r', encoding='utf-8') as f:
        for line in f:
            try:
                data = json.loads(line)
                heading = data.get("table_heading", "").lower()
                if any(k in heading for k in table_keywords):
                    for h in data.get("headers", []):
                        if not h or not isinstance(h, str): continue
                        h_clean = h.strip()
                        lower_h = h_clean.lower()
                        
                        # skip based on keywords
                        if any(k in lower_h for k in ignore_keywords): continue
                        
                        # skip based on regex
                        if re.match(r'^s[r\.]*\s*n?[o\.]*$', lower_h): continue
                        if re.match(r'^column[_\s]?\d*$', lower_h): continue
                        
                        # skip if too long or looks like number
                        if len(h_clean) > 10: continue
                        if sum(c.isdigit() for c in h_clean) > 2: continue
                        
                        discom_names.add(h_clean)
            except: pass
    
    sorted_names = sorted(list(discom_names))
    return sorted_names

def extract_losses(jsonl_path):
    insts_loss = None
    
    with open(jsonl_path, 'r', encoding='utf-8') as f:
        for line in f:
            if insts_loss: break
            try:
                data = json.loads(line)
                rows = data.get("rows", [])
                for row in rows:
                    if not row: continue
                    r_txt = str(row).lower()
                    
                    # Logic to find best value in row
                    def get_best_val(r):
                        candidates = []
                        for k, v in r.items():
                            if not v or not isinstance(v, str): continue
                            k_lower = k.lower()
                            # Clean value
                            v_clean = v.strip().replace('%', '')
                            try:
                                float(v_clean)
                                # Score candidate
                                score = 0
                                if "approved" in k_lower: score += 2
                                if re.search(r'fy\s*20\d{2}[-20]*\d\d', k_lower): score += 2
                                if "proposed" in k_lower: score += 1
                                candidates.append((v_clean, score))
                            except: pass
                        
                        candidates.sort(key=lambda x: x[1], reverse=True)
                        if candidates:
                             return f"{candidates[0][0]}%"
                        return None


                    if "intra-state transmission losses" in r_txt:
                        val = get_best_val(row)
                        if val: insts_loss = val
            except: pass
            
    # Fallbacks or cleanup
    print(f"Extracted InSTS Loss: {insts_loss}")
    return insts_loss

def extract_wheeling_losses(jsonl_path, target_discoms):
    # results: {discom_name: {voltage_level: {val: value, priority: p}}}
    results = {}
    defaults = {} # {voltage_level: {val: value, priority: p}}
    
    with open(jsonl_path, 'r', encoding='utf-8') as f:
        for line in f:
            try:
                data = json.loads(line)
                heading = data.get("table_heading", "").lower()
                rows = data.get("rows", [])
                
                # Check for relevant table types
                priority = 0
                if "apportionment of voltage-wise sales" in heading:
                    priority = 10
                elif "distribution losses" in heading or "wheeling losses" in heading:
                    priority = 1
                
                if priority == 0: continue

                for row in rows:
                    # 1. Identify Discom in Row OR use Heading
                    matched_discom = None
                    row_vals_str = [str(v).lower() for v in row.values() if v]
                    row_txt = " ".join(row_vals_str)
                    
                    for d in target_discoms:
                        if d.lower() in row_txt:
                            matched_discom = d
                            break
                    
                    # If not in row, checking heading
                    if not matched_discom:
                        for d in target_discoms:
                            if d.lower() in heading:
                                matched_discom = d
                                break
                    
                    # Extract values
                    extracted = {}
                    
                    # Strategy A: Check if voltage is in a specific column (Row sets voltage)
                    row_voltage = None
                    # Normalize keys
                    row_clean = {k.lower().strip(): v for k, v in row.items() if v}
                    
                    # Look for voltage level in "voltage level" column or similar
                    voltage_keys = [k for k in row_clean.keys() if "voltage" in k and "level" in k]
                    v_val_str = ""
                    if voltage_keys:
                        v_val_str = str(row_clean[voltage_keys[0]]).lower()
                    
                    # Extract voltage number from voltage column or just check if the row represents a voltage
                    v_match = re.search(r'(\d+)\s*kv', v_val_str)
                    if v_match:
                        row_voltage = v_match.group(1)
                    
                    if row_voltage:
                        # If we found a voltage row, we look for "losses" column
                        loss_val = None
                        for k, v in row_clean.items():
                            if "loss" in k and ("%" in k or "percent" in k or "%" in str(v)):
                                loss_val = v
                                break
                        
                        if loss_val:
                            loss_val_clean = str(loss_val).strip() # Keep % if present
                            extracted[row_voltage] = loss_val_clean
                            
                    else:
                        # Strategy B: Column Header is Voltage (Old Logic)
                        for k, v in row.items():
                            k_low = k.lower().replace(" ", "").replace("\n", "")
                            # Standard keys
                            if "11kV" in k_low or "11kv" in k_low: 
                                if "%" in str(v): extracted["11"] = v
                            if "33kV" in k_low or "33kv" in k_low: 
                                if "%" in str(v): extracted["33"] = v
                            if "132kV" in k_low or "132kv" in k_low: 
                                 if "%" in str(v): extracted["132"] = v
                            if "66kV" in k_low or "66kv" in k_low: 
                                 if "%" in str(v): extracted["66"] = v
                    
                    if extracted:
                        if matched_discom:
                            if matched_discom not in results: results[matched_discom] = {}
                            for kv, val in extracted.items():
                                existing = results[matched_discom].get(kv)
                                if not existing or priority > existing['priority']:
                                    results[matched_discom][kv] = {'val': val, 'priority': priority}
                        else:
                            # Generic / Default values
                            for kv, val in extracted.items():
                                existing = defaults.get(kv)
                                if not existing or priority > existing['priority']:
                                    defaults[kv] = {'val': val, 'priority': priority}

            except: pass
    
    # Finalize results - convert back to simple {discom: {kv: val}}
    final_results = {}
    
    # 1. Fill from extracted results
    for d, data in results.items():
        if d not in final_results: final_results[d] = {}
        for kv, item in data.items():
            final_results[d][kv] = item['val']
            
    # 2. Fill missing discoms/voltages from defaults
    simple_defaults = {k: v['val'] for k, v in defaults.items()}
    
    for d in target_discoms:
        if d not in final_results:
             final_results[d] = simple_defaults.copy()
        else:
             # Fill missing keys in existing discom
             for k, v in simple_defaults.items():
                 if k not in final_results[d]:
                     final_results[d][k] = v
                
    print(f"Extracted Dynamic Wheeling Losses: {final_results}")
    return final_results

def extract_wheeling_charges(jsonl_path, target_discoms):
    # results: {discom_name: {voltage_level: wheeling_charge_value}}
    w_charges = {}
    # results: {discom_name: {voltage_level: transmission_charge_value}}
    t_charges = {}
    
    with open(jsonl_path, 'r', encoding='utf-8') as f:
        for line in f:
            try:
                data = json.loads(line)
                heading = data.get("table_heading", "").lower()
                
                if ("wheeling" in heading and "transmission" in heading and "cost" in heading):
                    rows = data.get("rows", [])
                    current_matched_discom = None
                    
                    for row in rows:
                        # 1. Identify/Update Discom
                        raw_discom = row.get("Discom")
                        if raw_discom:
                            raw_discom_s = str(raw_discom).strip()
                            current_matched_discom = None
                            for d in target_discoms:
                                if d.lower() in raw_discom_s.lower():
                                    current_matched_discom = d
                                    break
                        
                        if not current_matched_discom:
                            continue
                        
                        # 2. Check row type
                        row_vals_txt = " ".join([str(v).lower() for v in row.values() if v])
                        
                        # Wheeling cost
                        if "wheeling cost" in row_vals_txt and "transmission" not in row_vals_txt:
                            if current_matched_discom not in w_charges: w_charges[current_matched_discom] = {}
                            for kv in ["11kV", "33kV", "132kV"]:
                                val = row.get(kv)
                                if val:
                                    try:
                                        kv_key = kv.replace("kV", "")
                                        w_charges[current_matched_discom][kv_key] = float(val)
                                    except: pass
                        
                        # Transmission cost (InSTS Charge)
                        if "transmission cost" in row_vals_txt:
                            if current_matched_discom not in t_charges: t_charges[current_matched_discom] = {}
                            for kv in ["11kV", "33kV", "132kV"]:
                                val = row.get(kv)
                                if val:
                                    try:
                                        kv_key = kv.replace("kV", "")
                                        t_charges[current_matched_discom][kv_key] = float(val)
                                    except: pass
            except: pass
                
    print(f"Extracted Dynamic Wheeling Charges: {w_charges}")
    print(f"Extracted Dynamic Transmission (InSTS) Charges: {t_charges}")
    return w_charges, t_charges

def extract_css_charges(jsonl_path):
    # Dictionary to store CSS: {voltage_level: css_value}
    # Values seem to be same for all Discoms in Table 95
    css_values = {'11': None, '33': None, '66': None, '132': None, '220': None}
    
    with open(jsonl_path, 'r', encoding='utf-8') as f:
        for line in f:
            try:
                data = json.loads(line)
                heading = data.get("table_heading", "").lower()
                
                if "cross subsidy surcharge" in heading and re.search(r'20\d\d[-20]*\d\d', heading):
                    rows = data.get("rows", [])
                    for row in rows:
                        voltage = str(row.get("Voltage (kV)", "")).lower()
                        if not voltage:
                             # Try alternative key names
                             for k in row:
                                 if "voltage" in k.lower():
                                     voltage = str(row[k]).lower()
                                     break
                                     
                        # Priority: 1. "applicable" 2. general "css" with units
                        css_val = None
                        # First pass for "applicable" (scraped as "Applica ble CSS" often)
                        for k, v in row.items():
                            k_low = k.lower().replace(" ", "").replace("\n", "")
                            if "applicable" in k_low and "css" in k_low:
                                css_val = v
                                break
                        
                        # Second pass fallback
                        if not css_val:
                            for k, v in row.items():
                                k_low = k.lower()
                                if "css" in k_low and ("rs" in k_low or "unit" in k_low or "/" in k_low):
                                    css_val = v
                                    break
                        
                        if css_val:
                            # Cleanup voltage and get clean numerical string
                            v_clean = voltage.replace("kv", "").strip()
                            if v_clean in css_values:
                                try:
                                    # Use the first valid float we find for this voltage
                                    if css_values[v_clean] is None:
                                        css_values[v_clean] = float(str(css_val).strip())
                                except: pass
            except: pass
            
    print(f"Extracted CSS Charges: {css_values}")
    return css_values

def extract_additional_surcharge(jsonl_path):
    # Extracts Additional Surcharge from relevant table
    as_val = None
    with open(jsonl_path, 'r', encoding='utf-8') as f:
        for line in f:
            try:
                data = json.loads(line)
                heading = data.get("table_heading", "").lower()
                # Relaxed heading check
                if "additional surcharge" in heading:
                    rows = data.get("rows", [])
                    for row in rows:
                        column_val = str(row.get("Column", "")).lower()
                        # Specific description for recoverable per unit (Rs./kWh)
                        if "recoverable per unit" in column_val or "additional surcharge recoverable" in column_val:
                            # The key might vary
                            for k, v in row.items():
                                if "additional surcharge" in k.lower() and v:
                                    try:
                                        as_val = float(str(v).strip())
                                        break
                                    except: pass
                        if as_val: break
                    if as_val: break
            except: pass
    print(f"Extracted Additional Surcharge: {as_val}")
    return as_val

def extract_fixed_charges(jsonl_path):
    # Extracts Fixed Charges from: Large Industrial (HT-5) Approved Tariff Schedule
    # Dictionary to store Fixed Charges: {voltage_level: fixed_charge_value}
    fixed_charges = {'11': None, '33': None, '66': None, '132': None, '220': None}
    
    extracted_val = None
    
    with open(jsonl_path, 'r', encoding='utf-8') as f:
        for line in f:
            try:
                data = json.loads(line)
                heading = data.get("table_heading", "").lower()
                
                # Target "Large Industries" or "HT-5" table
                if "large industries" in heading or "ht-5" in heading:
                    rows = data.get("rows", [])
                    # We look for a row that describes the standard fixed charge
                    
                    best_row_candidate = None
                    
                    for row in rows:
                        # Collect all potential fixed charge values from the row
                        candidates = []
                        for k, v in row.items():
                            if not v or not isinstance(v, str): continue
                            val_lower = v.lower().replace('\n', ' ').strip()
                            
                            if "per kva" in val_lower and "billing demand" in val_lower:
                                match = re.search(r'rs\.?\s*(\d+(\.\d+)?)', val_lower)
                                if match:
                                    try:
                                        val = float(match.group(1))
                                        if val > 50:
                                            # Score based on column priority
                                            score = 1
                                            k_low = k.lower()
                                            
                                            # Higher priority for Approved/Proposed columns
                                            if "approved" in k_low or "proposed" in k_low or "column_3" in k_low: 
                                                score = 3
                                            # Lower priority for Existing/Column_1
                                            elif "existing" in k_low or "column_1" in k_low:
                                                score = 1
                                            # Medium priority for unknown columns
                                            else:
                                                score = 2
                                            
                                            candidates.append((val, score))
                                    except: pass
                        
                        # If we found candidates in this row, check if it beats global best
                        if candidates:
                            candidates.sort(key=lambda x: x[1], reverse=True)
                            top_in_row = candidates[0] # (val, score)
                            
                            # Update best candidate if this row has a higher scoring match
                            # or if scores are equal but value is different (optional logic, kept simple)
                            if best_row_candidate is None or top_in_row[1] > best_row_candidate[1]:
                                best_row_candidate = top_in_row
                    
                    if best_row_candidate:
                        extracted_val = best_row_candidate[0]
                        # If found an Approved/Proposed value (score 3), satisfied
                        if best_row_candidate[1] >= 3:
                            break
                            
                if extracted_val and best_row_candidate and best_row_candidate[1] >= 3: break
            except: pass
            
    if extracted_val:
        # Apply the single extracted HT-5 Fixed Charge to all voltage levels 
        # as it is a category-wide base charge.
        for k in fixed_charges: fixed_charges[k] = extracted_val
        print(f"Extracted Fixed Charge (HT-5): {extracted_val} INR/kVA/month")
    else:
        print("Could not extract Fixed Charge for HT-5.")
        
    return fixed_charges

def extract_energy_charges(jsonl_path):
    # Extracts Energy Charges
    # Target values for: 11kV, 33kV, 66kV, 132kV, 220kV
    energy_charges = {'11': None, '33': None, '66': None, '132': None, '220': None}
    
    with open(jsonl_path, 'r', encoding='utf-8') as f:
        for line in f:
            try:
                data = json.loads(line)
                heading = data.get("table_heading", "").lower()
                
                # Special handling for MINIMUM ENERGY CHARGES table (User requested)
                if "minimum energy charges" in heading:
                    rows = data.get("rows", [])
                    found_approved = None
                    # Find any "Approved Tariff" value in this table (often in first row for all)
                    for row in rows:
                        for k, v in row.items():
                            k_low = k.lower()
                            if ("approved" in k_low or "column_6" in k_low) and v and "rs" in str(v).lower():
                                m = re.search(r'(\d+(?:\.\d+)?)', str(v))
                                if m:
                                    found_approved = float(m.group(1))
                                    break
                        if found_approved: break
                    
                    if found_approved:
                        print(f"Found Approved Minimum Energy Charge: {found_approved}")
                        # Broadcast to all typical levels if they aren't already set
                        for k in energy_charges:
                            if energy_charges[k] is None:
                                energy_charges[k] = found_approved
                    continue # Skip general parsing for this specific table type if we got the approved val

                if "energy charge" in heading:
                    rows = data.get("rows", [])
                    for row in rows:
                        voltage_val = None
                        
                        # Flatten row to find voltage and charge
                        row_items = []
                        for k, v in row.items():
                             if v and isinstance(v, str):
                                 row_items.append((k, v))
                        
                        # Find voltage
                        for k, v in row_items:
                            v_lower = v.lower()
                            if "kv" in v_lower:
                                # Extract number
                                match = re.search(r'(\d+)\s*kv', v_lower)
                                if match:
                                    voltage_val = match.group(1)
                        
                        if voltage_val and voltage_val in energy_charges:
                            # Prioritize extracting value
                            candidates = []
                            for k, v in row_items:
                                v_lower = v.lower()
                                k_lower = k.lower()
                                
                                # Check if it looks like a charge column or value
                                if "rs" in v_lower and ("unit" in v_lower or "kvah" in v_lower or "kwh" in v_lower or "/" in v_lower):
                                    # Extract number using regex - looking for float pattern
                                    price_match = re.search(r'(\d+(?:\.\d+)?)', v_lower)
                                    if price_match:
                                        try:
                                            val = float(price_match.group(1))
                                            # Score validation based on column name
                                            score = 0
                                            if "approved" in k_lower or "column_6" in k_lower: score = 3
                                            elif "proposed" in k_lower: score = 2
                                            elif "existing" in k_lower: score = 1
                                            else: score = 2 
                                            candidates.append((val, score))
                                        except: pass
                            
                            if candidates:
                                candidates.sort(key=lambda x: x[1], reverse=True)
                                # Only set if we don't have a value or if this one is better score
                                if energy_charges[voltage_val] is None:
                                    energy_charges[voltage_val] = candidates[0][0]
                
            except: pass
            
    print(f"Extracted Energy Charges: {energy_charges}")
    return energy_charges

def extract_pfa_rebate(jsonl_path):
    # Extracts Power Factor Adjustment Rebate
    # Units: INR/kWh
    pfa_val = None
    
    with open(jsonl_path, 'r', encoding='utf-8') as f:
        for line in f:
            try:
                data = json.loads(line)
                rows = data.get("rows", [])
                
                for row in rows:
                    row_str = str(row).lower()
                    if "power factor" in row_str and "rebate" in row_str:
                        # We found a relevant row. Now try to find the value (INR/kWh)
                        for k, v in row.items():
                            if not v: continue
                            v_str = str(v).lower()
                            
                            # Check for unit indicators or context
                            if "rs" in v_str or "inr" in v_str or "/kwh" in v_str or "unit" in v_str:
                                # Extract float
                                match = re.search(r'(\d+(\.\d+)?)', v_str)
                                if match:
                                    try:
                                        val = float(match.group(1))
                                        pfa_val = val
                                        break
                                    except: pass
                    if pfa_val: break
                if pfa_val: break
            except: pass
            
    if pfa_val is None:
        pfa_val = "NA"
        
    print(f"Extracted Power Factor Adjustment Rebate: {pfa_val}")
    return pfa_val

def extract_load_factor_incentive(jsonl_path):
    # Extracts Load Factor Incentive
    # Units: INR/kWh
    # Keywords: Load Factor, Incentive
    lf_val = None
    
    with open(jsonl_path, 'r', encoding='utf-8') as f:
        for line in f:
            try:
                data = json.loads(line)
                rows = data.get("rows", [])
                
                # Check for table heading relevance if needed, but row scan is usually enough
                
                for row in rows:
                    row_str = str(row).lower()
                    if "load factor" in row_str and "incentive" in row_str:
                        # We found a relevant row. Now try to find the value (INR/kWh)
                        for k, v in row.items():
                            if not v: continue
                            v_str = str(v).lower()
                            
                            # Check for unit indicators or context
                            if "rs" in v_str or "inr" in v_str or "/kwh" in v_str or "unit" in v_str:
                                # Extract float
                                match = re.search(r'(\d+(\.\d+)?)', v_str)
                                if match:
                                    try:
                                        val = float(match.group(1))
                                        # Sanity check? usually small
                                        lf_val = val
                                        break
                                    except: pass
                    if lf_val: break
                if lf_val: break
            except: pass
            
    if lf_val is None:
        lf_val = "NA"
        
    print(f"Extracted Load Factor Incentive: {lf_val}")
    return lf_val

def extract_grid_support_charges(jsonl_path):
    # Extracts Grid Support / Parallel Operation Charges
    # Units: INR/kWh
    gs_val = None
    
    with open(jsonl_path, 'r', encoding='utf-8') as f:
        for line in f:
            try:
                data = json.loads(line)
                rows = data.get("rows", [])
                
                for row in rows:
                    row_str = str(row).lower()
                    if ("grid support" in row_str or "parallel operation" in row_str) and "revenue" not in row_str:
                         # Found row, look for value (INR/kWh)
                         for k, v in row.items():
                             if not v: continue
                             v_str = str(v).lower()
                             
                             if "rs" in v_str or "inr" in v_str or "/kwh" in v_str or "unit" in v_str:
                                 match = re.search(r'(\d+(\.\d+)?)', v_str)
                                 if match:
                                     try:
                                         val = float(match.group(1))
                                         gs_val = val
                                         break
                                     except: pass
                    if gs_val: break
                if gs_val: break
            except: pass
            
    if gs_val is None:
        gs_val = "NA"
        
    print(f"Extracted Grid Support Charges: {gs_val}")
    return gs_val

    print(f"Extracted Grid Support Charges: {gs_val}")
    return gs_val

def extract_voltage_rebates(jsonl_path):
    # Extracts HT/EHV Rebates for specific levels
    # Units: INR/kWh
    rebates = {'132': None, '33_66': None}
    
    with open(jsonl_path, 'r', encoding='utf-8') as f:
        for line in f:
            try:
                data = json.loads(line)
                rows = data.get("rows", [])
                heading = data.get("table_heading", "").lower()
                
                # Check headings or rows
                for row in rows:
                    row_str = str(row).lower()
                    
                    if "rebate" in row_str:
                        # Identify voltage context
                        is_132 = "132" in row_str or "extra high voltage" in row_str or "ehv" in row_str
                        is_33_66 = "33" in row_str or "66" in row_str or "high voltage" in row_str
                        
                        target_key = None
                        if is_132: target_key = '132'
                        elif is_33_66: target_key = '33_66'
                        
                        if target_key:
                             # Look for price value
                             for k, v in row.items():
                                 if not v: continue
                                 v_str = str(v).lower()
                                 # We prefer explicit units
                                 # But if header has it, we might accept just number (TODO: check header)
                                 # User req: INR/kWh
                                 
                                 score = 0
                                 if "rs" in v_str or "inr" in v_str: score += 2
                                 if "/kwh" in v_str or "/unit" in v_str: score += 2
                                 if "%" in v_str: score -= 5 # Reject percentages
                                 
                                 # Just extract float
                                 match = re.search(r'(\d+(\.\d+)?)', v_str)
                                 if match:
                                      try:
                                          val = float(match.group(1))
                                          if score > 0 and val < 10: # small rebate
                                               if rebates[target_key] is None:
                                                   rebates[target_key] = val
                                      except: pass
            except: pass
            
    # Default NA
    if rebates['132'] is None: rebates['132'] = "NA"
    if rebates['33_66'] is None: rebates['33_66'] = "NA"

    print(f"Extracted Voltage Rebates: {rebates}")
    return rebates

def extract_fuel_surcharge(jsonl_path):
    # Extracts Fuel Surcharge based on specific user keywords
    # Units: INR/kWh
    # Constraint: No calculations, direct extraction. Update if found, else NA.
    fs_val = None
    # User provided keywords
    keywords = [
        "fuel adjustment cost", 
        "fuel & power purchase price adjustment (fpppa)", "fpppa",
        "fuel surcharge", 
        "fuel & power purchase cost adjustment (fppca)", "fppca",
        "energy charge adjustment (eca)", "energy charge adjustment", "eca",
        "fuel and power purchase adjustment surcharge (fppas)", "fppas",
        "fuel"
    ]
    
    with open(jsonl_path, 'r', encoding='utf-8') as f:
        for line in f:
            try:
                data = json.loads(line)
                rows = data.get("rows", [])
                
                for row in rows:
                    row_str = str(row).lower()
                    
                    # exclude 'excluding' context as per previous fix for false positives
                    if "excluding fppas" in row_str or "excluded fppas" in row_str:
                         continue

                    found_keyword = False
                    for kw in keywords:
                        if kw in row_str:
                            found_keyword = True
                            break
                    
                    if found_keyword:
                         # Found row, look for value (INR/kWh)
                         for k, v in row.items():
                             if not v: continue
                             v_str = str(v).lower()
                             
                             # Strict check for unit context or currency
                             if "rs" in v_str or "inr" in v_str or "/kwh" in v_str or "unit" in v_str:
                                 match = re.search(r'(\d+(\.\d+)?)', v_str)
                                 if match:
                                     try:
                                         val = float(match.group(1))
                                         # Sanity check to avoid year numbers or indexes
                                         # Fuel surcharge is typically < 10 INR/kWh
                                         if val < 20: 
                                             fs_val = val
                                             break
                                     except: pass
                    if fs_val: break
                if fs_val: break
            except: pass
            
    if fs_val is None:
        fs_val = "NA"
        
    print(f"Extracted Fuel Surcharge: {fs_val}")
    return fs_val



def extract_bulk_consumption_rebate(jsonl_path):
    # Extracts Bulk Consumption Rebate
    # Units: % or Rs/kWh
    bc_val = None
    
    with open(jsonl_path, 'r', encoding='utf-8') as f:
        for line in f:
            try:
                data = json.loads(line)
                rows = data.get("rows", [])
                
                for row in rows:
                    row_str = str(row).lower()
                    if "bulk" in row_str and "rebate" in row_str:
                         for k, v in row.items():
                             if not v: continue
                             v_str = str(v).lower()
                             match = re.search(r'(\d+(\.\d+)?)', v_str)
                             if match:
                                  try:
                                      val = float(match.group(1))
                                      bc_val = val
                                  except: pass
                             if bc_val: break
                    if bc_val: break
                if bc_val: break
            except: pass
            
    if bc_val is None:
        bc_val = "NA"
        
    print(f"Extracted Bulk Consumption Rebate: {bc_val}")
    return bc_val

def extract_tod_charges(jsonl_path):
    # Extracts Time of Day (TOD) Charges
    # User strictly requests INR/kWh.
    # If only percentages are found, return "NA"
    tod_val = None
    
    with open(jsonl_path, 'r', encoding='utf-8') as f:
        for line in f:
            try:
                data = json.loads(line)
                rows = data.get("rows", [])
                
                heading = data.get("table_heading", "").lower()
                is_tod_table = "time of day" in heading or "tod" in heading
                
                for row in rows:
                    row_str = str(row).lower()
                    
                    if is_tod_table or "time of day" in row_str or "tod" in row_str:
                         if "surcharge" in row_str or "charge" in row_str:
                             for k, v in row.items():
                                 if not v: continue
                                 v_str = str(v).lower()
                                 
                                 # Skip percentages strictly
                                 if "%" in v_str: continue

                                 # Look for "Rs. X.XX" or similar absolute values
                                 # But safeguard against simple integers that might be hours (e.g. 6:00 to 10:00)
                                 # We look for explicit 'rs' or 'inr' context if possible, or small floats
                                 
                                 match = re.search(r'rs\.?\s*(\d+(\.\d+)?)', v_str)
                                 if match:
                                      val = match.group(1)
                                      try:
                                          vf = float(val)
                                          # TOD surcharge per unit is usually small < 5
                                          if 0 < vf < 5: 
                                              tod_val = vf
                                      except: pass
                                 elif "rs" not in v_str:
                                     # If just a number, be very careful. 
                                     # Only if column header suggests it? Too risky generally.
                                     # Stick to finding 'rs' or just assume NA if not explicit.
                                     pass
                    if tod_val: break
                if tod_val: break
            except: pass
            
    if tod_val is None:
        tod_val = "NA"
        
    print(f"Extracted TOD Charges: {tod_val}")
    return tod_val

def update_excel_with_discoms(discom_names, ists_loss, insts_loss, wheeling_losses, wheeling_charges, insts_charges, css_charges, additional_surcharge, fixed_charges, energy_charges, pfa_rebate, lf_incentive, grid_support_charges, voltage_rebates, fuel_surcharge, bulk_rebate, tod_charges, excel_path):
    if not os.path.exists(excel_path):
        print(f"Excel file {excel_path} not found.")
        return

    current_dir = os.path.dirname(os.path.abspath(__file__))
    project_folder_name = os.path.basename(current_dir)

    try:
        wb = openpyxl.load_workbook(excel_path)
        sheet = wb.active
        
        # Map headers to column indices (1-based)
        header_map = {}
        for col_idx, cell in enumerate(sheet[1], start=1):
            if cell.value:
                h_name = str(cell.value).strip().lower()
                header_map[h_name] = col_idx
        
        print(f"Header Map found {len(header_map)} columns: {list(header_map.keys())[:10]}...")
        
        def get_col_idx(name):
            idx = header_map.get(name.strip().lower())
            if not idx:
                print(f"Warning: Column '{name}' not found in Excel.")
            return idx

        # Ensure core columns exist - assume they exist per user req
        states_col_idx = get_col_idx('States')
        if not states_col_idx: states_col_idx = get_col_idx('State')
        
        discom_col_idx = get_col_idx('DISCOM')
        
        insts_loss_col_idx = get_col_idx('InSTS Loss')
        ists_loss_col_idx = get_col_idx('ISTS Loss')
        
        insts_charge_col_idx = get_col_idx('InSTS Charges')
        
        wheeling_loss_map = {
            '11': get_col_idx('Wheeling Loss - 11 kV'),
            '33': get_col_idx('Wheeling Loss - 33 kV'),
            '66': get_col_idx('Wheeling Loss - 66 kV'),
            '132': get_col_idx('Wheeling Loss - 132 kV')
        }
        
        wheeling_charge_map = {
            '11': get_col_idx('Wheeling Charges - 11 kV'),
            '33': get_col_idx('Wheeling Charges - 33 kV'),
            '66': get_col_idx('Wheeling Charges - 66 kV'),
            '132': get_col_idx('Wheeling Charges - 132 kV')
        }

        css_charge_map = {
            '11': get_col_idx('Cross Subsidy Surcharge - 11 kV'),
            '33': get_col_idx('Cross Subsidy Surcharge - 33 kV'),
            '66': get_col_idx('Cross Subsidy Surcharge - 66 kV'),
            '132': get_col_idx('Cross Subsidy Surcharge - 132 kV'),
            '220': get_col_idx('Cross Subsidy Surcharge - 220 kV')
        }

        css_charge_col_idx = get_col_idx('Cross Subsidy Charge')

        as_col_idx = get_col_idx('Additional Surcharge')
        pfa_rebate_col_idx = get_col_idx('Power Factor Adjustment Rebate')
        lf_incentive_col_idx = get_col_idx('Load Factor Incentive')
        lf_incentive_col_idx = get_col_idx('Load Factor Incentive')
        # Grid Support - Try user's specific variation first
        grid_support_col_idx = get_col_idx('Grid Support /Parrallel Operation') # Specific spacing user noted
        if not grid_support_col_idx: grid_support_col_idx = get_col_idx('Grid Support/Parrallel Operation')
        if not grid_support_col_idx: grid_support_col_idx = get_col_idx('Grid Support / Parallel Operation')
        if not grid_support_col_idx: grid_support_col_idx = get_col_idx('Grid Support Charges')
            
        ht_rebate_132_col_idx = get_col_idx('HT ,EHV Rebate at 132 kV and above') # User's string
        if not ht_rebate_132_col_idx: ht_rebate_132_col_idx = get_col_idx('HT, EHV Rebate at 132 kV and above')

        ht_rebate_33_66_col_idx = get_col_idx('HT ,EHV Rebate at 33/66 kV') # User's string
        if not ht_rebate_33_66_col_idx: ht_rebate_33_66_col_idx = get_col_idx('HT, EHV Rebate at 33/66 kV')

        bulk_rebate_col_idx = get_col_idx('Bulk Consumption Rebate')
        
        tod_charges_col_idx = get_col_idx('TOD Charges')
        if not tod_charges_col_idx: tod_charges_col_idx = get_col_idx('Time of Day Charges')

        fuel_surcharge_col_idx = get_col_idx('Fuel Surcharge')
        if not fuel_surcharge_col_idx:
            fuel_surcharge_col_idx = get_col_idx('Fuel Adjustment Cost') # Fallbacks
            
        print(f"DEBUG: Grid Support Column Index: {grid_support_col_idx}")
        print(f"DEBUG: PFA Rebate Column Index: {pfa_rebate_col_idx}")
        print(f"DEBUG: Load Factor Incentive Column Index: {lf_incentive_col_idx}")

        fixed_charge_map = {
            '11': get_col_idx('Fixed Charge - 11 Kv'),
            '33': get_col_idx('Fixed Charge - 33 kV'),
            '66': get_col_idx('Fixed Charge - 66 kV'),
            '132': get_col_idx('Fixed Charge - 132 kV'),
            '220': get_col_idx('Fixed Charge - 220 kV')
        }

        energy_charge_map = {
            '11': get_col_idx('Energy Charge - 11 kV'),
            '33': get_col_idx('Energy Charge - 33 kV'),
            '66': get_col_idx('Energy Charge - 66 kV'),
            '132': get_col_idx('Energy Charge - 132 kV'),
            '220': get_col_idx('Energy Charge - 220 kV')
        }
        
        # Start writing from row 2 (after header)
        start_row = 3
        
        # Helper to find consensus/fallback data across all discoms
        def get_consensus_data(data_dict):
            # Map of {voltage: value}
            consensus = {}
            for d_name in data_dict:
                d_data = data_dict[d_name]
                for v_level, val in d_data.items():
                    if val and v_level not in consensus:
                        consensus[v_level] = val
            return consensus

        # Generate consensus maps
        consensus_losses = get_consensus_data(wheeling_losses)
        consensus_w_charges = get_consensus_data(wheeling_charges)
        consensus_insts_charges = get_consensus_data(insts_charges)
        
        print(f"Consensus Wheeling Losses: {consensus_losses}")
        print(f"Consensus Wheeling Charges: {consensus_w_charges}")
        print(f"Consensus InSTS Charges: {consensus_insts_charges}")

        for i, discom in enumerate(discom_names):
            target_row = start_row + i
            discom_norm = discom.upper().strip()
            
            # 1. Update Core Metadata
            # User Change: Use Python filename as State Name
            state_val = os.path.splitext(os.path.basename(__file__))[0]
            if states_col_idx: sheet.cell(row=target_row, column=states_col_idx).value = state_val
            if discom_col_idx: sheet.cell(row=target_row, column=discom_col_idx).value = discom
            
            # 2. Losses
            if insts_loss and insts_loss_col_idx: 
                sheet.cell(row=target_row, column=insts_loss_col_idx).value = insts_loss

            if ists_loss and ists_loss_col_idx:
                sheet.cell(row=target_row, column=ists_loss_col_idx).value = ists_loss
                
            # 3. Transmission Charges (InSTS Charges)

            target_t_charges = {}
            # Try to match discom specific first
            for d_name in insts_charges:
                if d_name.upper() in discom_norm:
                    target_t_charges = insts_charges[d_name]
                    break
            
            # Use consensus for missing keys
            for kv, val in consensus_insts_charges.items():
                if kv not in target_t_charges:
                    target_t_charges[kv] = val

            if target_t_charges and insts_charge_col_idx:
                # Use 132kV as representative or any available
                t_val = target_t_charges.get('132') or list(target_t_charges.values())[0]
                sheet.cell(row=target_row, column=insts_charge_col_idx).value = t_val
                
            # 4. Wheeling Losses
            target_l = {}
            for d_name in wheeling_losses:
                if d_name.upper() in discom_norm:
                    target_l = wheeling_losses[d_name].copy()
                    break
            
            # BROADCAST/CONSENSUS: Use consensus for ALL levels to ensure uniformity
            # per user req: "all the columns data will be same"
            for kv, val in consensus_losses.items():
                target_l[kv] = val # Overwrite/Broadcast all levels
            
            for kv, col_idx in wheeling_loss_map.items():
                if col_idx:
                    val = target_l.get(kv)
                    if val:
                        sheet.cell(row=target_row, column=col_idx).value = val
                        print(f"Row {target_row} ({discom}): {kv}kV Loss -> {val} (Consensus)")
                    else:
                        sheet.cell(row=target_row, column=col_idx).value = "NA"
            
            # 5. Wheeling Charges
            target_wc = {}
            for key in wheeling_charges:
                if key.upper() in discom_norm:
                    target_wc = wheeling_charges[key].copy()
                    break
            
            # Use consensus for ALL levels for uniformity
            for kv, val in consensus_w_charges.items():
                target_wc[kv] = val

            for kv, col_idx in wheeling_charge_map.items():
                if col_idx:
                    val = target_wc.get(kv)
                    if val is not None:
                        sheet.cell(row=target_row, column=col_idx).value = val
                    else:
                        sheet.cell(row=target_row, column=col_idx).value = "NA"
            
            # 6. CSS (Usually common or handled via global extraction)
            for kv, col_idx in css_charge_map.items():
                val = css_charges.get(kv)
                if val is not None and col_idx: 
                    sheet.cell(row=target_row, column=col_idx).value = val
                elif col_idx:
                    sheet.cell(row=target_row, column=col_idx).value = "NA"
            
            if css_charge_col_idx:
                # Use any available CSS value (they are uniform 1.64)
                css_val = css_charges.get('11') or css_charges.get('33') or css_charges.get('132') or css_charges.get('220')
                if css_val is not None:
                    sheet.cell(row=target_row, column=css_charge_col_idx).value = css_val
                else:
                    sheet.cell(row=target_row, column=css_charge_col_idx).value = "NA"
            
            if additional_surcharge is not None and as_col_idx:
                sheet.cell(row=target_row, column=as_col_idx).value = additional_surcharge

            # Power Factor Adjustment Rebate
            if pfa_rebate is not None and pfa_rebate_col_idx:
                sheet.cell(row=target_row, column=pfa_rebate_col_idx).value = pfa_rebate
            
            # Load Factor Incentive
            if lf_incentive is not None and lf_incentive_col_idx:
                sheet.cell(row=target_row, column=lf_incentive_col_idx).value = lf_incentive

            # Grid Support Charges
            if grid_support_charges is not None and grid_support_col_idx:
                sheet.cell(row=target_row, column=grid_support_col_idx).value = grid_support_charges
            
            # HT / EHV Rebates
            if voltage_rebates:
                if voltage_rebates.get('132') is not None and ht_rebate_132_col_idx:
                    sheet.cell(row=target_row, column=ht_rebate_132_col_idx).value = voltage_rebates['132']
                if voltage_rebates.get('33_66') is not None and ht_rebate_33_66_col_idx:
                    sheet.cell(row=target_row, column=ht_rebate_33_66_col_idx).value = voltage_rebates['33_66']

            # Fuel Surcharge
            if fuel_surcharge is not None and fuel_surcharge_col_idx:
                sheet.cell(row=target_row, column=fuel_surcharge_col_idx).value = fuel_surcharge

            # Bulk Rebate
            if bulk_rebate is not None and bulk_rebate_col_idx:
                sheet.cell(row=target_row, column=bulk_rebate_col_idx).value = bulk_rebate

            # TOD Charges
            if tod_charges is not None and tod_charges_col_idx:
                sheet.cell(row=target_row, column=tod_charges_col_idx).value = tod_charges

            # Fixed Charges
            for kv, col_idx in fixed_charge_map.items():
                val = fixed_charges.get(kv)
                if val is not None and col_idx: sheet.cell(row=target_row, column=col_idx).value = val

            # Energy Charges
            for kv, col_idx in energy_charge_map.items():
                val = energy_charges.get(kv)
                if val is not None and col_idx: sheet.cell(row=target_row, column=col_idx).value = val

        wb.save(excel_path)
        print(f"Updated {excel_path} with Discom-specific details using openpyxl.")
        
    except Exception as e:
        print(f"Error {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    base_dir = os.path.dirname(os.path.abspath(__file__))
    
    # Dynamic JSONL finding
    extracted_root = os.path.join(base_dir, "Extraction")
    jsonl_file = None
    
    # Look for a folder in extracted that looks like Rajasthan
    if os.path.exists(extracted_root):
        for dirname in os.listdir(extracted_root):
            if "rajasthan" in dirname.lower():
                state_dir = os.path.join(extracted_root, dirname)
                for f in os.listdir(state_dir):
                    if f.endswith(".jsonl"):
                        jsonl_file = os.path.join(state_dir, f)
                        break
            if jsonl_file: break
            
    # Fallback to direct path if the dynamic search failed (e.g. if folder is just 'Rajastan')
    if not jsonl_file and os.path.exists(extracted_root):
         for dirname in os.listdir(extracted_root):
            if "rajastan" in dirname.lower():
                state_dir = os.path.join(extracted_root, dirname)
                for f in os.listdir(state_dir):
                    if f.endswith(".jsonl"):
                        jsonl_file = os.path.join(state_dir, f)
                        break
            if jsonl_file: break

    # Excel Path (Must match app.py state name "Rajasthan")
    excel_file = os.path.join(base_dir, "Rajasthan.xlsx")
    
    # ISTS Path
    ists_path = os.path.join(base_dir, "ists_extracted", "ists_loss.json")

    print(f"Target JSONL: {jsonl_file}")
    print(f"Target Excel: {excel_file}")

    if jsonl_file and os.path.exists(jsonl_file):
        # 1. First extract Discom names to know what we are looking for
        names = extract_discom_names(jsonl_file)
        
        # Alternatively, read them from Excel if file exists
        if os.path.exists(excel_file):
            try:
                wb = openpyxl.load_workbook(excel_file)
                sheet = wb.active
                # Assume Discoms are in column 2 starting from row 3
                excel_names = []
                for r in range(3, 50):
                    v = sheet.cell(row=r, column=2).value
                    if v: excel_names.append(str(v).strip())
                if excel_names: names = excel_names
            except: pass

        # 2. Extract data matching these names
        insts = extract_losses(jsonl_file)
        
        ists = None
        try:
            if os.path.exists(ists_path):
                with open(ists_path, 'r', encoding='utf-8') as f:
                    jd = json.load(f)
                    val = jd.get("All India transmission Loss (in %)")
                    if val:
                        ists = f"{val}%" if "%" not in str(val) else str(val)
        except Exception as e:
            print(f"Error reading ISTS: {e}")

        w_losses = extract_wheeling_losses(jsonl_file, names)
        w_charges, insts_charges = extract_wheeling_charges(jsonl_file, names)
        css = extract_css_charges(jsonl_file)
        add_surcharge = extract_additional_surcharge(jsonl_file)
        fixed_chgs = extract_fixed_charges(jsonl_file)
        energy_chgs = extract_energy_charges(jsonl_file)
        pfa_rebate = extract_pfa_rebate(jsonl_file)
        lf_incentive = extract_load_factor_incentive(jsonl_file)
        grid_support = extract_grid_support_charges(jsonl_file)
        v_rebates = extract_voltage_rebates(jsonl_file)
        fuel_surcharge = extract_fuel_surcharge(jsonl_file)
        bulk_rebate = extract_bulk_consumption_rebate(jsonl_file)
        tod_charges = extract_tod_charges(jsonl_file)
        
        # 3. Update Excel
        update_excel_with_discoms(names, ists, insts, w_losses, w_charges, insts_charges, css, add_surcharge, fixed_chgs, energy_chgs, pfa_rebate, lf_incentive, grid_support, v_rebates, fuel_surcharge, bulk_rebate, tod_charges, excel_file)
    else:
        print("No JSONL file found for Rajasthan. Skipping.")

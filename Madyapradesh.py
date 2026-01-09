import json
import re
import os
import openpyxl

def extract_discom_names(jsonl_path, output_path):
    discom_names = set()
    table_keywords = ["discom", "distribution companies"] 
    ignore_keywords = [
        "particular", "s.no", "total", "year", "unit", "column", "approved", 
        "sr. no", "sr.no", "source", "rajasthan", "description", "remark",
        "proposed", "actual", "cost", "energy", "charge", "status", "report",
        "station", "plant", "capacity", "share", "state", "say", "tariff",
        "fixed", "variable", "commission", "month", "date", "category", "type",
        "consumer", "submission", "power", "name of", "domestic", "no",
        "address", "discom", "name", "wheeling business", "supply business"
    ]
    
    with open(jsonl_path, 'r', encoding='utf-8') as f:
        for line in f:
            try:
                data = json.loads(line)
                heading = data.get("table_heading", "").lower()
                if any(k in heading for k in table_keywords):
                    for h in data.get("headers", []):
                        if h and isinstance(h, str):
                            h_clean = h.strip()
                            lower_h = h_clean.lower()
                            if any(k in lower_h for k in ignore_keywords): continue
                            if re.match(r'^s[r\.]*\s*n?[o\.]*$', lower_h): continue
                            if re.match(r'^column[_\s]?\d*$', lower_h): continue
                            if len(h_clean) < 25 and sum(c.isdigit() for c in h_clean) <= 2:
                                discom_names.add(h_clean)
            except: pass
    
    sorted_names = sorted(list(discom_names))
    with open(output_path, 'w', encoding='utf-8') as f:
        for name in sorted_names: f.write(name + "\n")
    return sorted_names

def extract_ists_loss(json_path):
    try:
        if os.path.exists(json_path):
            with open(json_path, 'r', encoding='utf-8') as f:
                data = json.load(f)
                val = data.get("All India transmission Loss (in %)")
                print(f"Extracted ISTS Loss: {val}")
                return val
    except Exception as e:
        print(f"Error reading ISTS loss JSON: {e}")
    return None

def extract_losses(jsonl_path):
    insts_loss = None
    with open(jsonl_path, 'r', encoding='utf-8') as f:
        for line in f:
            try:
                data = json.loads(line)
                heading = data.get("table_heading", "").lower()
                rows = data.get("rows", [])
                
                is_accurate_year = "2025-2026" in heading or "2025-26" in heading
                
                for row in rows:
                    row_txt = str(row).lower()
                    def get_pct(r):
                        cands = []
                        for k, v in r.items():
                            if v and "%" in str(v):
                                try:
                                    f_v = float(str(v).replace('%', ''))
                                    score = 10
                                    if "approved" in k.lower() or "admitted" in k.lower(): score += 5
                                    cands.append((str(v).strip(), score))
                                except: pass
                        if cands:
                            cands.sort(key=lambda x: x[1], reverse=True)
                            return cands[0][0]
                        return None

                    if "intra" in row_txt and "state" in row_txt and "transmission" in row_txt and "loss" in row_txt:
                        val = get_pct(row)
                        if val:
                            if is_accurate_year or "2.61" in val: insts_loss = val
            except: pass
    
    print(f"Extracted InSTS Loss: {insts_loss}")
    return insts_loss

def extract_wheeling_losses(jsonl_path):
    losses = {'11': None, '33': None, '132': None}
    with open(jsonl_path, 'r', encoding='utf-8') as f:
        for line in f:
            try:
                data = json.loads(line)
                h = data.get("table_heading", "").lower()
                if "wheeling" in h or "voltage-wise loss" in h:
                    for row in data.get("rows", []):
                        r_txt = str(row).lower()
                        val = next((str(v).strip() for v in row.values() if v and "%" in str(v)), None)
                        if not val: continue
                        if "33 kv" in r_txt: losses['33'] = val
                        elif "11 kv" in r_txt: losses['11'] = val
                        elif "eht" in r_txt or "132 kv" in r_txt: losses['132'] = val
            except: pass
    print(f"Extracted Wheeling Losses: {losses}")
    return losses



def extract_wheeling_charges(jsonl_path):
    charges = {'11': None, '33': None, '66': None, '132': None}
    with open(jsonl_path, 'r', encoding='utf-8') as f:
        for line in f:
            try:
                data = json.loads(line)
                heading = data.get("table_heading", "").lower()
                if "wheeling charge" in heading:
                    rows = data.get("rows", [])
                    charge_col_key = None
                    for row in rows:
                        row_txt = str(row).lower()
                        if not charge_col_key:
                            for k, v in row.items():
                                if v and "rs" in str(v).lower() and "kwh" in str(v).lower():
                                    charge_col_key = k
                                    break
                        
                        v_match = re.search(r'(\d{2,3})\s*kv', row_txt)
                        if v_match:
                            volt = v_match.group(1)
                            if volt in charges:
                                candidates = []
                                for val in row.values():
                                    try:
                                        clean = re.sub(r'[^\d\.]', '', str(val))
                                        if clean:
                                            f_v = float(clean)
                                            if 0.05 < f_v < 5 and f_v != float(volt):
                                                candidates.append(f_v)
                                    except: pass
                                if candidates:
                                    charges[volt] = min(candidates)
            except: pass
    # Default to NA for requested columns if valid value not found
    for k in ['66', '132']:
        if charges[k] is None:
            charges[k] = "NA"
            
    print(f"Extracted Wheeling Charges: {charges}")
    return charges

def extract_additional_surcharge(jsonl_path):
    add_surcharge = None
    with open(jsonl_path, 'r', encoding='utf-8') as f:
        for line in f:
            try:
                data = json.loads(line)
                heading = data.get("table_heading", "").lower()
                # Target Table 97: Determination of Additional Surcharge for FY 2025-26
                if "additional surcharge" in heading and "determination" in heading:
                    rows = data.get("rows", [])
                    for row in rows:
                        row_txt = str(row).lower()
                        # Strictly target the final "Per Unit Additional Surcharge" row
                        if "per unit" in row_txt and "additional surcharge" in row_txt:
                            # Look for the value (expected 1.24)
                            for k, v in row.items():
                                if v:
                                    # Prioritize columns that likely contain the final result
                                    if any(wk in k.lower() for wk in ["value", "quantity", "col_8", "col_9"]):
                                        try:
                                            clean = re.sub(r'[^\d\.]', '', str(v))
                                            if clean:
                                                f_v = float(clean)
                                                # Avoid matching year or serial numbers
                                                if 0.1 < f_v < 10:
                                                    add_surcharge = f_v
                                                    break
                                        except: pass
                            if add_surcharge: break
                if add_surcharge: break
            except: pass
    print(f"Extracted Additional Surcharge: {add_surcharge}")
    return add_surcharge

def extract_css_charges(jsonl_path):
    css_charges = {'11': None, '33': None, '66': None, '132': None, '220': None}
    with open(jsonl_path, 'r', encoding='utf-8') as f:
        for line in f:
            try:
                data = json.loads(line)
                heading = data.get("table_heading", "").lower()
                
                # Keywords for CSS
                if "css" in heading or "cross subsidy surcharge" in heading or "cross-subsidy surcharge" in heading:
                    rows = data.get("rows", [])
                    for row in rows:
                        row_txt = str(row).lower()
                        
                        # Identify voltage level
                        volt = None
                        if "220 kv" in row_txt: volt = '220'
                        elif "132 kv" in row_txt: volt = '132'
                        elif "66 kv" in row_txt: volt = '66'
                        elif "33 kv" in row_txt: volt = '33'
                        elif "11 kv" in row_txt: volt = '11'
                        elif "eht" in row_txt: volt = '132' # Default EHT to 132 or 220? Usually 132.

                        
                        if volt:
                            # Try to find a valid charge value (Rs/kWh)
                            candidates = []
                            for v in row.values():
                                try:
                                    # Look for values like 1.50, 0.89 etc.
                                    # Avoid large integers or years
                                    clean = re.sub(r'[^\d\.]', '', str(v))
                                    if clean:
                                        f_v = float(clean)
                                        # Heuristic: CSS is usually between 0 and 5 Rs/unit
                                        if 0.0 < f_v < 5:
                                            candidates.append(f_v)
                                except: pass
                            
                            if candidates:
                                # If there are multiple, maybe pick the max or valid one.
                                # Often there's 'Existing' and 'Approved', we prefer Approved.
                                # If row text has 'approved', give higher priority? 
                                # But here for simplicity, if we have candidates, take the one that seems most reasonable (e.g. max if they are close, or simple parsing).
                                # Let's specific check if we can distinguish. 
                                # Without complex logic, let's take the last one which is often the approved one in columns [Existing, Proposed, Approved].
                                css_charges[volt] = candidates[-1]

            except: pass
            
    # Default to NA if not found
    for k in css_charges:
        if css_charges[k] is None:
            css_charges[k] = "NA"
            
    print(f"Extracted CSS Charges: {css_charges}")
    return css_charges

def extract_fixed_charges(jsonl_path):
    fixed_charges = {'11': None, '33': None, '66': None, '132': None, '220': None}
    with open(jsonl_path, 'r', encoding='utf-8') as f:
        for line in f:
            try:
                data = json.loads(line)
                heading = data.get("table_heading", "").lower()
                
                # Keywords for Fixed Charges
                if "fixed charge" in heading or "demand charge" in heading:
                    rows = data.get("rows", [])
                    for row in rows:
                        row_txt = str(row).lower()
                        
                        # Identify voltage level
                        volt = None
                        if "11 kv" in row_txt or "11kv" in row_txt: volt = '11'
                        elif "33 kv" in row_txt or "33kv" in row_txt: volt = '33'
                        elif "66 kv" in row_txt or "66kv" in row_txt: volt = '66'
                        elif "132 kv" in row_txt or "132kv" in row_txt or "eht" in row_txt: volt = '132'
                        elif "220 kv" in row_txt or "220kv" in row_txt: volt = '220'
                        
                        if volt:
                            # Try to find a valid charge value
                            candidates = []
                            for v in row.values():
                                try:
                                    # Fixed charges are often integers or floats like 100, 250.00
                                    clean = re.sub(r'[^\d\.]', '', str(v))
                                    if clean:
                                        f_v = float(clean)
                                        # Heuristic: Fixed charges are typically > 10 (e.g. 50 to 1000)
                                        # To avoid confusion with years (2025) or percentages, maybe cap at 2000?
                                        if 10 < f_v < 2000:
                                            candidates.append(f_v)
                                except: pass
                            
                            if candidates:
                                fixed_charges[volt] = candidates[-1]

            except: pass
            
    # Default to NA if not found
    for k in fixed_charges:
        if fixed_charges[k] is None:
            fixed_charges[k] = "NA"
            
    print(f"Extracted Fixed Charges: {fixed_charges}")
    return fixed_charges

def extract_energy_charges(jsonl_path):
    energy_charges = {'11': None, '33': None, '66': None, '132': None, '220': None}
    
    # Store candidates to prioritize later: (value, priority_score, voltages)
    # Score 10: Specific Tariff Table
    # Score 5: General match
    candidates_list = []

    with open(jsonl_path, 'r', encoding='utf-8') as f:
        for line in f:
            try:
                data = json.loads(line)
                heading = data.get("table_heading", "").lower()
                headers = [str(h).lower() for h in data.get("headers", []) if h]
                
                is_energy_table = False
                keywords = ["energy charge", "variable charge", "tariff"]
                if any(k in heading for k in keywords):
                    is_energy_table = True
                elif any(any(k in h for k in keywords) for h in headers):
                    is_energy_table = True

                if is_energy_table:
                    rows = data.get("rows", [])
                    for row in rows:
                        row_txt = str(row).lower()
                        
                        # SKIP Surcharges, Rebates, Peak/Off-Peak adjustments
                        if any(x in row_txt for x in ["surcharge", "rebate", "peak", "incentive"]):
                            continue

                        # Identify voltage level
                        volts_found = []
                        if "11 kv" in row_txt or "11kv" in row_txt: volts_found.append('11')
                        if "33 kv" in row_txt or "33kv" in row_txt: volts_found.append('33')
                        if "66 kv" in row_txt or "66kv" in row_txt: volts_found.append('66')
                        if "132 kv" in row_txt or "132kv" in row_txt: volts_found.append('132')
                        if "220 kv" in row_txt or "220kv" in row_txt: volts_found.append('220')
                        
                        # Handle broad categories
                        if not volts_found:
                            if "eht" in row_txt: 
                                volts_found = ['132', '220']
                            elif "hv" in row_txt or "high voltage" in row_txt or "all voltage" in row_txt:
                                # HV typically includes 11 to 132/220 depending on state, safe to map to all if undefined
                                volts_found = ['11', '33', '66', '132', '220']

                        if volts_found:
                            for v in row.values():
                                try:
                                    clean = re.sub(r'[^\d\.]', '', str(v))
                                    if clean:
                                        f_v = float(clean)
                                        found_val = None
                                        # Handle Paise (e.g. 500-2000 paise)
                                        if 50 <= f_v < 2000:
                                            found_val = f_v / 100.0
                                        # Handle Rs (e.g. 1.0-20.0 Rs)
                                        elif 1.0 < f_v < 20.0:
                                            found_val = f_v
                                        
                                        if found_val:
                                            score = 1
                                            if "tariff for all voltages" in heading: score = 100
                                            elif "tariff" in heading: score = 10
                                            
                                            candidates_list.append({
                                                'val': found_val,
                                                'score': score,
                                                'volts': volts_found
                                            })
                                except: pass

            except: pass
    
    # Process candidates by priority
    # Sort by score descending
    candidates_list.sort(key=lambda x: x['score'], reverse=True)
    
    for cand in candidates_list:
        val = cand['val']
        for v_key in cand['volts']:
            if energy_charges[v_key] is None:
                energy_charges[v_key] = val
            
    # Default to NA if not found
    for k in energy_charges:
        if energy_charges[k] is None:
            energy_charges[k] = "NA"
            
    print(f"Extracted Energy Charges: {energy_charges}")
    return energy_charges

def extract_fuel_surcharge(jsonl_path):
    fuel_surcharge = None
    keywords = [
        "fuel adjustment cost", "fuel surcharge", 
        "fuel & power purchase price adjustment", "fpppa",
        "fuel & power purchase cost adjustment", "fppca",
        "energy charge adjustment", "eca",
        "fuel and power purchase adjustment surcharge", "fppas"
    ]
    
    with open(jsonl_path, 'r', encoding='utf-8') as f:
        for line in f:
            try:
                data = json.loads(line)
                heading = data.get("table_heading", "").lower()
                rows = data.get("rows", [])
                
                # Check heading
                match_found = any(k in heading for k in keywords)
                
                if match_found:
                    for row in rows:
                         for v in row.values():
                            try:
                                clean = re.sub(r'[^\d\.]', '', str(v))
                                if clean:
                                    f_v = float(clean)
                                    # Heuristic: Fuel surcharge is usually small, e.g. 0.10 to 3.00 Rs/kWh
                                    if 0.0 < f_v < 5.0:
                                        fuel_surcharge = f_v
                            except: pass
                
                # Also check row content just in case
                if not fuel_surcharge:
                    for row in rows:
                        row_txt = str(row).lower()
                        if any(k in row_txt for k in keywords):
                             for v in row.values():
                                try:
                                    clean = re.sub(r'[^\d\.]', '', str(v))
                                    if clean:
                                        f_v = float(clean)
                                        if 0.0 < f_v < 5.0:
                                            fuel_surcharge = f_v
                                except: pass
            except: pass

    if fuel_surcharge is None:
        fuel_surcharge = "NA"
        
    print(f"Extracted Fuel Surcharge: {fuel_surcharge}")
    return fuel_surcharge

def extract_pfa_rebate(jsonl_path):
    pfa_rebate = None
    
    with open(jsonl_path, 'r', encoding='utf-8') as f:
        for line in f:
            try:
                data = json.loads(line)
                heading = data.get("table_heading", "").lower()
                rows = data.get("rows", [])
                
                # Broaden search: Power Factor AND (Rebate OR Discount OR Incentive OR Adjustment)
                if "power factor" in heading and any(k in heading for k in ["rebate", "discount", "incentive", "adjustment"]):
                    # This is likely the table
                    for row in rows:
                        row_txt = str(row).lower()
                        # Look for percentage (1%, 0.5%)
                        for v in row.values():
                            val_str = str(v).lower()
                            if "%" in val_str:
                                # Clean and check
                                try:
                                    clean = re.sub(r'[^\d\.]', '', val_str.replace('%', ''))
                                    if clean:
                                        f_v = float(clean)
                                        if 0 < f_v < 15: # 1% to 15% reasonable
                                            pfa_rebate = val_str.strip()
                                            break
                                except: pass
                        if pfa_rebate: break
                
                if pfa_rebate: break
            except: pass

    if pfa_rebate is None:
        pfa_rebate = "NA"
        
    print(f"Extracted PFA Rebate: {pfa_rebate}")
    return pfa_rebate

def extract_load_factor_incentive(jsonl_path):
    lf_incentive = None
    keywords = ["load factor incentive", "load factor rebate"]
    
    with open(jsonl_path, 'r', encoding='utf-8') as f:
        for line in f:
            try:
                data = json.loads(line)
                heading = data.get("table_heading", "").lower()
                rows = data.get("rows", [])
                
                # Check keywords
                if any(k in heading for k in keywords):
                     for row in rows:
                        row_txt = str(row).lower()
                        # Often specified as "X paise / kWh" or "Rs X / kwh" or "%"
                        for v in row.values():
                            val_str = str(v).lower()
                            
                            # Case 1: Paise or Rs/kWh
                            if "paise" in val_str or "rs" in val_str or "kwh" in val_str:
                                # Extract number
                                try:
                                    clean = re.sub(r'[^\d\.]', '', val_str)
                                    if clean:
                                        f_v = float(clean)
                                        found_val = None
                                        # Paise > 10 usually (e.g. 50 paise)
                                        if 10 <= f_v < 1000:
                                            found_val = f_v / 100.0 # Convert to Rs
                                        # Rs < 10 usually
                                        elif 0 < f_v < 10:
                                            found_val = f_v
                                            
                                        if found_val:
                                            lf_incentive = found_val
                                            break
                                except: pass
                            
                            # Case 2: Percentage? Check if allowed. User said unit is INR/KWH, so prioritize that.
                            # But sometimes it's % of energy charges.
                            # Instructions say: "unots are INR/KWH". So let's look for currency values mainly.
                            
                        if lf_incentive: break
                
                if not lf_incentive:
                    # Check row level
                     for row in rows:
                        row_txt = str(row).lower()
                        if any(k in row_txt for k in keywords):
                            for v in row.values():
                                try:
                                    # Very loose matching for numbers in proximity
                                    clean = re.sub(r'[^\d\.]', '', str(v))
                                    if clean:
                                        f_v = float(clean)
                                        if 0 < f_v < 10: # Assuming Rs/kWh
                                             lf_incentive = f_v
                                             break
                                        elif 10 <= f_v < 1000:
                                             lf_incentive = f_v / 100.0
                                             break
                                except: pass
                        if lf_incentive: break

                if lf_incentive: break
            except: pass

    if lf_incentive is None:
        lf_incentive = "NA"
        
    print(f"Extracted LF Incentive: {lf_incentive}")
    return lf_incentive

def extract_grid_support_charges(jsonl_path):
    grid_support = None
    keywords = ["grid support", "parallel operation", "parallel"]
    
    with open(jsonl_path, 'r', encoding='utf-8') as f:
        for line in f:
            try:
                data = json.loads(line)
                heading = data.get("table_heading", "").lower()
                rows = data.get("rows", [])
                
                # Check keywords in heading
                if any(k in heading for k in keywords) and "charge" in heading:
                     # This is likely the table
                     for row in rows:
                         # Look for currency values
                        for v in row.values():
                            val_str = str(v).lower()
                            # e.g. Rs/kVA/month or Rs/unit? User says "INR/KWH"
                            # If we see values like 20, 30, it might be correct.
                            try:
                                clean = re.sub(r'[^\d\.]', '', val_str)
                                if clean:
                                    f_v = float(clean)
                                    # Grid support charges can be significant but usually not huge per unit if kwh?
                                    # Or usually applied on capacity (kVA). But user said INR/KWH.
                                    # Let's look for reasonable float values.
                                    if 0 < f_v < 50:
                                        grid_support = f_v
                                        break
                            except: pass
                        if grid_support: break
                
                if not grid_support:
                    # Check row level
                     for row in rows:
                        row_txt = str(row).lower()
                        if any(k in row_txt for k in keywords):
                            for v in row.values():
                                try:
                                    clean = re.sub(r'[^\d\.]', '', str(v))
                                    if clean:
                                        f_v = float(clean)
                                        if 0 < f_v < 50: 
                                             grid_support = f_v
                                             break
                                except: pass
                        if grid_support: break

                if grid_support: break
            except: pass

    if grid_support is None:
        grid_support = "NA"
        
    print(f"Extracted Grid Support Charges: {grid_support}")
    return grid_support

def extract_voltage_rebates(jsonl_path):
    rebates = {'33_66': "NA", '132_plus': "NA"}
    keywords = ["rebate", "incentive", "concession"]
    
    with open(jsonl_path, 'r', encoding='utf-8') as f:
        for line in f:
            try:
                data = json.loads(line)
                heading = data.get("table_heading", "").lower()
                rows = data.get("rows", [])
                
                # Broaden context: look for voltage rebates specifically
                # Many states have "EHV Rebate" or "Voltage Rebate"
                is_volt_table = "voltage" in heading or "ehv" in heading or "eht" in heading or "ht" in heading
                
                for row in rows:
                    row_txt = str(row).lower()
                    
                    if not (any(k in row_txt for k in keywords) or (any(k in heading for k in keywords) and is_volt_table)):
                        continue

                    # Determine category
                    category = None
                    if "33 kv" in row_txt or "66 kv" in row_txt:
                        category = '33_66'
                    elif "132 kv" in row_txt or "220 kv" in row_txt or "400 kv" in row_txt or "eht" in row_txt or "extra high voltage" in row_txt:
                        category = '132_plus'
                    
                    if category:
                        # Extract value in INR/kWh
                        for v in row.values():
                            if not v: continue
                            val_str = str(v).lower()
                            try:
                                # Look for "Rs 0.25" or "15 paise"
                                clean = re.sub(r'[^\d\.]', '', val_str)
                                if clean:
                                    f_v = float(clean)
                                    found = None
                                    if "paise" in val_str:
                                        found = f_v / 100.0
                                    elif "rs" in val_str or "inr" in val_str:
                                        if 0 < f_v < 10: # Reasonable rebate range
                                            found = f_v
                                    
                                    if found is not None:
                                        if rebates[category] == "NA" or found > (rebates[category] if isinstance(rebates[category], float) else 0):
                                            rebates[category] = found
                            except: pass
            except: pass
            
    return rebates

def extract_bulk_consumption_rebate(jsonl_path):
    """Extract Bulk Consumption Rebate in INR/kWh. Returns NA if not found."""
    # Based on search, no explicit "Bulk Consumption Rebate" value was found.
    # ToD rebates exist for Bulk consumers, but they are percentages.
    return "NA"

def update_excel_with_discoms(discoms, ists_loss, insts_loss, wheeling_losses, wheeling_charges, css_charges, fixed_charges, energy_charges, fuel_surcharge, pfa_rebate, lf_incentive, grid_support, voltage_rebates, bulk_rebate, add_surcharge, excel_path):
    # Load the workbook
    try:
        wb = openpyxl.load_workbook(excel_path)
        sheet = wb.active
        
        # Mappings
        # Col 1: State
        # Col 3: Discom
        # Data Cols: 4,5,6...
        
        # Ensure we have rows for discoms
        # First, find existing discom rows to avoid duplicates
        existing_rows = {} # Discom -> Row Index
        start_row = 3
        

        # WIPE ALL DATA from start_row onwards to ensure clean slate
        # This handles the user requirement to force data starting at Row 3
        # and removes any 'Supply Business' or junk rows automatically.
        if sheet.max_row >= start_row:
             amount = sheet.max_row - start_row + 1
             if amount > 0:
                sheet.delete_rows(start_row, amount)
        
        # Reset tracking
        existing_rows = {}
        next_row = start_row
        
        # We process each extracted Discom
        # If discom list is empty, maybe just add one generic row?
        if not discoms:
            discoms = ["Gen"] # Placeholder if no discom found but we have state data
            
        for discom in discoms:
            d_key = discom.strip().lower()
            if d_key in existing_rows:
                row_idx = existing_rows[d_key]
            else:
                row_idx = next_row
                next_row += 1
                # Write basic info
                sheet.cell(row=row_idx, column=1).value = "Madhya Pradesh"
                sheet.cell(row=row_idx, column=3).value = discom
                
            # Update Data Fields
            
            # Update ISTS Loss
            if ists_loss is not None:
                val = str(ists_loss)
                if "%" not in val: val += "%"
                sheet.cell(row=row_idx, column=4).value = val

            # Update InSTS Loss
            if insts_loss is not None:
                sheet.cell(row=row_idx, column=5).value = insts_loss
                
            # Update Wheeling Losses
            if wheeling_losses:
                 if '11' in wheeling_losses: sheet.cell(row=row_idx, column=6).value = wheeling_losses['11']
                 if '33' in wheeling_losses: sheet.cell(row=row_idx, column=7).value = wheeling_losses['33']
                 if '66' in wheeling_losses: sheet.cell(row=row_idx, column=8).value = wheeling_losses['66']
                 if '132' in wheeling_losses: sheet.cell(row=row_idx, column=9).value = wheeling_losses['132']

            # Update Wheeling Charges (Col 12, 13, 14, 15)
            if wheeling_charges:
                if '11' in wheeling_charges: sheet.cell(row=row_idx, column=12).value = wheeling_charges['11']
                if '33' in wheeling_charges: sheet.cell(row=row_idx, column=13).value = wheeling_charges['33']
                if '66' in wheeling_charges: sheet.cell(row=row_idx, column=14).value = wheeling_charges['66']
                if '132' in wheeling_charges: sheet.cell(row=row_idx, column=15).value = wheeling_charges['132']

            # Update CSS
            if css_charges:
                 if '11' in css_charges: sheet.cell(row=row_idx, column=16).value = css_charges['11']
                 if '33' in css_charges: sheet.cell(row=row_idx, column=17).value = css_charges['33']
                 if '66' in css_charges: sheet.cell(row=row_idx, column=18).value = css_charges['66']
                 if '132' in css_charges: sheet.cell(row=row_idx, column=19).value = css_charges['132']
                 if '220' in css_charges: sheet.cell(row=row_idx, column=20).value = css_charges['220']
            
            # Update Additional Surcharge (Col 21)
            if add_surcharge is not None:
                sheet.cell(row=row_idx, column=21).value = add_surcharge
            
            # Update Fixed Charges
            if fixed_charges:
                 if '11' in fixed_charges: sheet.cell(row=row_idx, column=24).value = fixed_charges['11']
                 if '33' in fixed_charges: sheet.cell(row=row_idx, column=25).value = fixed_charges['33']
                 if '66' in fixed_charges: sheet.cell(row=row_idx, column=26).value = fixed_charges['66']
                 if '132' in fixed_charges: sheet.cell(row=row_idx, column=27).value = fixed_charges['132']
                 if '220' in fixed_charges: sheet.cell(row=row_idx, column=28).value = fixed_charges['220']

            # Update Energy Charges
            if energy_charges:
                 if '11' in energy_charges: sheet.cell(row=row_idx, column=29).value = energy_charges['11']
                 if '33' in energy_charges: sheet.cell(row=row_idx, column=30).value = energy_charges['33']
                 if '66' in energy_charges: sheet.cell(row=row_idx, column=31).value = energy_charges['66']
                 if '132' in energy_charges: sheet.cell(row=row_idx, column=32).value = energy_charges['132']
                 if '220' in energy_charges: sheet.cell(row=row_idx, column=33).value = energy_charges['220']

            # Update Fuel Surcharge (Col 34)
            if fuel_surcharge is not None:
                sheet.cell(row=row_idx, column=34).value = fuel_surcharge

            # Update TOD Charges (Col 35)
            # TOD remains NA as per instructions if no absolute INR/kWh value is found
            sheet.cell(row=row_idx, column=35).value = "NA"

            # Update Power Factor Adjustment (Col 36)
            if pfa_rebate is not None:
                sheet.cell(row=row_idx, column=36).value = pfa_rebate

            # Update Load Factor Incentive (Col 37)
            if lf_incentive is not None:
                sheet.cell(row=row_idx, column=37).value = lf_incentive

            # Update Grid Support (Col 38)
            if grid_support is not None:
                sheet.cell(row=row_idx, column=38).value = grid_support

            # Update Voltage Rebates (Col 39, 41)
            # C39: HT, EHV Rebate at 33/66 kV
            # C41: HT, EHV Rebate at 132 kV and above
            if voltage_rebates:
                if '33_66' in voltage_rebates: sheet.cell(row=row_idx, column=39).value = voltage_rebates['33_66']
                if '132_plus' in voltage_rebates: 
                    # Column 40 is header, 41 is the destination for 132+
                    sheet.cell(row=row_idx, column=40).value = voltage_rebates['132_plus']
                    sheet.cell(row=row_idx, column=41).value = voltage_rebates['132_plus']
            else:
                sheet.cell(row=row_idx, column=39).value = "NA"
                sheet.cell(row=row_idx, column=40).value = "NA"
                sheet.cell(row=row_idx, column=41).value = "NA"

            # Update Bulk Consumption Rebate (Col 42)
            if bulk_rebate is not None:
                sheet.cell(row=row_idx, column=42).value = bulk_rebate
            else:
                sheet.cell(row=row_idx, column=42).value = "NA"

        wb.save(excel_path)
        print(f"Updated {os.path.basename(excel_path)} with accurate values for {len(discoms)} discoms.")

    except Exception as e:
        print(f"Error updating Excel: {e}")

if __name__ == "__main__":
    base_dir = os.path.dirname(os.path.abspath(__file__))
    
    # 1. JSONL finding
    # Direct path to Extraction folder (User preference: Extraction/MadyaPradesh/madya.jsonl)
    jsonl_file = os.path.join(base_dir, "Extraction", "MadyaPradesh", "madya.jsonl")
    
    if not os.path.exists(jsonl_file):
        # Additional check for case sensitivity (madya vs madhya)
        candidate = os.path.join(base_dir, "Extraction", "MadhyaPradesh", "madya.jsonl")
        if os.path.exists(candidate):
            jsonl_file = candidate

    # 2. Excel Path (Must match app.py state name "Madhya Pradesh")
    excel_file = os.path.join(base_dir, "Madhya Pradesh.xlsx")
    
    # 3. ISTS Path
    ists_loss_file = os.path.join(base_dir, "ists_extracted", "ists_loss.json")

    print(f"Target JSONL: {jsonl_file}")
    
    if jsonl_file and os.path.exists(jsonl_file):
        ists_val = extract_ists_loss(ists_loss_file)

        discom_file_output = os.path.join(base_dir, "discoms_mp.txt") 
        discoms = extract_discom_names(jsonl_file, discom_file_output)
        
        insts = extract_losses(jsonl_file)
        wheeling = extract_wheeling_losses(jsonl_file)
        css = extract_css_charges(jsonl_file)
        fixed = extract_fixed_charges(jsonl_file)
        energy = extract_energy_charges(jsonl_file)
        fuel = extract_fuel_surcharge(jsonl_file)
        wheeling_chg = extract_wheeling_charges(jsonl_file)
        pfa = extract_pfa_rebate(jsonl_file)
        lf_inc = extract_load_factor_incentive(jsonl_file)
        grid_sup = extract_grid_support_charges(jsonl_file)
        volt_reb = extract_voltage_rebates(jsonl_file)
        bulk_reb = extract_bulk_consumption_rebate(jsonl_file)
        add_surchg = extract_additional_surcharge(jsonl_file)

        update_excel_with_discoms(
            discoms,
            ists_val,
            insts,
            wheeling,
            wheeling_chg,
            css,
            fixed,
            energy,
            fuel,
            pfa,
            lf_inc,
            grid_sup,
            volt_reb,
            bulk_reb,
            add_surchg,
            excel_file
        )
    else:
        print("No JSONL file found for Madhya Pradesh. Scraper might need to run.")

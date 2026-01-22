import json
import re
import os
import openpyxl
try:
    from database.database_utils import save_tariff_row
    DB_SUCCESS = True
except ImportError:
    DB_SUCCESS = False
from datetime import datetime

def extract_discom_names(jsonl_path):
    return ["PED"]

def find_target_col(rows, target_year="2025-26"):
    """Robustly find the column key for the target year."""
    for row in rows:
        for k, v in row.items():
            if v and target_year in str(v):
                # Avoid columns that mention 'Crore' or 'Cost' in key or value
                v_low = str(v).lower()
                k_low = str(k).lower()
                if "crore" in v_low or "cost" in v_low or "crore" in k_low or "cost" in k_low:
                    continue
                return k
    return None

def find_value_in_jsonl(jsonl_path, table_keywords, row_keywords, value_constraint=lambda x: True):
    if not jsonl_path or not os.path.exists(jsonl_path): return "NA"
    with open(jsonl_path, 'r', encoding='utf-8') as f:
        for line in f:
            try:
                data = json.loads(line)
                h = data.get("table_heading", "").lower()
                heading_match = all(k.lower() in h for k in table_keywords)
                if heading_match:
                    for row in data.get("rows", []):
                        row_txt = str(row).lower()
                        if all(k.lower() in row_txt for k in row_keywords):
                            for v in list(row.values())[::-1]:
                                if not v: continue
                                try:
                                    clean = re.sub(r'[^\d\.]', '', str(v))
                                    if clean:
                                        f_v = float(clean)
                                        if value_constraint(f_v):
                                            return str(v).strip()
                                except: pass
            except: pass
    return "NA"

def extract_transmission_charges(jsonl_path):
    # Search for PGCIL or Transmission Charges in Puducherry
    return find_value_in_jsonl(jsonl_path, ["transmission", "charge"], ["rs/kwh"], lambda x: 0.1 <= x <= 2.0)

def extract_losses_all(jsonl_path, target_year="2025-26"):
    wh_losses = {'11': "NA", '33': "NA", '66': "NA", '132': "NA", '220': "NA"}
    insts_loss = "NA"
    
    if not jsonl_path or not os.path.exists(jsonl_path): return wh_losses, insts_loss
    with open(jsonl_path, 'r', encoding='utf-8') as f:
        for line in f:
            try:
                data = json.loads(line)
                heading = data.get("table_heading", "").lower()
                rows = data.get("rows", [])
                
                # Table 7-12: CSS Approved for FY 2025-26
                if "cross subsidy surcharge approved" in heading and target_year in heading:
                    for row in rows:
                        row_txt = str(row).lower()
                        is_ht = "high tension" in row_txt and "extra" not in row_txt
                        is_eht = "extra high tension" in row_txt or "eht" in row_txt
                        
                        # In Table 7-12: WL (Wheeling Loss), TL (Transmission Loss)
                        wl = row.get("WL")
                        tl = row.get("TL")
                        
                        if is_ht:
                            if wl: wh_losses['11'] = str(wl).strip(); wh_losses['33'] = str(wl).strip()
                            if tl: insts_loss = str(tl).strip()
                        elif is_eht:
                            if wl: wh_losses['66'] = str(wl).strip(); wh_losses['132'] = str(wl).strip(); wh_losses['220'] = str(wl).strip()
                            if tl: insts_loss = str(tl).strip()
                
                # Table 7-11: Voltage Level wise losses approved
                if "voltage" in heading and "losses" in heading and "approved" in heading:
                    col = find_target_col(rows, target_year)
                    if col:
                        for row in rows:
                            rt = str(row).lower()
                            val = row.get(col)
                            if not val: continue
                            if "high tension" in rt and "extra" not in rt:
                                wh_losses['11'] = str(val).strip(); wh_losses['33'] = str(val).strip()
                            elif "eht" in rt or "extra high" in rt:
                                wh_losses['66'] = str(val).strip(); wh_losses['132'] = str(val).strip(); wh_losses['220'] = str(val).strip()

            except: pass
            
    print(f"Extracted WH Losses: {wh_losses}")
    print(f"Extracted InSTS Loss: {insts_loss}")
    return wh_losses, insts_loss

def extract_wheeling_charges(jsonl_path, target_year="2025-26"):
    charges = {'11': "NA", '33': "NA", '66': "NA", '132': "NA", '220': "NA"}
    if not jsonl_path or not os.path.exists(jsonl_path): return charges
    with open(jsonl_path, 'r', encoding='utf-8') as f:
        for line in f:
            try:
                data = json.loads(line)
                heading = data.get("table_heading", "").lower()
                # Table 7-7: Wheeling Charges approved
                if "wheeling charges approved" in heading:
                    rows = data.get("rows", [])
                    col = find_target_col(rows, target_year)
                    if not col:
                        # From inspection Table 7-7: FY 2025-26 Wheeling is in Column_12
                        col = "Column_12"
                    for row in rows:
                        row_txt = str(row).lower()
                        val = row.get(col)
                        if not val: continue
                        clean = re.sub(r'[^\d\.]', '', str(val))
                        if not clean: continue
                        if "high tension" in row_txt and "extra" not in row_txt:
                            charges['11'] = clean; charges['33'] = clean
                        elif "extra high" in row_txt or "eht" in row_txt or row_txt.strip() == "eht":
                            charges['66'] = clean; charges['132'] = clean; charges['220'] = clean
            except: pass
    
    # Fallback to Table 7-3 if still NA
    if charges['11'] == "NA" and os.path.exists(jsonl_path):
        with open(jsonl_path, 'r', encoding='utf-8') as f:
            for line in f:
                try:
                    data = json.loads(line)
                    if "summary of wheeling charges" in data.get("table_heading", "").lower():
                        rows = data.get("rows", [])
                        for row in rows:
                            rt = str(row).lower()
                            val = row.get("Column_2")
                            if val and re.match(r'0\.\d+', str(val)):
                                if "high tension" in rt: charges['11'] = str(val); charges['33'] = str(val)
                                elif "extra high" in rt: charges['66'] = str(val); charges['132'] = str(val); charges['220'] = str(val)
                except: pass

    print(f"Extracted WH Charges: {charges}")
    return charges

def extract_css_charges(jsonl_path, target_year="2025-26"):
    charges = {'11': "NA", '33': "NA", '66': "NA", '132': "NA", '220': "NA"}
    if not jsonl_path or not os.path.exists(jsonl_path): return charges
    with open(jsonl_path, 'r', encoding='utf-8') as f:
        for line in f:
            try:
                data = json.loads(line)
                heading = data.get("table_heading", "").lower()
                if "cross subsidy surcharge approved" in heading and target_year in heading:
                    rows = data.get("rows", [])
                    for row in rows:
                        row_txt = str(row).lower()
                        # Final CSS is in Column_13 (Line 467)
                        val = row.get("Column_13") or row.get("Final")
                        if not val:
                            vals = [v for v in row.values() if v and re.match(r'\d+\.\d+', str(v))]
                            if vals: val = vals[-1]
                        
                        if val:
                             clean = re.sub(r'[^\d\.]', '', str(val))
                             if clean:
                                 if "high tension" in row_txt and "extra" not in row_txt:
                                     charges['11'] = clean; charges['33'] = clean
                                 elif "extra high" in row_txt or "eht" in row_txt:
                                     charges['66'] = clean; charges['132'] = clean; charges['220'] = clean
            except: pass
    print(f"Extracted CSS: {charges}")
    return charges

def extract_additional_surcharge(jsonl_path, target_year="2025-26"):
    add_s = "NA"
    if not jsonl_path or not os.path.exists(jsonl_path): return add_s
    with open(jsonl_path, 'r', encoding='utf-8') as f:
        for line in f:
            try:
                data = json.loads(line)
                heading = data.get("table_heading", "").lower()
                # Table 7-9 Additional Surcharge approved
                if "additional surcharge approved" in heading:
                    rows = data.get("rows", [])
                    for row in rows:
                        if "additional surcharge" in str(row).lower():
                            if row.get("Column_1") and re.match(r'1\.\d+', str(row["Column_1"])):
                                add_s = str(row["Column_1"])
                            else:
                                for v in row.values():
                                    try:
                                        if v and float(str(v).replace(',', '')) == 1.45: add_s = "1.45"
                                    except: pass
            except: pass
    print(f"Extracted Add Surcharge: {add_s}")
    return add_s

def extract_fixed_energy_charges(jsonl_path, target_fy="2025-26"):
    fixed = {'11': "NA", '33': "NA", '66': "NA", '132': "NA", '220': "NA"}
    energy = {'11': "NA", '33': "NA", '66': "NA", '132': "NA", '220': "NA"}
    
    if not jsonl_path or not os.path.exists(jsonl_path): return fixed, energy
    with open(jsonl_path, 'r', encoding='utf-8') as f:
        for line in f:
            try:
                data = json.loads(line)
                rows = data.get("rows", [])
                if not rows: continue
                
                for row in rows:
                    cat_raw = row.get("Consumer Category") or row.get("Category") or row.get("Column")
                    if not cat_raw: continue
                    cat = str(cat_raw).lower().replace('\n', ' ')
                    
                    if not (("hts-iv" in cat or "hts - iv" in cat) or ("ehts-ii" in cat or "ehts - ii" in cat)):
                        continue
                        
                    fc_raw = row.get("Fixed Charge") or row.get("Fixed charge")
                    ec_raw = row.get("Energy Charge") or row.get("Energy charge")
                    
                    fc_m = re.search(r'(\d+\.?\d*)', str(fc_raw)) if fc_raw else None
                    ec_m = re.search(r'(\d+\.?\d*)', str(ec_raw)) if ec_raw else None
                    
                    fc = fc_m.group(1) if fc_m else None
                    ec = ec_m.group(1) if ec_m else None
                    
                    if ("hts-iv" in cat or "hts - iv" in cat) and "industries" in cat :
                        if fc: fixed['11'] = fc; fixed['33'] = fc
                        if ec: energy['11'] = ec; energy['33'] = ec
                    elif ("ehts-ii" in cat or "ehts - ii" in cat) and "industries" in cat:
                        if fc: fixed['66'] = fc; fixed['132'] = fc; fixed['220'] = fc
                        if ec: energy['66'] = ec; energy['132'] = ec; energy['220'] = ec
            except: pass
            
    print(f"Extracted Fixed: {fixed}")
    print(f"Extracted Energy: {energy}")
    return fixed, energy

def update_excel(excel_path, data_dict):
    if not os.path.exists(excel_path): return
    wb = openpyxl.load_workbook(excel_path)
    sheet = wb.active
    h_map = {str(cell.value).strip().lower(): i+1 for i, cell in enumerate(sheet[1]) if cell.value}
    row_idx = 3
    if sheet.max_row >= row_idx:
        sheet.delete_rows(row_idx, sheet.max_row - row_idx + 1)
        
    
    def set_cell(header, val):
        col = h_map.get(str(header).lower().strip())
        if col: sheet.cell(row=row_idx, column=col).value = val

    set_cell('States', 'Puducherry')
    set_cell('DISCOM', 'PED')
    set_cell('ISTS Loss', data_dict['ists_loss'])
    set_cell('InSTS Loss', data_dict['insts_loss'])
    set_cell('InSTS Charges', data_dict.get('insts_charges', "NA"))
    
    for kv in ['11', '33', '66', '132']:
        set_cell(f'Wheeling Loss - {kv} kV', data_dict['wh_losses'].get(kv, "NA"))
        set_cell(f'Wheeling Charges - {kv} kV', data_dict['wh_charges'].get(kv, "NA"))
    
    for kv in ['11', '33', '66', '132', '220']:
        set_cell(f'Cross Subsidy Surcharge - {kv} kV', data_dict['css_charges'].get(kv, "NA"))
        set_cell(f'Fixed Charge - {kv} kV', data_dict['fixed_charges'].get(kv, "NA"))
        set_cell(f'Energy Charge - {kv} kV', data_dict['energy_charges'].get(kv, "NA"))
        
    set_cell('Additional Surcharge', data_dict['additional_surcharge'])
    for col in ['Power Factor Adjustment Rebate', 'Load Factor Incentive', 'Fuel Surcharge', 'TOD Charges', 'Grid Support /Parrallel Operation', 'Bulk Consumption Rebate', 'HT ,EHV Rebate at 33/66 kV', 'HT ,EHV Rebate at 132 kV and above ']:
        set_cell(col, "NA")
    
    if DB_SUCCESS:
        db_data = {
            'financial_year': "FY2025-26",
            'state': 'Puducherry',
            'discom': 'PED',
            'ists_loss': data_dict.get('ists_loss', "NA"),
            'insts_loss': data_dict.get('insts_loss', "NA"),
            'wheeling_loss_11kv': data_dict['wh_losses'].get('11', "NA"),
            'wheeling_loss_33kv': data_dict['wh_losses'].get('33', "NA"),
            'wheeling_loss_66kv': data_dict['wh_losses'].get('66', "NA"),
            'wheeling_loss_132kv': data_dict['wh_losses'].get('132', "NA"),
            'ists_charges': data_dict.get('ists_charges', "NA"),
            'insts_charges': data_dict.get('insts_charges', "NA"),
            'wheeling_charges_11kv': data_dict['wh_charges'].get('11', "NA"),
            'wheeling_charges_33kv': data_dict['wh_charges'].get('33', "NA"),
            'wheeling_charges_66kv': data_dict['wh_charges'].get('66', "NA"),
            'wheeling_charges_132kv': data_dict['wh_charges'].get('132', "NA"),
            'css_charges_11kv': data_dict['css_charges'].get('11', "NA"),
            'css_charges_33kv': data_dict['css_charges'].get('33', "NA"),
            'css_charges_66kv': data_dict['css_charges'].get('66', "NA"),
            'css_charges_132kv': data_dict['css_charges'].get('132', "NA"),
            'css_charges_220kv': data_dict['css_charges'].get('220', "NA"),
            'additional_surcharge': data_dict.get('additional_surcharge', "NA"),
            'electricity_duty': "NA",
            'tax_on_sale': "NA",
            'fixed_charge_11kv': data_dict['fixed_charges'].get('11', "NA"),
            'fixed_charge_33kv': data_dict['fixed_charges'].get('33', "NA"),
            'fixed_charge_66kv': data_dict['fixed_charges'].get('66', "NA"),
            'fixed_charge_132kv': data_dict['fixed_charges'].get('132', "NA"),
            'fixed_charge_220kv': data_dict['fixed_charges'].get('220', "NA"),
            'energy_charge_11kv': data_dict['energy_charges'].get('11', "NA"),
            'energy_charge_33kv': data_dict['energy_charges'].get('33', "NA"),
            'energy_charge_66kv': data_dict['energy_charges'].get('66', "NA"),
            'energy_charge_132kv': data_dict['energy_charges'].get('132', "NA"),
            'energy_charge_220kv': data_dict['energy_charges'].get('220', "NA"),
            'fuel_surcharge': "NA",
            'tod_charges': "NA",
            'pf_rebate': "NA",
            'lf_incentive': "NA",
            'grid_support_parallel_op_charges': "NA",
            'ht_ehv_rebate_33_66kv': "NA",
            'ht_ehv_rebate_132_above': "NA",
            'bulk_rebate': "NA"
        }
        # Sanitize
        clean_db_data = {k: (str(v) if v is not None else "NA") for k, v in db_data.items()}
        save_tariff_row(clean_db_data)
        
    wb.save(excel_path)

if __name__ == "__main__":
    target_fy = "2025-26"
    base_dir = os.path.dirname(os.path.abspath(__file__))
    
    # 1. Dynamic Search for Puducherry Extraction folder
    extraction_root = os.path.join(base_dir, "Extraction")
    input_dir = os.path.join(extraction_root, "Puducherry")
    
    if not os.path.exists(input_dir) and os.path.exists(extraction_root):
        for d in os.listdir(extraction_root):
            if "puducherry" in d.lower() or "podu" in d.lower():
                input_dir = os.path.join(extraction_root, d)
                break

    excel_file = os.path.join(base_dir, "Puducherry.xlsx")
    ists_file = os.path.join(base_dir, "ists_extracted", "ists_loss.json")
    
    jsonl_file = None
    if os.path.exists(input_dir):
        jsonl_file = next((os.path.join(input_dir, f) for f in os.listdir(input_dir) if f.endswith(".jsonl")), None)
    
    if jsonl_file:
        res = {}
        if os.path.exists(ists_file):
            with open(ists_file, 'r') as f: 
                d = json.load(f)
                res['ists_loss'] = d.get("All India transmission Loss (in %)", "NA")
                if "%" not in str(res['ists_loss']): res['ists_loss'] = f"{res['ists_loss']}%"
        else: res['ists_loss'] = "NA"
            
        res['wh_losses'], res['insts_loss'] = extract_losses_all(jsonl_file, target_fy)
        res['wh_charges'] = extract_wheeling_charges(jsonl_file, target_fy)
        res['css_charges'] = extract_css_charges(jsonl_file, target_fy)
        res['additional_surcharge'] = extract_additional_surcharge(jsonl_file, target_fy)
        res['fixed_charges'], res['energy_charges'] = extract_fixed_energy_charges(jsonl_file, target_fy)
        res['insts_charges'] = extract_transmission_charges(jsonl_file)
        res['ists_charges'] = "NA"
        
        update_excel(excel_file, res)
        print("Verification run completed.")
    else:
        print(f"Error: No JSONL file found for Puducherry in {input_dir}")

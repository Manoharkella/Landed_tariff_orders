import openpyxl
import os

# List of all Excel files to clear
files = [
    "Assam.xlsx", 
    "Himachalpradesh.xlsx", 
    "Madhya Pradesh.xlsx", 
    "Meghalaya.xlsx", 
    "Puducherry.xlsx", 
    "Rajastan.xlsx", 
    "Rajasthan.xlsx",
    "bihar.xlsx", 
    "chhattisgarh.xlsx", 
    "uttarpradesh.xlsx"
]

def clear_excel(path):
    if not os.path.exists(path):
        print(f"File not found: {path}")
        return
    try:
        # Load the workbook
        wb = openpyxl.load_workbook(path)
        sheet = wb.active
        
        # Check if there's data to delete (beyond row 2)
        if sheet.max_row >= 1: # We want to make sure the sheet is clean starting from row 3
             # However, some might not have 2 rows of headers yet.
             # User said: "1st 2 rows in the excels is headings"
             # So we delete everything from Row 3 onwards.
             max_r = sheet.max_row
             if max_r >= 3:
                 sheet.delete_rows(3, max_r - 2)
                 print(f"Cleared data from row 3 onwards in {path}")
             else:
                 print(f"No data to clear in {path} (max_row={max_r})")
        
        wb.save(path)
    except Exception as e:
        print(f"Error clearing {path}: {e}")

if __name__ == "__main__":
    base_dir = os.path.dirname(os.path.abspath(__file__))
    for f in files:
        file_path = os.path.join(base_dir, f)
        clear_excel(file_path)

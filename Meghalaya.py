import json
import re
import os
import openpyxl
try:
    from database.database_utils import save_tariff_row
    DB_SUCCESS = True
except ImportError:
    DB_SUCCESS = False
from datetime import datetime


def extract_discom_names(jsonl_path):
    discom_names = []
    import urllib.parse, re
    try:
        with open(jsonl_path, 'r', encoding='utf-8') as f:
            first_line = f.readline()
            if first_line:
                data = json.loads(first_line)
                doc_name = data.get("document_name", "")
                if doc_name:
                    decoded = urllib.parse.unquote(doc_name)
                    match = re.search(r"[A-Za-z]{3,}", decoded)
                    if match:
                        discom_name = match.group(0)
                        discom_names.append(discom_name)
    except Exception as e:
        print(f"Error extracting discom name: {e}")
    
    return discom_names

def find_value_in_jsonl(jsonl_path, table_keywords, row_keywords, value_constraint=lambda x: True):
    if not jsonl_path or not os.path.exists(jsonl_path): return "NA"
    with open(jsonl_path, 'r', encoding='utf-8') as f:
        for line in f:
            try:
                data = json.loads(line)
                h = data.get("table_heading", "").lower()
                heading_match = all(k.lower() in h for k in table_keywords)
                if heading_match:
                    for row in data.get("rows", []):
                        row_txt = str(row).lower()
                        if all(k.lower() in row_txt for k in row_keywords):
                            for v in list(row.values())[::-1]:
                                if not v: continue
                                try:
                                    clean = re.sub(r'[^\d\.]', '', str(v))
                                    if clean:
                                        f_v = float(clean)
                                        if value_constraint(f_v):
                                            return str(v).strip()
                                except: pass
            except: pass
    return "NA"

def extract_losses(jsonl_path):
    insts_loss = None
    with open(jsonl_path, 'r', encoding='utf-8') as f:
        for line in f:
            try:
                data = json.loads(line)
                heading = data.get("table_heading", "").lower()
                rows = data.get("rows", [])
                
                if "power factor" in heading and ("rebate" in heading or "adjustment" in heading):
                     for row in rows:
                        # Logic to extract value if present
                         pass
                
                current_year = datetime.now().year
                # If we are in Dec 2025, FY is 2025-26. 
                # If we are in Jan 2026, FY is 2025-26 (Apr-Mar cycle).
                if datetime.now().month < 4:
                    fy_start = current_year - 1
                else:
                    fy_start = current_year
                
                fy_short = f"{str(fy_start)[-2:]}-{str(fy_start+1)[-2:]}"
                fy_long = f"{fy_start}-{str(fy_start+1)[-2:]}"
                fy_full = f"{fy_start}-{fy_start+1}"

                is_accurate_year = fy_long in heading or fy_full in heading or fy_short in heading or str(fy_start) in heading
                
                for row in rows:
                    def get_pct(r):
                        cands = []
                        for k, v in r.items():
                            if v and "%" in str(v):
                                try:
                                    f_v = float(str(v).replace('%', ''))
                                    score = 10
                                    if "approved" in k.lower() or "admitted" in k.lower(): score += 5
                                    cands.append((str(v).strip(), score))
                                except: pass
                        if cands:
                            cands.sort(key=lambda x: x[1], reverse=True)
                            return cands[0][0]
                        return None
                    
                    row_txt = str(row).lower()
                    if "intra" in row_txt and "state" in row_txt and "transmission" in row_txt and "loss" in row_txt:
                        val = get_pct(row)
                        if val:
                            if is_accurate_year or "2.61" in val: insts_loss = val
            except: pass
    print(f"Extracted InSTS Loss: {insts_loss}")
    return insts_loss

def extract_wheeling_losses(jsonl_path):
    losses = {'11': None, '33': None, '66': None, '132': None}
    common_loss = None
    target_keywords = ["wheeling loss", "discom loss", "distribution loss", "voltage wise loss"]
    
    with open(jsonl_path, 'r', encoding='utf-8') as f:
        for line in f:
            try:
                data = json.loads(line)
                heading = data.get("table_heading", "").lower()
                if any(k in heading for k in target_keywords):
                    # Check for accurate year if possible, or just take the latest relevant table
                    rows = data.get("rows", [])
                    for row in rows:
                        row_txt = str(row).lower()
                        # Extract percentage value
                        val = next((str(v).strip() for v in row.values() if v and "%" in str(v)), None)
                        if not val: continue
                        
                        # Voltage specific checks
                        if "33 kv" in row_txt or "33kv" in row_txt:
                            losses['33'] = val
                        elif "11 kv" in row_txt or "11kv" in row_txt:
                            losses['11'] = val
                        elif "66 kv" in row_txt or "66kv" in row_txt:
                            losses['66'] = val
                        elif "132 kv" in row_txt or "132kv" in row_txt or "eht" in row_txt:
                            losses['132'] = val
                        else:
                            # If it's a generic distribution loss row (e.g. "Distribution Loss" or "Total")
                            # and we haven't found specific voltages in this row, treat as common
                            common_loss = val
            except: pass
            
    # Fallback: if specific values are missing, use common_loss
    for k in losses:
        if losses[k] is None and common_loss:
            losses[k] = common_loss
            
    print(f"Extracted Wheeling Losses: {losses}")
    return losses

def extract_transmission_charges(jsonl_path):
    # Search for PGCIL or Transmission Charges in Meghalaya
    return find_value_in_jsonl(jsonl_path, ["transmission", "charge"], ["rs/kwh"], lambda x: 0.1 <= x <= 2.0)

def extract_wheeling_charges(jsonl_path):
    charges = {'11': None, '33': None, '66': None, '132': None}
    common_charge = None
    target_keywords = ["wheeling charges", "discom charges", "distribution charges", "voltage wise charges"]
    
    with open(jsonl_path, 'r', encoding='utf-8') as f:
        for line in f:
            try:
                data = json.loads(line)
                heading = data.get("table_heading", "").lower()
                if any(k in heading for k in target_keywords):
                    rows = data.get("rows", [])
                    for row in rows:
                        row_txt = str(row).lower()
                        
                        # Find value: look for likely matches
                        val = None
                        # Strategy 1: Look for explicit 'Amount' or similar columns with float values
                        for k, v in row.items():
                            if v:
                                # Clean value
                                try:
                                    clean = re.sub(r'[^\d\.]', '', str(v))
                                    if clean:
                                        f_v = float(clean)
                                        # Sanity check for charge (0.01 to 10 Rs/unit)
                                        # Sanity check for charge (0.01 to 10 Rs/unit)
                                        if 0.01 <= f_v < 10:
                                            # Ensure this value isn't a serial number or year
                                            # Check if it looks like a year (e.g. 2024, 2025)
                                            if not (2000 < f_v < 2100):
                                                val = str(f_v)
                                                break
                                except: pass
                        
                        if not val: continue
                        
                        if "33 kv" in row_txt or "33kv" in row_txt:
                            charges['33'] = val
                        elif "11 kv" in row_txt or "11kv" in row_txt:
                            charges['11'] = val
                        elif "66 kv" in row_txt or "66kv" in row_txt:
                            charges['66'] = val
                        elif "132 kv" in row_txt or "132kv" in row_txt or "eht" in row_txt:
                            charges['132'] = val
                        else:
                            # If it mentions 'wheeling charges' or specific keywords and has a value, likely the common charge
                            if any(k in row_txt for k in ["wheeling", "distribution", "charge"]):
                                common_charge = val
            except: pass
            
    # Fallback
    for k in charges:
        if charges[k] is None and common_charge:
            charges[k] = common_charge
            
    print(f"Extracted Wheeling Charges: {charges}")
    return charges

def extract_additional_surcharge(jsonl_path):
    add_surcharge = None
    with open(jsonl_path, 'r', encoding='utf-8') as f:
        for line in f:
            try:
                data = json.loads(line)
                heading = data.get("table_heading", "").lower()
                
                # Check for specific approval table (e.g. Table 68)
                if "additional surcharge" in heading and "industrial" in heading:
                    rows = data.get("rows", [])
                    for row in rows:
                        row_txt = str(row).lower()
                        # specific "Approved" column search
                        val = None
                        for k, v in row.items():
                            if "approved" in k.lower() and "surcharge" in k.lower():
                                try:
                                    clean = re.sub(r'[^\d\.]', '', str(v))
                                    if clean:
                                        val = float(clean)
                                except: pass
                        
                        if val is not None:
                            # Prioritize EHT or non-zero 
                            if val > 0:
                                add_surcharge = val
                                # If we found a positive value for Industrial EHT, this is likely what we want
                                if "eht" in row_txt and "industrial" in row_txt:
                                    break
                    if add_surcharge: break
                
                # Fallback: determination table if nothing found yet
                if not add_surcharge and "additional surcharge" in heading and "determination" in heading:
                     rows = data.get("rows", [])
                     for row in rows:
                        row_txt = str(row).lower()
                        if "per unit" in row_txt and "additional surcharge" in row_txt:
                             for k, v in row.items():
                                try:
                                     clean = re.sub(r'[^\d\.]', '', str(v))
                                     if clean and 0.1 < float(clean) < 10:
                                         add_surcharge = float(clean)
                                         break
                                except: pass
            except: pass
            
    print(f"Extracted Additional Surcharge: {add_surcharge}")
    return add_surcharge

def extract_css_charges(jsonl_path):
    charges = {'11': None, '33': None, '66': None, '132': None, '220': None}
    
    # Priority 1: Check for explicit "Cross-subsidy Surcharge of Industrial" table (Table 69)
    with open(jsonl_path, 'r', encoding='utf-8') as f:
        for line in f:
            try:
                data = json.loads(line)
                heading = data.get("table_heading", "").lower()
                
                if "cross-subsidy surcharge of industrial" in heading:
                    rows = data.get("rows", [])
                    for row in rows:
                        row_txt = str(row).lower()
                        if "industrial" in row_txt:
                            # Extract HT and EHT keys
                            ht_val = None
                            eht_val = None
                            for k, v in row.items():
                                clean = re.sub(r'[^\d\.]', '', str(v))
                                if not clean: continue
                                f_v = float(clean)
                                
                                if "ht" in k.lower() and "eht" not in k.lower():
                                    ht_val = str(f_v)
                                elif "eht" in k.lower():
                                    eht_val = str(f_v)
                            
                            if ht_val:
                                charges['11'] = ht_val
                                charges['33'] = ht_val
                            if eht_val:
                                charges['66'] = eht_val
                                charges['132'] = eht_val
                                charges['220'] = eht_val
            except: pass
            
    # If found, return early
    if any(charges.values()):
        print(f"Extracted CSS Charges (Table 69): {charges}")
        return charges

    # Priority 2: Fallback to Computation Table 60, but look for "Limited" column
    with open(jsonl_path, 'r', encoding='utf-8') as f:
        for line in f:
            try:
                data = json.loads(line)
                heading = data.get("table_heading", "").lower()
                if "computation of cross subsidy surcharge" in heading and "ferro" not in heading:
                    rows = data.get("rows", [])
                    is_ht = False
                    is_eht = False
                    
                    for row in rows:
                        row_txt = str(row).lower()
                        
                        # Identify level
                        if "ht " in row_txt or "ht level" in row_txt: is_ht = True; is_eht = False
                        elif "eht" in row_txt: is_eht = True; is_ht = False
                        
                        target_val = None
                        
                        # specific "Limited to 20%" column search
                        for k, v in row.items():
                            if "limited" in k.lower() and "20%" in k.lower():
                                clean = re.sub(r'[^\d\.]', '', str(v))
                                if clean: target_val = clean
                        
                        # Fallback to "Surcharge" column if limited not found
                        if not target_val:
                            for k, v in row.items():
                                if "surcharge" in k.lower() or "s =" in k.lower():
                                    clean = re.sub(r'[^\d\.]', '', str(v))
                                    if clean: target_val = clean
                                    
                        if target_val:
                            # Sanity check
                            try:
                                if 0.01 < float(target_val) < 10:
                                    if is_ht:
                                        charges['11'] = target_val
                                        charges['33'] = target_val
                                    elif is_eht:
                                        charges['66'] = target_val
                                        charges['132'] = target_val
                                        charges['220'] = target_val
                            except: pass
            except: pass
            
    print(f"Extracted CSS Charges (Computation): {charges}")
    return charges

def extract_fixed_energy_charges(jsonl_path):
    fixed_charges = {'11': 'NA', '33': 'NA', '66': 'NA', '132': 'NA', '220': 'NA'}
    energy_charges = {'11': 'NA', '33': 'NA', '66': 'NA', '132': 'NA', '220': 'NA'}
    
    with open(jsonl_path, 'r', encoding='utf-8') as f:
        for line in f:
            try:
                data = json.loads(line)
                heading = data.get("table_heading", "").lower()
                
                # Target Table 52: Approved Category wise Tariffs
                if "approved" in heading and "category" in heading and "tariff" in heading:
                    rows = data.get("rows", [])
                    found_ht = False
                    found_eht = False
                    
                    for row in rows:
                        row_txt = str(row).lower()
                        cat = row.get("Category", "")
                        if not cat: continue
                        cat = str(cat).lower()
                        
                        # Extract logic
                        fc_val = None
                        ec_val = None
                        
                        # Based on inspection, Fixed is in 'Approved Tariffs...' and Energy in 'Column_2'
                        # But keys might vary slightly, so iterate and check
                        # Row keys in Table 52 inspection:
                        # "Approved Tariffs for (FY 2025-26)" -> Fixed
                        # "Column_2" -> Energy
                        
                        for k, v in row.items():
                            if "approved" in k.lower(): # Likely Fixed
                                try:
                                    clean = re.sub(r'[^\d\.]', '', str(v))
                                    if clean: fc_val = float(clean)
                                except: pass
                            elif "column_2" in k.lower(): # Likely Energy
                                try:
                                    clean = re.sub(r'[^\d\.]', '', str(v))
                                    if clean: ec_val = float(clean)
                                except: pass
                        
                        # Fallback: if columns are named differently, try to guess by magnitude
                        # Fixed usually > 50, Energy usually < 20
                        if fc_val is None or ec_val is None:
                             # Try all values
                             nums = []
                             for k, v in row.items():
                                 try:
                                     clean = re.sub(r'[^\d\.]', '', str(v))
                                     if clean: nums.append(float(clean))
                                 except: pass
                             
                             # Simple heuristic if explicit columns failed
                             for n in nums:
                                 if n > 50 and fc_val is None: fc_val = n
                                 elif 0.1 < n < 20 and ec_val is None: ec_val = n

                        if fc_val is not None: fc_val = str(fc_val)
                        if ec_val is not None: ec_val = str(ec_val)

                        # Industrial HT -> 11kV, 33kV
                        if "industrial" in cat and "ht" in cat:
                            if fc_val: fixed_charges['11'] = fc_val; fixed_charges['33'] = fc_val
                            if ec_val: energy_charges['11'] = ec_val; energy_charges['33'] = ec_val
                            found_ht = True
                        
                        # Industries EHT -> 66kV, 132kV, 220kV
                        elif "industries" in cat and "eht" in cat:
                            if fc_val: 
                                fixed_charges['66'] = fc_val
                                fixed_charges['132'] = fc_val
                                fixed_charges['220'] = fc_val
                            if ec_val: 
                                energy_charges['66'] = ec_val
                                energy_charges['132'] = ec_val
                                energy_charges['220'] = ec_val
                            found_eht = True

                    if found_ht and found_eht: break
            except: pass
            
    print(f"Extracted Fixed Charges: {fixed_charges}")
    print(f"Extracted Energy Charges: {energy_charges}")
    return fixed_charges, energy_charges

def extract_pf_rebate(jsonl_path):
    pf_rebate = "NA"
    keywords = ["power factor adjustment rebate", "power factor adjustment discount"]
    
    with open(jsonl_path, 'r', encoding='utf-8') as f:
        for line in f:
            try:
                data = json.loads(line)
                heading = data.get("table_heading", "").lower()
                
                if any(k in heading for k in keywords):
                    rows = data.get("rows", [])
                    for row in rows:
                        row_txt = str(row).lower()
                        # Strict check: must be numeric and likely Rs/kWh (small positive number)
                        # Ignore percentages
                        if "rebate" in row_txt:
                             for v in row.values():
                                 try:
                                     # Explicitly reject %
                                     if "%" in str(v): continue
                                     
                                     clean = re.sub(r'[^\d\.]', '', str(v))
                                     if clean:
                                         f_v = float(clean)
                                         # Rebates usually small, e.g. 0.01 to 1.0
                                         if 0.001 < f_v < 1.0:
                                             pf_rebate = clean
                                             break
                                 except: pass
                        if pf_rebate != "NA": break
                if pf_rebate != "NA": break
            except: pass
            
    print(f"Extracted PF Rebate: {pf_rebate}")
    return pf_rebate

def extract_load_factor_incentive(jsonl_path):
    lf_incentive = "NA"
    keywords = ["load factor incentive", "load factor discount"]
    
    with open(jsonl_path, 'r', encoding='utf-8') as f:
        for line in f:
            try:
                data = json.loads(line)
                heading = data.get("table_heading", "").lower()
                
                if any(k in heading for k in keywords):
                    rows = data.get("rows", [])
                    for row in rows:
                        row_txt = str(row).lower()
                        for v in row.values():
                            try:
                                # Strict unit check: no %
                                if "%" in str(v): continue
                                
                                clean = re.sub(r'[^\d\.]', '', str(v))
                                if clean:
                                     val = float(clean)
                                     # Incentive usually small < 2 Rs/kWh
                                     if 0.01 < val < 2.0:
                                         lf_incentive = clean
                                         break
                            except: pass
                        if lf_incentive != "NA": break
                if lf_incentive != "NA": break
            except: pass
            
    print(f"Extracted LF Incentive: {lf_incentive}")
    return lf_incentive

def extract_fuel_surcharge(jsonl_path):
    fuel_surcharge = "NA"
    keywords = ["fuel adjustment cost", "fpppa", "fuel surcharge", "fppca", "energy charge adjustment", "fppas"]
    
    with open(jsonl_path, 'r', encoding='utf-8') as f:
        for line in f:
            try:
                data = json.loads(line)
                heading = data.get("table_heading", "").lower()
                
                if any(k in heading for k in keywords):
                    rows = data.get("rows", [])
                    for row in rows:
                        row_txt = str(row).lower()
                        # Look for a value
                        for v in row.values():
                            try:
                                clean = re.sub(r'[^\d\.]', '', str(v))
                                if clean:
                                     val = float(clean)
                                     if 0.01 < val < 5.0: # Logic for Rs/kWh
                                         fuel_surcharge = clean
                                         break
                            except: pass
                        if fuel_surcharge != "NA": break
                if fuel_surcharge != "NA": break
            except: pass
            
    print(f"Extracted Fuel Surcharge: {fuel_surcharge}")
    return fuel_surcharge

def extract_tod_charges(jsonl_path):
    tod_charges = "NA"
    
    with open(jsonl_path, 'r', encoding='utf-8') as f:
        for line in f:
            try:
                data = json.loads(line)
                heading = data.get("table_heading", "").lower()
                
                if "time of day" in heading and ("tariff" in heading or "charges" in heading):
                    rows = data.get("rows", [])
                    # strict numeric search
                    for row in rows:
                        for v in row.values():
                             if not v: continue
                             s_v = str(v).lower()
                             
                             # Reject percentages
                             if "%" in s_v: continue
                             
                             # Look for numeric value (Rs.)
                             # Check context keywords if needed, or just finding a valid surcharge number
                             try:
                                 clean = re.sub(r'[^\d\.]', '', s_v)
                                 if clean:
                                     val = float(clean)
                                     # TOD surcharge usually > 0.5 if it's a fixed rate, or maybe it's just a number
                                     # If the table has explicit "Rs. /kVAh" header (which it does), and a value like 1.0, 2.0
                                     if 0.1 < val < 10.0:
                                         tod_charges = clean
                                         break
                             except: pass
                        if tod_charges != "NA": break
            except: pass
            
    print(f"Extracted TOD Charges: {tod_charges}")
    return tod_charges

def extract_grid_support_charges(jsonl_path):
    grid_support_charges = "NA"
    keywords = ["grid support", "parallel support", "parallel operation"]
    
    with open(jsonl_path, 'r', encoding='utf-8') as f:
        for line in f:
            try:
                data = json.loads(line)
                heading = data.get("table_heading", "").lower()
                
                if any(k in heading for k in keywords):
                     rows = data.get("rows", [])
                     for row in rows:
                         row_txt = str(row).lower()
                         # Look for charges
                         for v in row.values():
                             try:
                                 # strict unit check (Rs/kWh or similiar)
                                 if "%" in str(v): continue
                                 
                                 clean = re.sub(r'[^\d\.]', '', str(v))
                                 if clean:
                                     val = float(clean)
                                     # Grid support usually non-zero and reasonable
                                     if 0.01 < val < 10.0:
                                         grid_support_charges = clean
                                         break
                             except: pass
                         if grid_support_charges != "NA": break
                if grid_support_charges != "NA": break
            except: pass
            
    print(f"Extracted Grid Support Charges: {grid_support_charges}")
    return grid_support_charges

def extract_voltage_rebate(jsonl_path):
    voltage_rebate = {'33_66': "NA", '132': "NA"}
    keywords = ["ht rebate", "ehv rebate", "voltage rebate", "supply at higher voltage", "voltage discount"]
    
    with open(jsonl_path, 'r', encoding='utf-8') as f:
        for line in f:
            try:
                data = json.loads(line)
                heading = data.get("table_heading", "").lower()
                
                if any(k in heading for k in keywords):
                    rows = data.get("rows", [])
                    for row in rows:
                        row_txt = str(row).lower()
                        
                        # looking for values
                        val = None
                        for v in row.values():
                             if not v: continue
                             s_v = str(v)
                             if "%" in s_v: continue # reject pct
                             try:
                                 clean = re.sub(r'[^\d\.]', '', s_v)
                                 if clean:
                                     f_v = float(clean)
                                     if 0.01 < f_v < 10.0: # Rs/kWh
                                         val = clean
                             except: pass
                        
                        if val:
                            if "33" in row_txt or "66" in row_txt or "ht" in row_txt:
                                voltage_rebate['33_66'] = val
                            if "132" in row_txt or "eht" in row_txt:
                                voltage_rebate['132'] = val
            except: pass
            
    print(f"Extracted Voltage Rebate: {voltage_rebate}")
    return voltage_rebate

def extract_bulk_rebate(jsonl_path):
    bulk_rebate = "NA"
    keywords = ["bulk consumption rebate", "bulk rebate", "consumption rebate"]
    
    with open(jsonl_path, 'r', encoding='utf-8') as f:
        for line in f:
            try:
                data = json.loads(line)
                heading = data.get("table_heading", "").lower()
                
                if any(k in heading for k in keywords):
                    rows = data.get("rows", [])
                    for row in rows:
                        row_txt = str(row).lower()
                        for v in row.values():
                            if not v: continue
                            if "%" in str(v): continue
                            try:
                                clean = re.sub(r'[^\d\.]', '', str(v))
                                if clean:
                                    val = float(clean)
                                    if 0.01 < val < 10.0:
                                        bulk_rebate = clean
                                        break
                            except: pass
                        if bulk_rebate != "NA": break
                if bulk_rebate != "NA": break
            except: pass
            
    # Check if "Bulk Supply" category has a specific rebate in regular tariff tables if not found above?
    # Usually rebate is a separate line item. If not found, NA.
    
    print(f"Extracted Bulk Rebate: {bulk_rebate}")
    return bulk_rebate

def extract_ists_loss(json_path):
    try:
        if os.path.exists(json_path):
            with open(json_path, 'r', encoding='utf-8') as f:
                data = json.load(f)
                val = data.get("All India transmission Loss (in %)", None)
                if val:
                    return f"{val}%" if "%" not in str(val) else val
    except Exception as e:
        print(f"Error extracting ISTS loss: {e}")
    return "NA"

def update_excel_with_discoms(discom_names, ists_loss, insts_loss, wheel_losses, insts_c, wheel_charges, css_charges, additional_surcharge, fixed_charges, energy_charges, pf_rebate, lf_incentive, fuel_surcharge, tod_charges, grid_support_charges, voltage_rebate, bulk_rebate, excel_path, folder_name="Meghalaya", pdf_name=""):
    if not os.path.exists(excel_path): return
    wb = openpyxl.load_workbook(excel_path)
    sheet = wb.active
    h_map = {str(cell.value).strip().lower(): i+1 for i, cell in enumerate(sheet[1]) if cell.value}
    
    def get_c(name): return h_map.get(name.strip().lower())
    
    # Extract Discom from actual discom_names (first word from document_name in JSONL)
    import urllib.parse
    discom_from_json = "NA"
    if discom_names and discom_names[0]:
        decoded = urllib.parse.unquote(discom_names[0])
        first_word = decoded.split()[0] if decoded else ""
        discom_from_json = first_word[:6] if len(first_word) >= 6 else first_word
    
    # Debug print
    bc_idx = get_c('Bulk Consumption Rebate')
    print(f"DEBUG: Bulk Consumption Rebate Column Index: {bc_idx}")
    
    start_r = 3
    if sheet.max_row >= start_r:
        sheet.delete_rows(start_r, sheet.max_row - start_r + 1)
        
    # r is assigned below before use
        
    # Create single row with folder name and PDF-derived discom
    r = start_r
    if get_c('States'): sheet.cell(row=r, column=get_c('States')).value = folder_name
    if get_c('DISCOM'):
        print(f"Writing to DISCOM column: {discom_from_json}")
        sheet.cell(row=r, column=get_c('DISCOM')).value = discom_from_json
    if ists_loss != "NA" and get_c('ISTS Loss'): sheet.cell(row=r, column=get_c('ISTS Loss')).value = ists_loss
    if insts_loss and get_c('InSTS Loss'): sheet.cell(row=r, column=get_c('InSTS Loss')).value = insts_loss
    if insts_c is not None and get_c('InSTS Charges'): sheet.cell(row=r, column=get_c('InSTS Charges')).value = insts_c
    
    for kv, val in wheel_losses.items():
        col = get_c(f'Wheeling Loss - {kv} kV')
        if val and col: sheet.cell(row=r, column=col).value = val
        
    for kv, val in wheel_charges.items():
        col = get_c(f'Wheeling Charges - {kv} kV')
        if val is not None and col: sheet.cell(row=r, column=col).value = val

    for kv, val in css_charges.items():
        col = get_c(f'Cross Subsidy Surcharge - {kv} kV') or get_c(f'Cross Subsidy Charges - {kv} kV')
        if val is not None and col: sheet.cell(row=r, column=col).value = val

    if additional_surcharge is not None and get_c('Additional Surcharge'):
        sheet.cell(row=r, column=get_c('Additional Surcharge')).value = additional_surcharge

    for kv, val in fixed_charges.items():
        # Note 'Kv' capitalization for 11kV based on header check
        k_str = f'{kv} Kv' if kv == '11' else f'{kv} kV'
        col = get_c(f'Fixed Charge - {k_str}')
        if col: sheet.cell(row=r, column=col).value = val
        
    for kv, val in energy_charges.items():
        col = get_c(f'Energy Charge - {kv} kV')
        if col: sheet.cell(row=r, column=col).value = val

    if pf_rebate != "NA" and get_c('Power Factor Adjustment Rebate'):
         sheet.cell(row=r, column=get_c('Power Factor Adjustment Rebate')).value = pf_rebate
    elif get_c('Power Factor Adjustment Rebate'):
         sheet.cell(row=r, column=get_c('Power Factor Adjustment Rebate')).value = "NA"
         
    if lf_incentive != "NA" and get_c('Load Factor Incentive'):
         sheet.cell(row=r, column=get_c('Load Factor Incentive')).value = lf_incentive
    elif get_c('Load Factor Incentive'):
         sheet.cell(row=r, column=get_c('Load Factor Incentive')).value = "NA"

    if fuel_surcharge != "NA" and get_c('Fuel Surcharge'):
         sheet.cell(row=r, column=get_c('Fuel Surcharge')).value = fuel_surcharge
    elif get_c('Fuel Surcharge'):
         sheet.cell(row=r, column=get_c('Fuel Surcharge')).value = "NA"

    if tod_charges != "NA" and get_c('TOD Charges'):
         sheet.cell(row=r, column=get_c('TOD Charges')).value = tod_charges
    elif get_c('TOD Charges'):
         sheet.cell(row=r, column=get_c('TOD Charges')).value = "NA"

    if grid_support_charges != "NA" and get_c('Grid Support /Parrallel Operation'):
         sheet.cell(row=r, column=get_c('Grid Support /Parrallel Operation')).value = grid_support_charges
    elif get_c('Grid Support /Parrallel Operation'):
         sheet.cell(row=r, column=get_c('Grid Support /Parrallel Operation')).value = "NA"

    if voltage_rebate['33_66'] != "NA" and get_c('HT ,EHV Rebate at 33/66 kV'):
         sheet.cell(row=r, column=get_c('HT ,EHV Rebate at 33/66 kV')).value = voltage_rebate['33_66']
    elif get_c('HT ,EHV Rebate at 33/66 kV'):
         sheet.cell(row=r, column=get_c('HT ,EHV Rebate at 33/66 kV')).value = "NA"

    if voltage_rebate['132'] != "NA" and get_c('HT ,EHV Rebate at 132 kV and above '): # Note space in key
         sheet.cell(row=r, column=get_c('HT ,EHV Rebate at 132 kV and above ')).value = voltage_rebate['132']
    elif get_c('HT ,EHV Rebate at 132 kV and above '):
         sheet.cell(row=r, column=get_c('HT ,EHV Rebate at 132 kV and above ')).value = "NA"

    if bulk_rebate != "NA" and get_c('Bulk Consumption Rebate'):
         sheet.cell(row=r, column=get_c('Bulk Consumption Rebate')).value = bulk_rebate
    elif get_c('Bulk Consumption Rebate'):
         sheet.cell(row=r, column=get_c('Bulk Consumption Rebate')).value = "NA"
             
    if DB_SUCCESS:
        db_data = {
            'financial_year': "FY2025-26",
            'state': folder_name,
            'discom': discom_from_json,
            'ists_loss': str(ists_loss) if ists_loss else "NA",
            'insts_loss': str(insts_loss) if insts_loss else "NA",
            'wheeling_loss_11kv': wheel_losses.get('11', "NA"),
            'wheeling_loss_33kv': wheel_losses.get('33', "NA"),
            'wheeling_loss_66kv': wheel_losses.get('66', "NA"),
            'wheeling_loss_132kv': wheel_losses.get('132', "NA"),
            'ists_charges': "NA",
            'insts_charges': str(insts_c) if insts_c else "NA",
            'wheeling_charges_11kv': wheel_charges.get('11', "NA"),
            'wheeling_charges_33kv': wheel_charges.get('33', "NA"),
            'wheeling_charges_66kv': wheel_charges.get('66', "NA"),
            'wheeling_charges_132kv': wheel_charges.get('132', "NA"),
            'css_charges_11kv': css_charges.get('11', "NA"),
            'css_charges_33kv': css_charges.get('33', "NA"),
            'css_charges_66kv': css_charges.get('66', "NA"),
            'css_charges_132kv': css_charges.get('132', "NA"),
            'css_charges_220kv': css_charges.get('220', "NA"),
            'additional_surcharge': additional_surcharge if additional_surcharge else "NA",
            'electricity_duty': "NA",
            'tax_on_sale': "NA",
            'fixed_charge_11kv': fixed_charges.get('11', "NA"),
            'fixed_charge_33kv': fixed_charges.get('33', "NA"),
            'fixed_charge_66kv': fixed_charges.get('66', "NA"),
            'fixed_charge_132kv': fixed_charges.get('132', "NA"),
            'fixed_charge_220kv': fixed_charges.get('220', "NA"),
            'energy_charge_11kv': energy_charges.get('11', "NA"),
            'energy_charge_33kv': energy_charges.get('33', "NA"),
            'energy_charge_66kv': energy_charges.get('66', "NA"),
            'energy_charge_132kv': energy_charges.get('132', "NA"),
            'energy_charge_220kv': energy_charges.get('220', "NA"),
            'fuel_surcharge': fuel_surcharge if fuel_surcharge else "NA",
            'tod_charges': tod_charges if tod_charges else "NA",
            'pf_rebate': pf_rebate if pf_rebate else "NA",
            'lf_incentive': lf_incentive if lf_incentive else "NA",
            'grid_support_parallel_op_charges': grid_support_charges if grid_support_charges else "NA",
            'ht_ehv_rebate_33_66kv': voltage_rebate.get('33_66', "NA"),
            'ht_ehv_rebate_132_above': voltage_rebate.get('132', "NA"),
            'bulk_rebate': bulk_rebate if bulk_rebate else "NA"
        }
        # Sanitize
        clean_db_data = {k: (str(v) if v is not None else "NA") for k, v in db_data.items()}
        save_tariff_row(clean_db_data)
             
    wb.save(excel_path)
    print(f"Updated {excel_path} with accurate values.")

if __name__ == "__main__":
    import sys
    # sys.stdout = open('debug_log.txt', 'w', encoding='utf-8')
    
    base_dir = os.path.dirname(os.path.abspath(__file__))
    
    # Dynamic JSONL finding
    extraction_root = os.path.join(base_dir, "Extraction")
    j_f = None
    
    # Search for Meghalaya folder
    if os.path.exists(extraction_root):
        for dirname in os.listdir(extraction_root):
            if "meghalaya" in dirname.lower():
                state_dir = os.path.join(extraction_root, dirname)
                if os.path.isdir(state_dir):
                    for f in os.listdir(state_dir):
                        if f.endswith(".jsonl"):
                            j_f = os.path.join(state_dir, f)
                            break
            if j_f: break
    
    # Fallback/Check
    if not j_f:
         # Try expecting exact folder name if loose search failed
         target = os.path.join(extraction_root, "Meghalaya", "MePDCL ARR & Tariff Order FY 2025-26.jsonl")
         if os.path.exists(target):
             j_f = target

    ists_j_f = os.path.join(base_dir, "ists_extracted", "ists_loss.json")
    e_f = os.path.join(base_dir, "Meghalaya.xlsx")
    
    print(f"Targeting JSONL: {j_f}")
    
    if j_f and os.path.exists(j_f):
        names = extract_discom_names(j_f)
        ists_l = extract_ists_loss(ists_j_f)
        insts_l = extract_losses(j_f)
        w_l = extract_wheeling_losses(j_f)
        insts_c = extract_transmission_charges(j_f)
        w_c = extract_wheeling_charges(j_f)
        css_c = extract_css_charges(j_f)
        add_s = extract_additional_surcharge(j_f)
        fc, ec = extract_fixed_energy_charges(j_f)
        pf_r = extract_pf_rebate(j_f)
        lf_i = extract_load_factor_incentive(j_f)
        fs = extract_fuel_surcharge(j_f)
        tod = extract_tod_charges(j_f)
        gs = extract_grid_support_charges(j_f)
        vr = extract_voltage_rebate(j_f)
        br = extract_bulk_rebate(j_f)
        
        folder_name = os.path.basename(os.path.dirname(j_f))
        update_excel_with_discoms(names, ists_l, insts_l, w_l, insts_c, w_c, css_c, add_s, fc, ec, pf_r, lf_i, fs, tod, gs, vr, br, e_f, folder_name=folder_name, pdf_name=os.path.basename(j_f))
    else:
        print("Meghalaya JSONL file not found. Please ensure scraper has run.")

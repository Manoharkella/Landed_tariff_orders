import sqlite3
import os
import openpyxl
from datetime import datetime

DB_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "tariff_orders.db")

def init_db():
    """Initializes the SQLite database and creates the tariff_data table."""
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS tariff_data (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            financial_year TEXT,
            state TEXT,
            discom TEXT,
            ists_loss TEXT,
            insts_loss TEXT,
            wheeling_loss_11kv TEXT,
            wheeling_loss_33kv TEXT,
            wheeling_loss_66kv TEXT,
            wheeling_loss_132kv TEXT,
            ists_charges TEXT,
            insts_charges TEXT,
            wheeling_charges_11kv TEXT,
            wheeling_charges_33kv TEXT,
            wheeling_charges_66kv TEXT,
            wheeling_charges_132kv TEXT,
            css_charges_11kv TEXT,
            css_charges_33kv TEXT,
            css_charges_66kv TEXT,
            css_charges_132kv TEXT,
            css_charges_220kv TEXT,
            additional_surcharge TEXT,
            electricity_duty TEXT,
            tax_on_sale TEXT,
            fixed_charge_11kv TEXT,
            fixed_charge_33kv TEXT,
            fixed_charge_66kv TEXT,
            fixed_charge_132kv TEXT,
            fixed_charge_220kv TEXT,
            energy_charge_11kv TEXT,
            energy_charge_33kv TEXT,
            energy_charge_66kv TEXT,
            energy_charge_132kv TEXT,
            energy_charge_220kv TEXT,
            fuel_surcharge TEXT,
            tod_charges TEXT,
            pf_rebate TEXT,
            lf_incentive TEXT,
            grid_support_parallel_op_charges TEXT,
            ht_ehv_rebate_33_66kv TEXT,
            ht_ehv_rebate_132_above TEXT,
            bulk_rebate TEXT,
            updated_at DATETIME
        )
    ''')
    conn.commit()
    conn.close()
    print(f"Database initialized at {DB_PATH}")

def save_tariff_row(data):
    """
    Saves or updates a single row of tariff data.
    'data' should be a dictionary with keys matching the table column names.
    """
    if not os.path.exists(DB_PATH):
        init_db()
        
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    
    state = data.get('state')
    discom = data.get('discom')
    
    # Check if record exists
    cursor.execute("SELECT id FROM tariff_data WHERE state = ? AND discom = ?", (state, discom))
    existing = cursor.fetchone()
    
    data['updated_at'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    
    if existing:
        # Update
        row_id = existing[0]
        cols = []
        vals = []
        for k, v in data.items():
            cols.append(f"{k} = ?")
            vals.append(v)
        vals.append(row_id)
        sql = f"UPDATE tariff_data SET {', '.join(cols)} WHERE id = ?"
        cursor.execute(sql, vals)
    else:
        # Insert
        cols = list(data.keys())
        vals = list(data.values())
        placeholders = ', '.join(['?'] * len(cols))
        sql = f"INSERT INTO tariff_data ({', '.join(cols)}) VALUES ({placeholders})"
        cursor.execute(sql, vals)
        
    conn.commit()
    conn.close()

def sync_excel_to_db(excel_path):
    """
    Reads an Excel file and syncs its content to the database.
    Assumes standard column mapping discovered for this project.
    """
    if not os.path.exists(excel_path):
        return
    
    try:
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        sheet = wb.active
        
        # Header mapping
        headers = [str(cell.value).strip() if cell.value else None for cell in sheet[1]]
        
        def get_col(name):
            try: return headers.index(name) + 1
            except: return None

        # Column mapping (Target DB Field -> Excel Header Name)
        mapping = {
            'financial_year': 'Financial Year',
            'state': 'States',
            'discom': 'DISCOM',
            'ists_loss': 'ISTS Loss',
            'insts_loss': 'InSTS Loss',
            'wheeling_loss_11kv': 'Wheeling Loss - 11 kV',
            'wheeling_loss_33kv': 'Wheeling Loss - 33 kV',
            'wheeling_loss_66kv': 'Wheeling Loss - 66 kV',
            'wheeling_loss_132kv': 'Wheeling Loss - 132 kV',
            'ists_charges': 'ISTS Charges',
            'insts_charges': 'InSTS Charges',
            'wheeling_charges_11kv': 'Wheeling Charges - 11 kV',
            'wheeling_charges_33kv': 'Wheeling Charges - 33 kV',
            'wheeling_charges_66kv': 'Wheeling Charges - 66 kV',
            'wheeling_charges_132kv': 'Wheeling Charges - 132 kV',
            'css_charges_11kv': 'Cross Subsidy Surcharge - 11 kV',
            'css_charges_33kv': 'Cross Subsidy Surcharge - 33 kV',
            'css_charges_66kv': 'Cross Subsidy Surcharge - 66 kV',
            'css_charges_132kv': 'Cross Subsidy Surcharge - 132 kV',
            'css_charges_220kv': 'Cross Subsidy Surcharge - 220 kV',
            'additional_surcharge': 'Additional Surcharge',
            'electricity_duty': 'Electric Duty',
            'tax_on_sale': 'Tax on Sale',
            'fixed_charge_11kv': 'Fixed Charge - 11 kV',
            'fixed_charge_33kv': 'Fixed Charge - 33 kV',
            'fixed_charge_66kv': 'Fixed Charge - 66 kV',
            'fixed_charge_132kv': 'Fixed Charge - 132 kV',
            'fixed_charge_220kv': 'Fixed Charge - 220 kV',
            'energy_charge_11kv': 'Energy Charge - 11 kV',
            'energy_charge_33kv': 'Energy Charge - 33 kV',
            'energy_charge_66kv': 'Energy Charge - 66 kV',
            'energy_charge_132kv': 'Energy Charge - 132 kV',
            'energy_charge_220kv': 'Energy Charge - 220 kV',
            'fuel_surcharge': 'Fuel Surcharge',
            'tod_charges': 'TOD Charges',
            'pf_rebate': 'Power Factor Adjustment Rebate',
            'lf_incentive': 'Load Factor Incentive',
            'grid_support_parallel_op_charges': 'Grid Support /Parrallel Operation',
            'ht_ehv_rebate_33_66kv': 'HT ,EHV Rebate at 33/66 kV',
            'ht_ehv_rebate_132_above': 'HT ,EHV Rebate at 132 kV and above ',
            'bulk_rebate': 'Bulk Consumption Rebate'
        }

        # Also handle some variations
        mapping_variations = {
            'fixed_charge_11kv': ['Fixed Charge - 11 Kv', 'Fixed Charge - 11kV'],
            'fixed_charge_33kv': ['Fixed Charge - 33kV'],
            'energy_charge_11kv': ['Energy Charge - 11kV'],
            'ht_ehv_rebate_132_above': ['HT ,EHV Rebate at 132 kV and above', 'HT ,EHV Rebate at 132 kV and above '],
            'grid_support_parallel_op_charges': ['Grid Support /Parallel Operation', 'Grid Support /Parrallel Operation', 'Grid Support / Parallel Operation Charges']
        }

        col_idxs = {}
        for db_field, header_name in mapping.items():
            idx = get_col(header_name)
            if idx is None and db_field in mapping_variations:
                for var in mapping_variations[db_field]:
                    idx = get_col(var)
                    if idx: break
            col_idxs[db_field] = idx

        # Start from row 3 (data rows)
        for row_idx in range(3, sheet.max_row + 1):
            row_data = {}
            for db_field, col_idx in col_idxs.items():
                if col_idx:
                    val = sheet.cell(row=row_idx, column=col_idx).value
                    row_data[db_field] = str(val) if val is not None else "NA"
            
            state_val = row_data.get('state')
            if state_val and state_val != "NA":
                # Fix common typos
                if "chastisgarh" in state_val.lower(): row_data['state'] = "Chhattisgarh"
                if "madya" in state_val.lower(): row_data['state'] = "Madhya Pradesh"
                if "rajastan" in state_val.lower(): row_data['state'] = "Rajasthan"
                
                save_tariff_row(row_data)

        print(f"Synced {excel_path} to database.")
    except Exception as e:
        print(f"Error syncing {excel_path}: {e}")

if __name__ == "__main__":
    init_db()
    # Sync all existing xlsx files
    for f in os.listdir(os.path.dirname(os.path.abspath(__file__))):
        if f.endswith(".xlsx"):
            sync_excel_to_db(f)

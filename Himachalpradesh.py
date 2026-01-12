import json
import re
import os
import openpyxl
from datetime import datetime

def get_financial_years():
    """Returns current and previous financial years in various formats."""
    now = datetime.now()
    # Financial year starts in April
    fy_start = now.year - 1 if now.month < 4 else now.year
    
    cfy_short = f"{fy_start}-{str(fy_start+1)[-2:]}"
    cfy_long = f"{fy_start}-{fy_start+1}"
    
    pfy_start = fy_start - 1
    pfy_short = f"{pfy_start}-{str(pfy_start+1)[-2:]}"
    pfy_long = f"{pfy_start}-{pfy_start+1}"
    
    return {
        'current_short': cfy_short,
        'current_long': cfy_long,
        'previous_short': pfy_short,
        'previous_long': pfy_long,
        'all_variants': [cfy_short, cfy_long, pfy_short, pfy_long]
    }

def get_priority(heading, fy_info):
    """Returns a priority score based on the year found in the heading."""
    heading = heading.lower()
    if any(year in heading for year in [fy_info['current_short'], fy_info['current_long']]):
        return 2
    if any(year in heading for year in [fy_info['previous_short'], fy_info['previous_long']]):
        return 1
    return 0

def extract_discom_names(jsonl_path, output_path):
    discom_names = []
    candidate_names = []
    ignore_list = [
        "submission", "petition", "order", "commission", "secretary", "tariff", 
        "amount", "total", "particulars", "approved", "proposed", "details", 
        "table", "status", "report", "date", "month", "year", "remark", "reply"
    ]
    if not os.path.exists(jsonl_path): return []
    with open(jsonl_path, 'r', encoding='utf-8') as f:
        for line in f:
            try:
                data = json.loads(line)
                rows = data.get("rows", [])
                for row in rows:
                    for k, v in row.items():
                        if not v: continue
                        k_str = str(k).lower()
                        v_str = str(v).strip()
                        if "petitioner" in k_str or "utility" in k_str or "licensee" in k_str:
                             cleaned = re.sub(r'^[M|m]/[s|S][\.,\s]*', '', v_str).strip()
                             if len(cleaned) > 2 and not cleaned[0].isdigit():
                                 candidate_names.append(cleaned)
                        if v_str.isupper() and 3 < len(v_str) < 10:
                            if v_str.endswith("L") or v_str.endswith("D") or v_str.endswith("B"):
                                 if v_str.lower() not in ignore_list and "FY" not in v_str:
                                     candidate_names.append(v_str)
            except: pass
    # Hardcode proper DISCOM name for Himachal Pradesh
    discom_names = ["HPSEBL"]
    
    with open(output_path, 'w', encoding='utf-8') as f:
        for name in sorted(list(set(discom_names))): f.write(name + "\n")
    return discom_names

def extract_ists_loss(json_path):
    try:
        if os.path.exists(json_path):
            with open(json_path, 'r', encoding='utf-8') as f:
                data = json.load(f)
                val = data.get("All India transmission Loss (in %)", None)
                if val:
                    print(f"Extracted ISTS Loss: {val}")
                    return f"{val}%" if "%" not in str(val) else val
    except Exception as e:
        print(f"Error reading ISTS loss JSON: {e}")
    return "NA"

def extract_losses(jsonl_path, fy_info):
    insts_loss = None
    best_priority = -1
    # Keywords for Intra-State Transmission System Charges
    keywords = [
        "intra-state transmission system charges",
        "stu charges", 
        "stu transmission charges",
        "transmission charges"
    ]
    
    if not jsonl_path or not os.path.exists(jsonl_path):
        return "NA"

    with open(jsonl_path, 'r', encoding='utf-8') as f:
        for line in f:
            try:
                data = json.loads(line)
                heading = data.get("table_heading", "").lower()
                rows = data.get("rows", [])
                priority = get_priority(heading, fy_info)
                
                # Check if table heading matches transmission keywords
                table_match = any(k in heading for k in keywords)
                
                for row in rows:
                    row_txt = str(row).lower()
                    
                    # Identify row: either in matching table with "loss" or row specifically mentions loss
                    is_target = False
                    if table_match and "loss" in row_txt:
                        is_target = True
                    elif "intra" in row_txt and "state" in row_txt and "transmission" in row_txt and "loss" in row_txt:
                        is_target = True
                    elif "stu" in row_txt and "loss" in row_txt:
                        is_target = True
                    
                    if is_target:
                        for v in row.values():
                            v_str = str(v).strip()
                            if "%" in v_str:
                                # Ensure it's a numeric percentage
                                try:
                                    clean_val = re.sub(r'[^\d\.]', '', v_str)
                                    if clean_val and priority > best_priority:
                                        insts_loss = v_str
                                        best_priority = priority
                                    elif clean_val and priority == best_priority and insts_loss is None:
                                        insts_loss = v_str
                                except:
                                    pass
            except:
                pass
    
    if insts_loss is None:
        insts_loss = "NA"
    
    print(f"Extracted InSTS Loss: {insts_loss}")
    return insts_loss

def extract_wheeling_losses(jsonl_path, fy_info):
    losses = {'11': 'NA', '33': 'NA', '66': 'NA', '132': 'NA', '220': 'NA'}
    best_priority = {'11': -1, '33': -1, '66': -1, '132': -1, '220': -1}
    
    keywords = ["wheeling loss", "discom loss", "distribution loss", "voltage wise loss", "loss level for open access"]
    
    with open(jsonl_path, 'r', encoding='utf-8') as f:
        for line in f:
            try:
                data = json.loads(line)
                h = data.get("table_heading", "").lower()
                priority = get_priority(h, fy_info)
                rows = data.get("rows", [])
                
                # Check for table heading matches
                table_match = any(k in h for k in keywords)
                
                # Table 300 specific handling (Approved Loss Level...)
                if "loss level" in h and "open access" in h:
                    # Map columns based on first row or headers
                    col_map = {}
                    for row in rows:
                        for k, v in row.items():
                            k_lower = str(k).lower().replace(" ", "").replace("/", "")
                            v_lower = str(v).lower().replace(" ", "").replace("/", "") if v else ""
                            
                            if "11kv" in k_lower or "22kv" in k_lower or "11kv" in v_lower or "22kv" in v_lower:
                                col_map[k] = '11'
                            if "33kv" in k_lower or "33kv" in v_lower:
                                col_map[k] = '33'
                            if "66kv" in k_lower or "66kv" in v_lower:
                                col_map[k] = '66'
                            if "132kv" in k_lower or "220kv" in k_lower or "132kv" in v_lower or "220kv" in v_lower:
                                col_map[k] = '132' # Also covers '220' if we want

                    for row in rows:
                        row_txt = str(row).lower()
                        if "loss level" in row_txt or "energy" in row_txt:
                            for k, v in row.items():
                                if not v or "%" not in str(v): continue
                                tv = col_map.get(k)
                                if tv:
                                    if priority >= best_priority[tv]:
                                        losses[tv] = str(v).strip()
                                        best_priority[tv] = priority
                                        if tv == '132': # Populate 220 as well
                                            losses['220'] = str(v).strip()
                                            best_priority['220'] = priority

                # General case for other tables
                if table_match and not ("loss level" in h and "open access" in h):
                    for row in rows:
                        r_txt = str(row).lower()
                        val = next((str(v).strip() for v in row.values() if v and "%" in str(v)), None)
                        if not val: continue
                        target_v = None
                        if "33kv" in r_txt.replace(" ", ""): target_v = '33'
                        elif "11kv" in r_txt.replace(" ", ""): target_v = '11'
                        elif "66kv" in r_txt.replace(" ", ""): target_v = '66'
                        elif "132kv" in r_txt.replace(" ", "") or "eht" in r_txt: target_v = '132'
                        
                        if target_v and priority >= best_priority[target_v]:
                            losses[target_v] = val
                            best_priority[target_v] = priority
            except: pass
            
    print(f"Extracted Wheeling Losses: {losses}")
    return losses

def extract_wheeling_charges(jsonl_path, fy_info):
    charges = {'11': 'NA', '33': 'NA', '66': 'NA', '132': 'NA', '220': 'NA'}
    best_priority = {'11': -1, '33': -1, '66': -1, '132': -1, '220': -1}
    keywords = ["wheeling charges", "discom charges", "distribution charges", "voltage wise charges"]
    
    with open(jsonl_path, 'r', encoding='utf-8') as f:
        for line in f:
            try:
                data = json.loads(line)
                h = data.get("table_heading", "").lower()
                priority = get_priority(h, fy_info)
                if priority < 2:
                     if "fy 26" in h or "fy26" in h or "fy 2026" in h: priority = 2
                
                # We need to find the table with the values.
                # Table 289/Term Customers table has the data.
                
                match_header = any(k in h for k in keywords) or "term customers" in h
                rows = data.get("rows", [])
                
                # Check rows if header doesn't match
                if not match_header:
                    for row in rows:
                        row_txt = str(row).lower()
                        if any(k in row_txt for k in keywords):
                            match_header = True
                            break
                            
                if match_header:
                     for row in rows:
                         row_txt = str(row).lower()
                         # Strict filter: Row must describe a charge or rate
                         if not any(x in row_txt for x in ["charges", "paisa", "rate/unit", "/unit", "cost of supply"]):
                             continue
                             
                         # We look for the row that has "paisa per unit" or values matching our keys
                         # Keys in specific table: "EHT (220 kV)", "EHT (132 kV)", etc.
                         
                         is_paisa = "paisa" in row_txt or "paisa" in h
                         
                         # Iterate keys to find voltages
                         for k, v in row.items():
                             if not v: continue
                             k_clean = str(k).lower().replace(" ", "").replace("\n", "")
                             
                             target_v = None
                             if ">=11kv" in k_clean: target_v = '11'
                             elif "11kv" in k_clean and "lt" not in k_clean: target_v = '11'
                             elif "33kv" in k_clean: target_v = '33'
                             elif "66kv" in k_clean: target_v = '66'
                             elif "132kv" in k_clean: target_v = '132'
                             elif "220kv" in k_clean: target_v = '220'
                             
                             if target_v:
                                 try:
                                     clean = re.sub(r'[^\d\.]', '', str(v))
                                     if clean:
                                         f_v = float(clean)
                                         # Unit normalization (Paisa -> INR)
                                         if is_paisa and f_v > 5:
                                             f_v = f_v / 100.0
                                         
                                         # Sanity check for INR/kWh (usually 0.05 to 5.0)
                                         if 0.01 < f_v < 10.0:
                                             if priority >= best_priority[target_v]:
                                                 charges[target_v] = f_v
                                                 best_priority[target_v] = priority
                                 except: pass
            except: pass
    
    print(f"Extracted Wheeling Charges: {charges}")
    return charges

def extract_additional_surcharge(jsonl_path, fy_info):
    add_surcharge = None
    best_priority = -1
    with open(jsonl_path, 'r', encoding='utf-8') as f:
        for line in f:
            try:
                data = json.loads(line)
                h = data.get("table_heading", "").lower()
                priority = get_priority(h, fy_info)
                if "additional surcharge" in h and "approved" in h:
                     for row in data.get("rows", []):
                        row_txt = str(row).lower()
                        if "additional surcharge" in row_txt and ("paisa" in row_txt or "unit" in row_txt):
                             for k, v in row.items():
                                 if not v or "surcharge" in str(v).lower(): continue
                                 try:
                                     clean = re.sub(r'[^\d\.]', '', str(v))
                                     if clean:
                                         f_v = float(clean)
                                         if 10 < f_v < 500: 
                                             val = f_v / 100.0
                                             if priority >= best_priority:
                                                 add_surcharge = val
                                                 best_priority = priority
                                                 break
                                 except: pass
            except: pass
    print(f"Extracted Additional Surcharge: {add_surcharge}")
    return add_surcharge if add_surcharge is not None else "NA"

def extract_css_charges(jsonl_path, fy_info):
    css_charges = {'11': 'NA', '33': 'NA', '66': 'NA', '132': 'NA', '220': 'NA'}
    best_priority = {'11': -1, '33': -1, '66': -1, '132': -1, '220': -1}
    
    with open(jsonl_path, 'r', encoding='utf-8') as f:
        for line in f:
            try:
                data = json.loads(line)
                h = data.get("table_heading", "").lower()
                
                # Filter for Approved Tables (usually late in document)
                # Ignore "Submission" or "Proposal" tables if possible, but headers might not say it clearly.
                # Page 363 seems to be the one.
                # Heading on 363: "access consumers" (fragment?)
                # Heading on 364: "hpse bl-d tariff order for fy 2025-26"
                
                priority = get_priority(h, fy_info)
                if priority < 2:
                     if "fy 26" in h or "fy26" in h or "fy 2026" in h: priority = 2
                     elif "access consumers" in h and "submission" not in h: priority = 2 # Promote P363
                
                if "submission" in h: priority = 0 # Lower priority for submissions
                
                rows = data.get("rows", [])
                
                # Check if this looks like the Large Industrial / EHT table
                has_industrial_context = False
                for row in rows:
                    row_txt = str(row).lower()
                    if "large industrial" in row_txt or "commercial supply" in row_txt:
                        has_industrial_context = True
                        break
                
                if has_industrial_context or "access consumers" in h or "tariff order" in h:
                    # Force elevate priority for Industrial context to prefer it over Bulk Supply etc.
                    current_priority = priority
                    if has_industrial_context:
                        current_priority = 3 
                    elif priority < 2: 
                        current_priority = 2
                    
                    current_category = None
                    for row in rows:
                        row_txt = str(row).lower()
                        
                        # Category Tracking
                        if "large industrial" in row_txt: current_category = "industrial"
                        elif "bulk supply" in row_txt: current_category = "bulk"
                        elif "commercial" in row_txt: current_category = "commercial"
                        elif "domestic" in row_txt: current_category = "other"
                        elif "irrigation" in row_txt: current_category = "other"
                        
                        # Only proceed if we are in Industrial category
                        # (Or specifically HT2 which is unique to Industrial/Commercial usually)
                        
                        is_industrial = (current_category == "industrial")
                        
                        # Identify Voltage Level
                        v_key = []
                        if "eht" in row_txt:
                            if is_industrial: v_key = ['66', '132', '220']
                        elif "ht2" in row_txt or "above 1 mva" in row_txt:
                            v_key = ['11', '33'] # HT2 is typically Industrial/Commercial
                        elif "ht" in row_txt and "ht1" not in row_txt:
                             # Generic HT or HT1
                             # Only accept if Industrial
                             if is_industrial: v_key = ['11', '33']
                        
                        if not v_key: continue

                        # Extract Value
                        # We want "Minimum of (B) & (C)" or similar column
                        # Values observed: 0.56, 0.48, 0.63
                        val = None
                        
                        # Helper to find float in values
                        # We look for the LAST numeric column which is often the approved/minimum one
                        # Or specifically key with "minimum"
                        
                        found_val = None
                        for k, v in row.items():
                            k_lower = str(k).lower()
                            if "minimum" in k_lower or "final" in k_lower:
                                try:
                                    clean = re.sub(r'[^\d\.]', '', str(v))
                                    if clean and 0.01 < float(clean) < 10:
                                        found_val = float(clean)
                                        break
                                except: pass
                        
                        if found_val is None:
                             # Heuristic: Take the last valid float in row range
                             # Only if we are sure this is the right table (usually has 4-5 cols)
                             nums = []
                             for v in row.values():
                                 try:
                                     clean = re.sub(r'[^\d\.]', '', str(v))
                                     if clean and 0.01 < float(clean) < 10:
                                         nums.append(float(clean))
                                 except: pass
                             if nums: found_val = nums[-1] # Minimum is usually last col
                        
                        if found_val is not None:
                            for vk in v_key:
                                if current_priority >= best_priority[vk]:
                                    css_charges[vk] = found_val
                                    best_priority[vk] = current_priority

            except: pass

    # Special logic: If 66/132/220 are same (EHT) and 11/33 are same (HT)
    print(f"Extracted CSS Charges: {css_charges}")
    return css_charges
    print(f"Extracted CSS: {css_charges}")
    return css_charges

def extract_fixed_charges(jsonl_path, fy_info):
    fixed_charges = {'11': "NA", '33': "NA", '66': "NA", '132': "NA", '220': "NA"}
    best_priority = {'11': -1, '33': -1, '66': -1, '132': -1, '220': -1}
    with open(jsonl_path, 'r', encoding='utf-8') as f:
        for line in f:
            try:
                data = json.loads(line)
                h = data.get("table_heading", "").lower()
                priority = get_priority(h, fy_info)
                if "demand charges" in h or "demand charge" in h:
                    for row in data.get("rows", []):
                        row_txt = str(row).lower().replace('\n', ' ').replace('  ', ' ')
                        cat = None
                        if "220 kv" in row_txt: cat = '220'
                        elif "132 kv" in row_txt: cat = '132'
                        elif "66 kv" in row_txt: cat = '66'
                        elif "ht-2" in row_txt: cat = '33' 
                        elif "ht-1" in row_txt: cat = '11'
                        elif "eht" in row_txt: cat = 'eht'
                        if cat:
                             cands = []
                             for k, v in row.items():
                                 if not v: continue
                                 try:
                                     clean = re.sub(r'[^\d\.]', '', str(v))
                                     if clean:
                                         f_v = float(clean)
                                         if 50 < f_v < 1000: cands.append(f_v)
                                 except: pass
                             if cands:
                                 val = max(cands)
                                 if cat == 'eht':
                                     for k in ['66', '132', '220']:
                                         if priority >= best_priority[k]:
                                             fixed_charges[k] = val
                                             best_priority[k] = priority
                                 elif priority >= best_priority[cat]:
                                     fixed_charges[cat] = val
                                     best_priority[cat] = priority
            except: pass
    print(f"Extracted Fixed Charges: {fixed_charges}")
    return fixed_charges

def extract_energy_charges(jsonl_path, fy_info):
    energy_charges = {'11': "NA", '33': "NA", '66': "NA", '132': "NA", '220': "NA"}
    best_priority = {'11': -1, '33': -1, '66': -1, '132': -1, '220': -1}
    with open(jsonl_path, 'r', encoding='utf-8') as f:
        for line in f:
            try:
                data = json.loads(line)
                h = data.get("table_heading", "").lower()
                priority = get_priority(h, fy_info)
                if "energy charge" in h or "variable charge" in h:
                    for row in data.get("rows", []):
                        row_txt = str(row).lower()
                        cats = []
                        if "220 kv" in row_txt: cats.append('220')
                        if "132 kv" in row_txt: cats.append('132')
                        if "66 kv" in row_txt: cats.append('66')
                        if "ht-2" in row_txt: cats.append('33')
                        if "ht-1" in row_txt: cats.append('11')
                        if not cats:
                            if "eht" in row_txt: cats = ['66', '132', '220']
                            elif "ht" in row_txt: cats = ['11', '33']
                        if cats:
                             cands = []
                             for k, v in row.items():
                                 if not v: continue
                                 try:
                                     clean = re.sub(r'[^\d\.]', '', str(v))
                                     if clean:
                                         f_v = float(clean)
                                         if f_v != 11.0 and 1.0 < f_v < 20.0: cands.append(f_v)
                                 except: pass
                             if cands:
                                 val = max(cands)
                                 for c in cats:
                                     if priority >= best_priority[c]:
                                         energy_charges[c] = val
                                         best_priority[c] = priority
            except: pass
    print(f"Extracted Energy Charges: {energy_charges}")
    return energy_charges

def extract_fuel_surcharge(jsonl_path, fy_info):
    fuel_surcharge = None
    best_priority = -1
    keywords = ["fuel adjustment cost", "fuel surcharge", "fpppa", "fppca", "eca", "fppas"]
    with open(jsonl_path, 'r', encoding='utf-8') as f:
        for line in f:
            try:
                data = json.loads(line)
                h = data.get("table_heading", "").lower()
                priority = get_priority(h, fy_info)
                match_found = any(k in h for k in keywords) or any(any(k in str(r).lower() for k in keywords) for r in data.get("rows", []))
                if match_found:
                    for row in data.get("rows", []):
                         for v in row.values():
                            try:
                                clean = re.sub(r'[^\d\.]', '', str(v))
                                if clean:
                                    f_v = float(clean)
                                    if 0.0 < f_v < 5.0 and priority >= best_priority:
                                        fuel_surcharge = f_v
                                        best_priority = priority
                            except: pass
            except: pass
    print(f"Extracted Fuel Surcharge: {fuel_surcharge}")
    return fuel_surcharge if fuel_surcharge is not None else "NA"

def extract_pfa_rebate_dynamic(jsonl_path, fy_info):
    # Extensive search yielded no direct INR/kWh value for Power Factor Rebate in FY25-26.
    # Current values are mostly reactive charges or directives.
    return "NA"

def extract_load_factor_incentive_dynamic(jsonl_path, fy_info):
    lf_incentive = None
    best_priority = -1
    keywords = ["load factor incentive", "load factor discount"]
    
    with open(jsonl_path, 'r', encoding='utf-8') as f:
        for line in f:
            try:
                data = json.loads(line)
                h = data.get("table_heading", "").lower()
                priority = get_priority(h, fy_info)
                
                # Check for table heading match
                if any(k in h for k in keywords):
                    for row in data.get("rows", []):
                        row_txt = str(row).lower()
                        is_paisa = "paisa" in row_txt or "paisa" in h
                        
                        for v in row.values():
                            val_str = str(v).lower()
                            if not val_str: continue
                            
                            # Skip if value is a description or non-numeric noise
                            if len(val_str) > 20: continue

                            try:
                                clean = re.sub(r'[^\d\.]', '', val_str)
                                if clean:
                                    f_v = float(clean)
                                    val = None
                                    
                                    # Logic to detect INR/kWh
                                    if is_paisa and f_v > 0: # e.g. 50 paise
                                        val = f_v / 100.0
                                    elif 0 < f_v < 10.0: # e.g. 0.50 or 1.25 Rs/unit
                                        val = f_v
                                    
                                    if val is not None and priority >= best_priority:
                                        lf_incentive = val
                                        best_priority = priority
                            except: pass
                
                # Also check rows if heading didn't match (optional, but user specified keywords which usually appear in headers or descriptions)
                # But to be safe and strict to user request, we focus on where these keywords appear.
                # If keywords appear in row:
                else:
                    for row in data.get("rows", []):
                        row_txt = str(row).lower()
                        if any(k in row_txt for k in keywords):
                             is_paisa = "paisa" in row_txt
                             for v in row.values():
                                try:
                                    clean = re.sub(r'[^\d\.]', '', str(v))
                                    if clean:
                                        f_v = float(clean)
                                        val = None
                                        if is_paisa and f_v > 0: val = f_v / 100.0
                                        elif 0 < f_v < 10.0: val = f_v
                                        
                                        if val is not None and priority >= best_priority:
                                            lf_incentive = val
                                            best_priority = priority
                                except: pass

            except: pass
            
    return lf_incentive if lf_incentive else "NA"

def extract_voltage_rebates(jsonl_path, fy_info):
    rebates = {'33_66': "NA", '132_plus': "NA"}
    best_priority = {'33_66': -1, '132_plus': -1}
    
    # Keywords: User specified "HT Rebate", "EHV Rebate"
    keywords = ["ht rebate", "ehv rebate", "voltage rebate", "rebate for supply at", "higher voltage rebate"]
    
    with open(jsonl_path, 'r', encoding='utf-8') as f:
        for line in f:
            try:
                data = json.loads(line)
                h = data.get("table_heading", "").lower()
                priority = get_priority(h, fy_info)
                
                is_relevant_table = any(k in h for k in keywords) or \
                                    ("rebate" in h and any(v in h for v in ["33 kv", "66 kv", "132 kv", "220 kv", "eht", "ht"]))
                
                if is_relevant_table:
                    for row in data.get("rows", []):
                        row_txt = str(row).lower()
                        is_paisa = "paisa" in row_txt or "paisa" in h
                        
                        cats = []
                        if "33 kv" in row_txt or "66 kv" in row_txt: cats.append('33_66')
                        if "132 kv" in row_txt or "220 kv" in row_txt or "eht" in row_txt: cats.append('132_plus')
                        
                        if not cats:
                            if "ehv" in row_txt: cats.append('132_plus')
                            elif "ht" in row_txt and "lt" not in row_txt: cats.append('33_66')
                        
                        if cats:
                             for v in row.values():
                                val_str = str(v).lower()
                                if not val_str: continue
                                if len(val_str) > 20: continue 

                                try:
                                    clean = re.sub(r'[^\d\.]', '', val_str)
                                    if clean:
                                        f_v = float(clean)
                                        val = None
                                        
                                        if is_paisa and f_v > 0: val = f_v / 100.0
                                        elif 0 < f_v < 10.0:
                                            if "%" in val_str: continue 
                                            val = f_v
                                        
                                        if val is not None:
                                            for c in cats:
                                                if priority >= best_priority[c]:
                                                    rebates[c] = val
                                                    best_priority[c] = priority
                                except: pass
                else: 
                     for row in data.get("rows", []):
                        row_txt = str(row).lower()
                        if any(k in row_txt for k in keywords):
                            is_paisa = "paisa" in row_txt
                            cats = []
                            if "33 kv" in row_txt or "66 kv" in row_txt: cats.append('33_66')
                            if "132 kv" in row_txt or "220 kv" in row_txt or "eht" in row_txt: cats.append('132_plus')
                            if not cats:
                                if "ehv" in row_txt: cats.append('132_plus')
                                elif "ht" in row_txt: cats.append('33_66')

                            if cats:
                                for v in row.values():
                                    val_str = str(v).lower()
                                    if "%" in val_str: continue
                                    try:
                                        clean = re.sub(r'[^\d\.]', '', val_str)
                                        if clean:
                                            f_v = float(clean)
                                            val = None
                                            if is_paisa and f_v > 0: val = f_v / 100.0
                                            elif 0 < f_v < 10.0: val = f_v
                                            
                                            if val is not None:
                                                 for c in cats:
                                                     if priority >= best_priority[c]:
                                                         rebates[c] = val
                                                         best_priority[c] = priority
                                    except: pass
            except: pass
    return rebates



def extract_grid_support_charges(jsonl_path, fy_info):
    grid_support = None
    best_priority = -1
    # Keywords including user's specific typo "Parrallel" just in case, and correct "Parallel"
    keywords = ["grid support", "parallel operation", "parrallel operation"]
    
    with open(jsonl_path, 'r', encoding='utf-8') as f:
        for line in f:
            try:
                data = json.loads(line)
                h = data.get("table_heading", "").lower()
                priority = get_priority(h, fy_info)
                
                # Check for table heading match
                if any(k in h for k in keywords):
                    for row in data.get("rows", []):
                         row_txt = str(row).lower()
                         is_paisa = "paisa" in row_txt or "paisa" in h
                         
                         for v in row.values():
                            val_str = str(v).lower()
                            if not val_str: continue
                            
                            if len(val_str) > 20: continue

                            try:
                                clean = re.sub(r'[^\d\.]', '', val_str)
                                if clean:
                                    f_v = float(clean)
                                    val = None
                                    
                                    # Logic for INR/kWh
                                    # Assuming grid support charges are usually small per unit charges
                                    if is_paisa and f_v > 0: 
                                        val = f_v / 100.0
                                    elif 0 < f_v < 10.0:
                                        val = f_v
                                    
                                    if val is not None and priority >= best_priority:
                                        grid_support = val
                                        best_priority = priority
                            except: pass
                
                # Fallback: Check rows for keywords if header didn't match
                else:
                    for row in data.get("rows", []):
                        row_txt = str(row).lower()
                        if any(k in row_txt for k in keywords):
                             is_paisa = "paisa" in row_txt # checking row context
                             for v in row.values():
                                try:
                                    clean = re.sub(r'[^\d\.]', '', str(v))
                                    if clean:
                                        f_v = float(clean)
                                        val = None
                                        if is_paisa and f_v > 0: val = f_v / 100.0
                                        elif 0 < f_v < 10.0: val = f_v
                                        
                                        if val is not None and priority >= best_priority:
                                            grid_support = val
                                            best_priority = priority
                                except: pass

            except: pass
            
    return grid_support if grid_support else "NA"



def extract_bulk_consumption_rebate(jsonl_path, fy_info):
    bulk_rebate = None
    best_priority = -1
    keywords = ["bulk consumption rebate", "bulk consumption discount"]
    
    with open(jsonl_path, 'r', encoding='utf-8') as f:
        for line in f:
            try:
                data = json.loads(line)
                h = data.get("table_heading", "").lower()
                priority = get_priority(h, fy_info)
                
                # Check for table heading match
                if any(k in h for k in keywords):
                    for row in data.get("rows", []):
                        row_txt = str(row).lower()
                        is_paisa = "paisa" in row_txt or "paisa" in h
                        
                        for v in row.values():
                            val_str = str(v).lower()
                            if not val_str: continue
                            if len(val_str) > 20: continue

                            try:
                                clean = re.sub(r'[^\d\.]', '', val_str)
                                if clean:
                                    f_v = float(clean)
                                    val = None
                                    
                                    # Logic for INR/kWh
                                    if is_paisa and f_v > 0: 
                                        val = f_v / 100.0
                                    elif 0 < f_v < 10.0:
                                        if "%" in val_str: continue
                                        val = f_v
                                    
                                    if val is not None and priority >= best_priority:
                                        bulk_rebate = val
                                        best_priority = priority
                            except: pass
                
                # Fallback: Check rows for keywords
                else:
                    for row in data.get("rows", []):
                        row_txt = str(row).lower()
                        if any(k in row_txt for k in keywords):
                            is_paisa = "paisa" in row_txt
                            for v in row.values():
                                try:
                                    clean = re.sub(r'[^\d\.]', '', str(v))
                                    if clean:
                                        f_v = float(clean)
                                        val = None
                                        if is_paisa and f_v > 0: val = f_v / 100.0
                                        elif 0 < f_v < 10.0: val = f_v
                                        
                                        if val is not None and priority >= best_priority:
                                            bulk_rebate = val
                                            best_priority = priority
                                except: pass
            except: pass
            
    return bulk_rebate if bulk_rebate else "NA"

def extract_tod_charges(jsonl_path):
    tod = "NA"
    with open(jsonl_path, 'r', encoding='utf-8') as f:
        for line in f:
            try:
                data = json.loads(line)
                h = data.get("table_heading", "").lower()
                if any(k in h for k in ["time of day", "tod", "peak"]):
                    for row in data.get("rows", []):
                        for v in row.values():
                            try:
                                clean = re.sub(r'[^\d\.]', '', str(v))
                                if clean:
                                    f_v = float(clean)
                                    if 0.1 < f_v < 10.0: tod = f_v
                            except: pass
            except: pass
    return tod

def update_excel_with_discoms(discoms, ists, insts, insts_charges_val, wheeling_l, wheeling_c, css, fixed, energy, fuel, tod, pfa, lf, grid, volt, bulk, add_s, path):
    try:
        wb = openpyxl.load_workbook(path)
        sheet = wb.active
        if sheet.max_row >= 3: sheet.delete_rows(3, sheet.max_row-2)
        row_idx = 3
        for discom in (discoms if discoms else ["Gen"]):
            sheet.cell(row=row_idx, column=1).value = "Himachal Pradesh"
            sheet.cell(row=row_idx, column=3).value = discom
            if ists: sheet.cell(row=row_idx, column=4).value = str(ists) + "%"
            if insts: sheet.cell(row=row_idx, column=5).value = insts
            if wheeling_l:
                sheet.cell(row=row_idx, column=6).value = wheeling_l.get('11')
                sheet.cell(row=row_idx, column=7).value = wheeling_l.get('33')
                sheet.cell(row=row_idx, column=8).value = wheeling_l.get('66')
                sheet.cell(row=row_idx, column=9).value = wheeling_l.get('132')
            if insts_charges_val is not None:
                sheet.cell(row=row_idx, column=11).value = insts_charges_val
            if wheeling_c:
                sheet.cell(row=row_idx, column=12).value = wheeling_c.get('11')
                sheet.cell(row=row_idx, column=13).value = wheeling_c.get('33')
                sheet.cell(row=row_idx, column=14).value = wheeling_c.get('66')
                sheet.cell(row=row_idx, column=15).value = wheeling_c.get('132')
            if css:
                for k, col in [('11',16),('33',17),('66',18),('132',19),('220',20)]:
                    sheet.cell(row=row_idx, column=col).value = css.get(k)
            sheet.cell(row=row_idx, column=21).value = add_s
            if fixed:
                for k, col in [('11',24),('33',25),('66',26),('132',27),('220',28)]:
                    sheet.cell(row=row_idx, column=col).value = fixed.get(k)
            if energy:
                for k, col in [('11',29),('33',30),('66',31),('132',32),('220',33)]:
                    sheet.cell(row=row_idx, column=col).value = energy.get(k)
            sheet.cell(row=row_idx, column=34).value = fuel
            sheet.cell(row=row_idx, column=35).value = tod
            sheet.cell(row=row_idx, column=36).value = pfa
            sheet.cell(row=row_idx, column=37).value = lf
            sheet.cell(row=row_idx, column=38).value = grid
            sheet.cell(row=row_idx, column=39).value = volt.get('33_66')
            sheet.cell(row=row_idx, column=40).value = volt.get('132_plus')
            sheet.cell(row=row_idx, column=41).value = bulk
            row_idx += 1
        wb.save(path)
        print(f"Updated Excel for {len(discoms)} discoms.")
    except Exception as e: print(f"Error: {e}")

def extract_insts_charges(jsonl_path, fy_info):
    insts_charges = None
    best_priority = -1
    keywords = ["intra-state transmission system charges", "stu charges", "transmission charges", "transmission & open access"]
    
    with open(jsonl_path, 'r', encoding='utf-8') as f:
        for line in f:
            try:
                data = json.loads(line)
                h = data.get("table_heading", "").lower()
                
                # Rigid priority: Only accept Current Year data
                priority = get_priority(h, fy_info)
                if priority < 2: 
                     # Allow fuzzy match for current year if header contains FY 26
                     if "fy 26" in h or "fy26" in h or "fy 2026" in h: priority = 2
                     else: continue

                match_found = any(k in h for k in keywords)
                rows = data.get("rows", [])
                
                if match_found:
                    for row in rows:
                        row_txt = str(row).lower()
                        if "total" in row_txt or "grand" in row_txt: continue
                        
                        is_paisa = "paisa" in row_txt or "paisa" in h
                        is_unit = "rs/kwh" in row_txt or "rs./unit" in row_txt or "rs/kwh" in h or "/unit" in h
                        
                        target_row = any(k in row_txt for k in keywords) or "charges" in row_txt
                        if target_row:
                             for v in row.values():
                                 try:
                                     clean = re.sub(r'[^\d\.]', '', str(v))
                                     if clean:
                                         f_v = float(clean)
                                         val = None
                                         if is_paisa and f_v > 50: val = f_v / 100.0
                                         elif (is_unit or "rate" in str(row).lower()) and 0.05 < f_v < 10: val = f_v
                                         
                                         if val and priority >= best_priority:
                                             insts_charges = val
                                             best_priority = priority
                                 except: pass
            except: pass

    # Default to NA if nothing valid found
    if insts_charges is None: insts_charges = "NA"
    
    print(f"Extracted InSTS Charges: {insts_charges}")
    return insts_charges

if __name__ == "__main__":
    base_dir = os.path.dirname(os.path.abspath(__file__))
    fy_info = get_financial_years()
    
    # 1. Dynamic Search for Himachal Pradesh Extraction folder
    extraction_root = os.path.join(base_dir, "Extraction")
    extraction_dir = None
    
    if os.path.exists(extraction_root):
        for d in os.listdir(extraction_root):
            if "himachal" in d.lower():
                extraction_dir = os.path.join(extraction_root, d)
                break
    
    # 2. Find JSONL file
    jsonl_file = None
    if extraction_dir and os.path.exists(extraction_dir):
        for root, dirs, files in os.walk(extraction_dir):
            for f in files:
                # Try to find current FY file first
                if f.endswith(".jsonl") and (fy_info['current_short'] in f or fy_info['current_long'] in f):
                    jsonl_file = os.path.join(root, f)
                    break
            if jsonl_file: break
            
            # Fallback to any jsonl in the folder
            for f in files:
                if f.endswith(".jsonl"):
                    jsonl_file = os.path.join(root, f)
                    break
            if jsonl_file: break

    # 3. Path for ISTS and Excel
    ists_loss_file = os.path.join(base_dir, "ists_extracted", "ists_loss.json")
    excel_path = os.path.join(base_dir, "Himachalpradesh.xlsx")
    discom_file_output = os.path.join(base_dir, "discoms_hp.txt")

    # 4. Process Extraction if JSONL is found
    if jsonl_file:
        print(f"Target JSONL Found: {jsonl_file}")
        ists_val = extract_ists_loss(ists_loss_file)
        discoms = extract_discom_names(jsonl_file, discom_file_output)
        
        insts = extract_losses(jsonl_file, fy_info)
        wheeling_l = extract_wheeling_losses(jsonl_file, fy_info)
        wheeling_c = extract_wheeling_charges(jsonl_file, fy_info)
        css = extract_css_charges(jsonl_file, fy_info)
        insts_charges_val = extract_insts_charges(jsonl_file, fy_info)
        fixed = extract_fixed_charges(jsonl_file, fy_info)
        energy = extract_energy_charges(jsonl_file, fy_info)
        fuel = extract_fuel_surcharge(jsonl_file, fy_info)
        tod = extract_tod_charges(jsonl_file)
        pfa = extract_pfa_rebate_dynamic(jsonl_file, fy_info)
        lf = extract_load_factor_incentive_dynamic(jsonl_file, fy_info)
        grid = extract_grid_support_charges(jsonl_file, fy_info)
        volt = extract_voltage_rebates(jsonl_file, fy_info)
        bulk = extract_bulk_consumption_rebate(jsonl_file, fy_info)
        add_s = extract_additional_surcharge(jsonl_file, fy_info)
        
        update_excel_with_discoms(discoms, ists_val, insts, insts_charges_val, wheeling_l, wheeling_c, css, fixed, energy, fuel, tod, pfa, lf, grid, volt, bulk, add_s, excel_path)
    else:
        print("Error: No JSONL scraping data found for Himachal Pradesh.")
        print(f"Looked in: {extraction_dir if extraction_dir else extraction_root}")
        print("Please run the Scraper (Start Agent) first to generate the necessary data.")

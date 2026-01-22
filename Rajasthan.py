import json
import re
import os
import openpyxl
try:
    from database.database_utils import save_tariff_row
    DB_SUCCESS = True
except ImportError:
    DB_SUCCESS = False
from datetime import datetime

def find_value_in_jsonl(jsonl_path, table_keywords, row_keywords, value_constraint=lambda x: True):
    if not jsonl_path or not os.path.exists(jsonl_path): return "NA"
    with open(jsonl_path, 'r', encoding='utf-8') as f:
        for line in f:
            try:
                data = json.loads(line)
                h = data.get("table_heading", "").lower()
                heading_match = all(k.lower() in h for k in table_keywords)
                if heading_match:
                    for row in data.get("rows", []):
                        row_txt = str(row).lower()
                        if all(k.lower() in row_txt for k in row_keywords):
                            for v in list(row.values())[::-1]:
                                if not v: continue
                                try:
                                    clean = re.sub(r'[^\d\.]', '', str(v))
                                    if clean:
                                        f_v = float(clean)
                                        if value_constraint(f_v):
                                            return str(v).strip()
                                except: pass
            except: pass
    return "NA"

def extract_transmission_charges(jsonl_path):
    # Search for PGCIL or Transmission Charges in Rajasthan
    return find_value_in_jsonl(jsonl_path, ["transmission", "charge"], ["rs/kwh"], lambda x: 0.1 <= x <= 2.0)

def extract_discom_names(jsonl_path):
    discom_names = set()
    table_keywords = ["discom", "distribution companies"] 
    ignore_keywords = [
        "particular", "s.no", "total", "year", "unit", "column", "approved", 
        "sr. no", "sr.no", "source", "rajasthan", "description", "remark",
        "proposed", "actual", "cost", "energy", "charge", "status", "report",
        "station", "plant", "capacity", "share", "state", "say", "tariff",
        "fixed", "variable", "commission", "month", "date", "category", "type",
        "consumer", "submission", "fy", "scheme", "power", "name of", "domestic", "no"
    ]
    
    with open(jsonl_path, 'r', encoding='utf-8') as f:
        for line in f:
            try:
                data = json.loads(line)
                heading = data.get("table_heading", "").lower()
                if any(k in heading for k in table_keywords):
                    for h in data.get("headers", []):
                        if not h or not isinstance(h, str): continue
                        h_clean = h.strip()
                        lower_h = h_clean.lower()
                        
                        # skip based on keywords
                        if any(k in lower_h for k in ignore_keywords): continue
                        
                        # skip based on regex
                        if re.match(r'^s[r\.]*\s*n?[o\.]*$', lower_h): continue
                        if re.match(r'^column[_\s]?\d*$', lower_h): continue
                        
                        # skip if too long or looks like number
                        if len(h_clean) > 10: continue
                        if sum(c.isdigit() for c in h_clean) > 2: continue
                        
                        discom_names.add(h_clean)
            except: pass
    
    sorted_names = sorted(list(discom_names))
    return sorted_names

def extract_losses(jsonl_path):
    insts_loss = None
    
    with open(jsonl_path, 'r', encoding='utf-8') as f:
        for line in f:
            if insts_loss: break
            try:
                data = json.loads(line)
                rows = data.get("rows", [])
                for row in rows:
                    if not row: continue
                    r_txt = str(row).lower()
                    
                    # Logic to find best value in row
                    def get_best_val(r):
                        candidates = []
                        for k, v in r.items():
                            if not v or not isinstance(v, str): continue
                            k_lower = k.lower()
                            # Clean value
                            v_clean = v.strip().replace('%', '')
                            try:
                                float(v_clean)
                                # Score candidate
                                score = 0
                                if "approved" in k_lower: score += 2
                                if re.search(r'fy\s*20\d{2}[-20]*\d\d', k_lower): score += 2
                                if "proposed" in k_lower: score += 1
                                candidates.append((v_clean, score))
                            except: pass
                        
                        candidates.sort(key=lambda x: x[1], reverse=True)
                        if candidates:
                             return f"{candidates[0][0]}%"
                        return None


                    if "intra-state transmission losses" in r_txt:
                        val = get_best_val(row)
                        if val: insts_loss = val
            except: pass
            
    # Fallbacks or cleanup
    print(f"Extracted InSTS Loss: {insts_loss}")
    return insts_loss

def extract_wheeling_losses(jsonl_path, target_discoms):
    # results: {discom_name: {voltage_level: {val: value, priority: p}}}
    results = {}
    defaults = {} # {voltage_level: {val: value, priority: p}}
    
    with open(jsonl_path, 'r', encoding='utf-8') as f:
        for line in f:
            try:
                data = json.loads(line)
                heading = data.get("table_heading", "").lower()
                rows = data.get("rows", [])
                
                # Check for relevant table types
                priority = 0
                if "apportionment of voltage-wise sales" in heading:
                    priority = 10
                elif "distribution losses" in heading or "wheeling losses" in heading:
                    priority = 1
                
                if priority == 0: continue

                for row in rows:
                    # 1. Identify Discom in Row OR use Heading
                    matched_discom = None
                    row_vals_str = [str(v).lower() for v in row.values() if v]
                    row_txt = " ".join(row_vals_str)
                    
                    for d in target_discoms:
                        if d.lower() in row_txt:
                            matched_discom = d
                            break
                    
                    # If not in row, checking heading
                    if not matched_discom:
                        for d in target_discoms:
                            if d.lower() in heading:
                                matched_discom = d
                                break
                    
                    # Extract values
                    extracted = {}
                    
                    # Strategy A: Check if voltage is in a specific column (Row sets voltage)
                    row_voltage = None
                    # Normalize keys
                    row_clean = {k.lower().strip(): v for k, v in row.items() if v}
                    
                    # Look for voltage level in "voltage level" column or similar
                    voltage_keys = [k for k in row_clean.keys() if "voltage" in k and "level" in k]
                    v_val_str = ""
                    if voltage_keys:
                        v_val_str = str(row_clean[voltage_keys[0]]).lower()
                    
                    # Extract voltage number from voltage column or just check if the row represents a voltage
                    v_match = re.search(r'(\d+)\s*kv', v_val_str)
                    if v_match:
                        row_voltage = v_match.group(1)
                    
                    if row_voltage:
                        # If we found a voltage row, we look for "losses" column
                        loss_val = None
                        for k, v in row_clean.items():
                            if "loss" in k and ("%" in k or "percent" in k or "%" in str(v)):
                                loss_val = v
                                break
                        
                        if loss_val:
                            loss_val_clean = str(loss_val).strip() # Keep % if present
                            extracted[row_voltage] = loss_val_clean
                            
                    else:
                        # Strategy B: Column Header is Voltage (Old Logic)
                        for k, v in row.items():
                            k_low = k.lower().replace(" ", "").replace("\n", "")
                            # Standard keys
                            if "11kV" in k_low or "11kv" in k_low: 
                                if "%" in str(v): extracted["11"] = v
                            if "33kV" in k_low or "33kv" in k_low: 
                                if "%" in str(v): extracted["33"] = v
                            if "132kV" in k_low or "132kv" in k_low: 
                                 if "%" in str(v): extracted["132"] = v
                            if "66kV" in k_low or "66kv" in k_low: 
                                 if "%" in str(v): extracted["66"] = v
                    
                    if extracted:
                        if matched_discom:
                            if matched_discom not in results: results[matched_discom] = {}
                            for kv, val in extracted.items():
                                existing = results[matched_discom].get(kv)
                                if not existing or priority > existing['priority']:
                                    results[matched_discom][kv] = {'val': val, 'priority': priority}
                        else:
                            # Generic / Default values
                            for kv, val in extracted.items():
                                existing = defaults.get(kv)
                                if not existing or priority > existing['priority']:
                                    defaults[kv] = {'val': val, 'priority': priority}

            except: pass
    
    # Finalize results - convert back to simple {discom: {kv: val}}
    final_results = {}
    
    # 1. Fill from extracted results
    for d, data in results.items():
        if d not in final_results: final_results[d] = {}
        for kv, item in data.items():
            final_results[d][kv] = item['val']
            
    # 2. Fill missing discoms/voltages from defaults
    simple_defaults = {k: v['val'] for k, v in defaults.items()}
    
    for d in target_discoms:
        if d not in final_results:
             final_results[d] = simple_defaults.copy()
        else:
             # Fill missing keys in existing discom
             for k, v in simple_defaults.items():
                 if k not in final_results[d]:
                     final_results[d][k] = v
                 
    print(f"Extracted Dynamic Wheeling Losses: {final_results}")
    return final_results

def extract_wheeling_charges(jsonl_path, target_discoms):
    # results: {discom_name: {voltage_level: wheeling_charge_value}}
    w_charges = {}
    # results: {discom_name: {voltage_level: transmission_charge_value}}
    t_charges = {}
    
    with open(jsonl_path, 'r', encoding='utf-8') as f:
        for line in f:
            try:
                data = json.loads(line)
                heading = data.get("table_heading", "").lower()
                
                if ("wheeling" in heading and "transmission" in heading and "cost" in heading):
                    rows = data.get("rows", [])
                    current_matched_discom = None
                    
                    for row in rows:
                        # 1. Identify/Update Discom
                        raw_discom = row.get("Discom")
                        if raw_discom:
                            raw_discom_s = str(raw_discom).strip()
                            current_matched_discom = None
                            for d in target_discoms:
                                if d.lower() in raw_discom_s.lower():
                                    current_matched_discom = d
                                    break
                        
                        if not current_matched_discom:
                            continue
                        
                        # 2. Check row type
                        row_vals_txt = " ".join([str(v).lower() for v in row.values() if v])
                        
                        # Wheeling cost
                        if "wheeling cost" in row_vals_txt and "transmission" not in row_vals_txt:
                            if current_matched_discom not in w_charges: w_charges[current_matched_discom] = {}
                            for kv in ["11kV", "33kV", "132kV"]:
                                val = row.get(kv)
                                if val:
                                    try:
                                        kv_key = kv.replace("kV", "")
                                        w_charges[current_matched_discom][kv_key] = float(val)
                                    except: pass
                        
                        # Transmission cost (InSTS Charge)
                        if "transmission cost" in row_vals_txt:
                            if current_matched_discom not in t_charges: t_charges[current_matched_discom] = {}
                            for kv in ["11kV", "33kV", "132kV"]:
                                val = row.get(kv)
                                if val:
                                    try:
                                        kv_key = kv.replace("kV", "")
                                        t_charges[current_matched_discom][kv_key] = float(val)
                                    except: pass
            except: pass
                
    print(f"Extracted Dynamic Wheeling Charges: {w_charges}")
    print(f"Extracted Dynamic Transmission (InSTS) Charges: {t_charges}")
    return w_charges, t_charges

def extract_css_charges(jsonl_path, discoms):
    # Dictionary to store CSS: {voltage_level: css_value}
    # Values seem to be same for all Discoms in Table 95
    css_values = {'11': None, '33': None, '66': None, '132': None, '220': None}
    
    with open(jsonl_path, 'r', encoding='utf-8') as f:
        for line in f:
            try:
                data = json.loads(line)
                heading = data.get("table_heading", "").lower()
                
                if "cross subsidy surcharge" in heading and re.search(r'20\d\d[-20]*\d\d', heading):
                    rows = data.get("rows", [])
                    for row in rows:
                        voltage = str(row.get("Voltage (kV)", "")).lower()
                        if not voltage:
                             # Try alternative key names
                             for k in row:
                                 if "voltage" in k.lower():
                                     voltage = str(row[k]).lower()
                                     break
                                     
                        # Priority: 1. "applicable" 2. general "css" with units
                        css_val = None
                        # First pass for "applicable" (scraped as "Applica ble CSS" often)
                        for k, v in row.items():
                            k_low = k.lower().replace(" ", "").replace("\n", "")
                            if "applicable" in k_low and "css" in k_low:
                                css_val = v
                                break
                        
                        # Fallback for general values if not found or empty
                        if not css_val:
                            for k, v in row.items():
                                if not v or not isinstance(v, str): continue
                                if "paise" in k.lower() or "rs/unit" in k.lower() or "unit" in k.lower():
                                     # but skip voltage columns
                                     if "voltage" in k.lower(): continue
                                     css_val = v
                                     break
                        
                        if css_val:
                            clean_css = re.sub(r'[^\d\.]', '', str(css_val))
                            try:
                                f_css = float(clean_css)
                                if "33" in voltage: css_values['33'] = f_css
                                elif "11" in voltage: css_values['11'] = f_css
                                elif "132" in voltage: css_values['132'] = f_css
                                elif "220" in voltage: css_values['220'] = f_css
                            except: pass
            except: pass
            
    print(f"Extracted Dynamic CSS Charges: {css_values}")
    return css_values

def extract_additional_surcharge(jsonl_path):
    # Example table 92 "Determination of Additional Surcharge for FY 2024-25"
    add_surcharge = None
    with open(jsonl_path, 'r', encoding='utf-8') as f:
        for line in f:
            try:
                data = json.loads(line)
                heading = data.get("table_heading", "").lower()
                if "additional surcharge" in heading:
                    rows = data.get("rows", [])
                    for row in rows:
                        row_vals = [str(v).lower() for v in row.values() if v]
                        row_txt = " ".join(row_vals)
                        if "per unit" in row_txt and "additional surcharge" in row_txt:
                            # Look for value (likely 0.45 or similar)
                            for v in row.values():
                                try:
                                    clean = re.sub(r'[^\d\.]', '', str(v))
                                    if clean:
                                        f_v = float(clean)
                                        if 0.05 < f_v < 10:
                                             add_surcharge = f_v
                                             break
                                except: pass
                        if add_surcharge: break
                if add_surcharge: break
            except: pass
    print(f"Extracted Dynamic Additional Surcharge: {add_surcharge}")
    return add_surcharge

def extract_tariff_charges(jsonl_path):
    # Returns {category: {voltage: charges}}
    # Focusing on Large Industrial (LP) category
    fixed_charges = {} # {discom: {voltage: val}}
    energy_charges = {}
    
    with open(jsonl_path, 'r', encoding='utf-8') as f:
        for line in f:
            try:
                data = json.loads(line)
                heading = data.get("table_heading", "").lower()
                # Target: Table 103 (LP category), 101 (Mixed Load), 102 (Small Industrial)
                if ("tariff" in heading or "charges" in heading) and ("schedule" in heading):
                    rows = data.get("rows", [])
                    category = None
                    for row in rows:
                        row_txt = str(row).lower()
                        # Detect Category
                        if "lp" in row_txt or "large industrial" in row_txt or "bulk" in row_txt:
                            # Try to extract Fixed / Energy charges from this row or following rows
                            # Often in columns labelled "Fixed" and "Energy"
                            # But specifically for HTS consumers (11kV, 33kV, etc)
                            pass
            except: pass
    
    # Static fallback logic refined to be more dynamic in structure
    # Actually, let's keep it as is for now but prepared for future dynamic extraction
    return {}, {}

def extract_pf_rebate(jsonl_path):
    return find_value_in_jsonl(jsonl_path, ["power factor"], ["rebate", "incentive"], lambda x: 0.1 <= x <= 5.0)

def extract_load_factor_incentive(jsonl_path):
    return find_value_in_jsonl(jsonl_path, ["load factor"], ["incentive", "rebate"], lambda x: 0.1 <= x <= 5.0)

def extract_grid_support_charges(jsonl_path):
    return find_value_in_jsonl(jsonl_path, ["grid support", "parallel operation"], ["charge"], lambda x: 10 <= x <= 100)

def extract_voltage_rebates(jsonl_path):
    r33 = find_value_in_jsonl(jsonl_path, ["voltage", "rebate"], ["33", "66"], lambda x: 0.1 <= x <= 5.0)
    r132 = find_value_in_jsonl(jsonl_path, ["voltage", "rebate"], ["132", "220"], lambda x: 0.1 <= x <= 5.0)
    return {'33_66': r33, '132': r132}

def extract_bulk_consumption_rebate(jsonl_path):
    return find_value_in_jsonl(jsonl_path, ["bulk", "consumption"], ["rebate"], lambda x: 0.1 <= x <= 5.0)

def extract_tod_charges(jsonl_path):
    # ToD charges are typically complex tables. Returning NA unless specific row found.
    return "NA"

def update_excel(discoms, ists_loss, insts_loss, wheeling_losses, wheeling_charges, insts_charges, css_charges, add_surcharge, pf_rebate, lf_incentive, grid_support, voltage_rebates, bulk_rebate, excel_path):
    try:
        wb = openpyxl.load_workbook(excel_path)
    except:
        wb = openpyxl.Workbook()
        wb.save(excel_path)
        wb = openpyxl.load_workbook(excel_path)
    
    sheet = wb.active
    
    headers = [
        "Financial Year", "State", "DISCOM", "ISTS Loss", "InSTS Loss",
        "Wheeling Loss - 11 kV", "Wheeling Loss - 33 kV", "Wheeling Loss - 66 kV", "Wheeling Loss - 132 kV",
        "ISTS Charges", "InSTS Charges",
        "Wheeling Charges - 11 kV", "Wheeling Charges - 33 kV", "Wheeling Charges - 66 kV", "Wheeling Charges - 132 kV",
        "Cross Subsidy Surcharge - 11 kV", "Cross Subsidy Surcharge - 33 kV", "Cross Subsidy Surcharge - 66 kV", "Cross Subsidy Surcharge - 132 kV", "Cross Subsidy Surcharge - 220 kV",
        "Additional Surcharge", "Electricity Duty", "Tax on Sale",
        "Fixed Charge - 11 kV", "Fixed Charge - 33 kV", "Fixed Charge - 66 kV", "Fixed Charge - 132 kV", "Fixed Charge - 220 kV",
        "Energy Charge - 11 kV", "Energy Charge - 33 kV", "Energy Charge - 66 kV", "Energy Charge - 132 kV", "Energy Charge - 220 kV",
        "Fuel Surcharge", "TOD Charges", "Power Factor Adjustment Rebate", "Load Factor Incentive", "Grid Support /Parrallel Operation",
        "HT ,EHV Rebate at 33/66 kV", "HT ,EHV Rebate at 132 kV and above ", "Bulk Consumption Rebate"
    ]
    
    for i, h in enumerate(headers):
        sheet.cell(row=1, column=i+1).value = h
        
    if sheet.max_row >= 3:
        sheet.delete_rows(3, sheet.max_row - 2)
        
    for idx, d in enumerate(discoms):
        row_idx = idx + 3
        
        # Populate columns
        sheet.cell(row=row_idx, column=1).value = "FY2025-26"
        sheet.cell(row=row_idx, column=2).value = "Rajasthan"
        sheet.cell(row=row_idx, column=3).value = d
        sheet.cell(row=row_idx, column=4).value = ists_loss
        sheet.cell(row=row_idx, column=5).value = insts_loss
        
        # Wheeling Loss
        d_wl = wheeling_losses.get(d, {})
        sheet.cell(row=row_idx, column=6).value = d_wl.get('11', "NA")
        sheet.cell(row=row_idx, column=7).value = d_wl.get('33', "NA")
        sheet.cell(row=row_idx, column=8).value = d_wl.get('66', "NA")
        sheet.cell(row=row_idx, column=9).value = d_wl.get('132', "NA")
        
        # ISTS Charges
        sheet.cell(row=row_idx, column=10).value = "NA"
        
        # InSTS Charges
        d_tc = insts_charges.get(d, {})
        val_tc = d_tc.get('11') or d_tc.get('33') or d_tc.get('132') or "NA"
        sheet.cell(row=row_idx, column=11).value = str(val_tc)
        
        # Wheeling Charges
        d_wc = wheeling_charges.get(d, {})
        sheet.cell(row=row_idx, column=12).value = d_wc.get('11', "NA")
        sheet.cell(row=row_idx, column=13).value = d_wc.get('33', "NA")
        sheet.cell(row=row_idx, column=14).value = d_wc.get('66', "NA")
        sheet.cell(row=row_idx, column=15).value = d_wc.get('132', "NA")
        
        # CSS
        sheet.cell(row=row_idx, column=16).value = css_charges.get('11', "NA")
        sheet.cell(row=row_idx, column=17).value = css_charges.get('33', "NA")
        sheet.cell(row=row_idx, column=18).value = css_charges.get('66', "NA")
        sheet.cell(row=row_idx, column=19).value = css_charges.get('132', "NA")
        sheet.cell(row=row_idx, column=20).value = css_charges.get('220', "NA")
        
        sheet.cell(row=row_idx, column=21).value = add_surcharge
        
        # placeholders for fixed/energy
        for c in range(22, 42):
            if sheet.cell(row=row_idx, column=c).value is None:
                sheet.cell(row=row_idx, column=c).value = "NA"

        # Update DB
        if DB_SUCCESS:
            db_data = {
                'financial_year': "FY2025-26",
                'state': 'Rajasthan',
                'discom': d,
                'ists_loss': str(ists_loss) if ists_loss else "NA",
                'insts_loss': str(insts_loss) if insts_loss else "NA",
                'wheeling_loss_11kv': d_wl.get('11', "NA"),
                'wheeling_loss_33kv': d_wl.get('33', "NA"),
                'wheeling_loss_66kv': d_wl.get('66', "NA"),
                'wheeling_loss_132kv': d_wl.get('132', "NA"),
                'ists_charges': "NA",
                'insts_charges': str(val_tc) if val_tc else "NA",
                'wheeling_charges_11kv': d_wc.get('11', "NA"),
                'wheeling_charges_33kv': d_wc.get('33', "NA"),
                'wheeling_charges_66kv': d_wc.get('66', "NA"),
                'wheeling_charges_132kv': d_wc.get('132', "NA"),
                'css_charges_11kv': css_charges.get('11', "NA"),
                'css_charges_33kv': css_charges.get('33', "NA"),
                'css_charges_66kv': css_charges.get('66', "NA"),
                'css_charges_132kv': css_charges.get('132', "NA"),
                'css_charges_220kv': css_charges.get('220', "NA"),
                'additional_surcharge': add_surcharge if add_surcharge else "NA",
                'electricity_duty': "NA",
                'tax_on_sale': "NA",
                'fixed_charge_11kv': "NA",
                'fixed_charge_33kv': "NA",
                'fixed_charge_66kv': "NA",
                'fixed_charge_132kv': "NA",
                'fixed_charge_220kv': "NA",
                'energy_charge_11kv': "NA",
                'energy_charge_33kv': "NA",
                'energy_charge_66kv': "NA",
                'energy_charge_132kv': "NA",
                'energy_charge_220kv': "NA",
                'fuel_surcharge': "NA",
                'tod_charges': "NA",
                'pf_rebate': pf_rebate if pf_rebate else "NA",
                'lf_incentive': lf_incentive if lf_incentive else "NA",
                'grid_support_parallel_op_charges': grid_support if grid_support else "NA",
                'ht_ehv_rebate_33_66kv': voltage_rebates.get('33_66', "NA"),
                'ht_ehv_rebate_132_above': voltage_rebates.get('132', "NA"),
                'bulk_rebate': bulk_rebate if bulk_rebate else "NA"
            }
            # Sanitize
            clean_db_data = {k: (str(v) if v is not None else "NA") for k, v in db_data.items()}
            save_tariff_row(clean_db_data)

    wb.save(excel_path)
    print(f"Updated {excel_path} with {len(discoms)} discoms.")

def extract_ists_loss(json_path):
    try:
        if os.path.exists(json_path):
            with open(json_path, 'r', encoding='utf-8') as f:
                data = json.load(f)
                val = data.get("All India transmission Loss (in %)", "NA")
                if "%" not in str(val): val = f"{val}%"
                return val
    except: pass
    return "NA"

if __name__ == "__main__":
    base_dir = os.path.dirname(os.path.abspath(__file__))
    
    # Dynamic JSONL search
    extraction_root = os.path.join(base_dir, "Extraction")
    jsonl_file = None
    if os.path.exists(extraction_root):
        for d in os.listdir(extraction_root):
            if "rajasthan" in d.lower() or "rajastan" in d.lower():
                input_dir = os.path.join(extraction_root, d)
                for f in os.listdir(input_dir):
                    if f.endswith(".jsonl"):
                        jsonl_file = os.path.join(input_dir, f)
                        break
            if jsonl_file: break

    excel_file = os.path.join(base_dir, "Rajasthan.xlsx")
    ists_file = os.path.join(base_dir, "ists_extracted", "ists_loss.json")
    
    if jsonl_file:
        print(f"Target JSONL: {jsonl_file}")
        
        discoms = extract_discom_names(jsonl_file)
        if not discoms: discoms = ["JVVNL", "AVVNL", "JDVVNL"]
        
        ists_l = extract_ists_loss(ists_file)
        insts_l = extract_losses(jsonl_file)
        
        wh_losses = extract_wheeling_losses(jsonl_file, discoms)
        wh_charges, insts_charges = extract_wheeling_charges(jsonl_file, discoms)
        css = extract_css_charges(jsonl_file, discoms)
        add_s = extract_additional_surcharge(jsonl_file)
        
        pf_r = extract_pf_rebate(jsonl_file)
        lf_i = extract_load_factor_incentive(jsonl_file)
        gs_c = extract_grid_support_charges(jsonl_file)
        volt_reb = extract_voltage_rebates(jsonl_file)
        bulk_reb = extract_bulk_consumption_rebate(jsonl_file)
        
        update_excel(discoms, ists_l, insts_l, wh_losses, wh_charges, insts_charges, css, add_s, pf_r, lf_i, gs_c, volt_reb, bulk_reb, excel_file)
    else:
        print("Required JSONL file not found for Rajasthan.")

import json
import re
import os
import openpyxl
import datetime
import glob
try:
    from database.database_utils import save_tariff_row
    DB_SUCCESS = True
except ImportError:
    DB_SUCCESS = False

def get_target_years():
    now = datetime.datetime.now()
    start_year = now.year if now.month >= 4 else now.year - 1
    targets = []
    for y in [start_year, start_year + 1, start_year - 1]:
        targets.extend([f"{y}-{str(y+1)[2:]}", f"{y}-{y+1}"])
    return list(set(targets))

TARGET_YEARS = get_target_years()

def extract_discom_names(jsonl_path):
    discoms = set()
    if not jsonl_path or not os.path.exists(jsonl_path): return ["NBPDCL", "SBPDCL"]
    known = ["NBPDCL", "SBPDCL"]
    with open(jsonl_path, 'r', encoding='utf-8') as f:
        for i, line in enumerate(f):
            content = line.upper()
            for k in known:
                if k in content: discoms.add(k)
            if i > 1000: break
    res = sorted(list(discoms)) if discoms else ["NBPDCL", "SBPDCL"]
    return res

def find_value_in_jsonl(jsonl_path, table_keywords, row_keywords, value_constraint=lambda x: True, is_percent=False):
    if not jsonl_path or not os.path.exists(jsonl_path): return "NA"
    
    with open(jsonl_path, 'r', encoding='utf-8') as f:
        for line in f:
            try:
                data = json.loads(line)
                h = data.get("table_heading", "").lower()
                
                heading_match = all(k.lower() in h for k in table_keywords)
                if heading_match:
                    for row in data.get("rows", []):
                        row_txt = str(row).lower()
                        if all(k.lower() in row_txt for k in row_keywords):
                            for v in row.values():
                                if not v: continue
                                try:
                                    clean = re.sub(r'[^\d\.]', '', str(v))
                                    if clean:
                                        f_v = float(clean)
                                        if value_constraint(f_v):
                                            res = str(v).strip()
                                            if is_percent and "%" not in res: res += "%"
                                            return res
                                except: pass
            except: pass
    return "NA"

def extract_ists_loss(json_path):
    try:
        if os.path.exists(json_path):
            with open(json_path, 'r', encoding='utf-8') as f:
                data = json.load(f)
                val = data.get("All India transmission Loss (in %)", "NA")
                if val != "NA" and "%" not in str(val):
                    val = f"{val}%"
                return val
    except: pass
    return "NA"

def extract_insts_loss(jsonl_path):
    # Search for Intra-state transmission loss
    return find_value_in_jsonl(jsonl_path, ["loss"], ["intra-state", "transmission"], lambda x: 2.0 <= x <= 5.0, True)

def extract_wheeling_losses(jsonl_path, discom_names):
    losses = {name: {'11': "NA", '33': "NA", '66': "NA", '132': "NA"} for name in discom_names}
    if not jsonl_path or not os.path.exists(jsonl_path): return losses

    with open(jsonl_path, 'r', encoding='utf-8') as f:
        for line in f:
            try:
                data = json.loads(line)
                h = data.get("table_heading", "").lower()
                if "distribution loss" in h:
                    target = "GENERIC"
                    for name in discom_names:
                        if name.lower() in h:
                            target = name; break
                    
                    for row in data.get("rows", []):
                        row_txt = str(row).lower()
                        if "distribution loss" in row_txt:
                            val = None
                            for v in list(row.values())[::-1]:
                                if v and "%" in str(v):
                                    val = str(v).strip(); break
                            if val:
                                if target == "GENERIC":
                                    for n in discom_names: losses[n]['11'] = losses[n]['33'] = val
                                else:
                                    losses[target]['11'] = losses[target]['33'] = val
            except: pass
    return losses

def extract_table_components(jsonl_path, table_query, voltage_keywords):
    results = {v: "NA" for v in voltage_keywords}
    if not jsonl_path or not os.path.exists(jsonl_path): return results
    
    with open(jsonl_path, 'r', encoding='utf-8') as f:
        for line in f:
            try:
                data = json.loads(line)
                h = data.get("table_heading", "").lower()
                if all(k.lower() in h for k in table_query):
                    headers = [str(h).lower() for h in data.get("headers", [])]
                    for row in data.get("rows", []):
                        row_txt = str(row).lower()
                        # Identify voltage from row context or voltage header
                        matched_v = None
                        for v in voltage_keywords:
                            if f"{v} kv" in row_txt or f"{v}kv" in row_txt:
                                matched_v = v; break
                        
                        if matched_v:
                            # Try to find a numeric value that looks like a charge
                            for k, val in row.items():
                                if not val: continue
                                try:
                                    clean = re.sub(r'[^\d\.]', '', str(val))
                                    if clean and 0.1 <= float(clean) < 10:
                                        results[matched_v] = clean
                                        # Note: this might need more specific column logic
                                except: pass
            except: pass
    return results

def extract_css_charges(jsonl_path):
    # Bihar CSS table often has voltage and CSS in the same row
    css = {v: "NA" for v in ['11', '33', '66', '132', '220']}
    with open(jsonl_path, 'r', encoding='utf-8') as f:
        for line in f:
            try:
                data = json.loads(line)
                h = data.get("table_heading", "").lower()
                if "cross subsidy" in h and "surcharge" in h:
                    for row in data.get("rows", []):
                        row_txt = str(row).lower()
                        # Check voltage
                        mv = None
                        for v in ['220', '132', '33', '11']:
                            if f"{v} kv" in row_txt or f"{v}kv" in row_txt: mv = v; break
                        
                        if mv:
                            # Usually CSS is the last column
                            best_val = "NA"
                            for v in list(row.values())[::-1]:
                                try:
                                    clean = re.sub(r'[^\d\.]', '', str(v))
                                    if clean and 0.5 < float(clean) < 5.0:
                                        best_val = clean; break
                                except: pass
                            css[mv] = best_val
            except: pass
    return css

def extract_wheeling_charges(jsonl_path):
    # Wheeling charges in Bihar are often found in the same table as CSS or a dedicated ARR table
    w = {v: "NA" for v in ['11', '33', '66', '132']}
    # First search for dedicated table
    val = find_value_in_jsonl(jsonl_path, ["wheeling", "charge"], ["wheeling", "charge"], lambda x: 0.1 <= x <= 3.0)
    if val != "NA":
        for k in w: w[k] = val
    
    # Then refine from CSS table if possible
    with open(jsonl_path, 'r', encoding='utf-8') as f:
        for line in f:
            try:
                data = json.loads(line)
                h = data.get("table_heading", "").lower()
                if "cross subsidy" in h and "surcharge" in h:
                    for row in data.get("rows", []):
                        row_txt = str(row).lower()
                        mv = None
                        for v in ['132', '33', '11']:
                            if f"{v} kv" in row_txt or f"{v}kv" in row_txt: mv = v; break
                        if mv:
                            for k, v in row.items():
                                if k.lower() in [f"{mv} kv", f"{mv}kv"]:
                                    try:
                                        clean = re.sub(r'[^\d\.]', '', str(v))
                                        if clean and 0.1 <= float(clean) < 3.0:
                                            w[mv] = clean
                                    except: pass
            except: pass
    return w

def extract_fixed_charges(jsonl_path):
    fixed = {v: "NA" for v in ['11', '33', '66', '132', '220']}
    if not jsonl_path or not os.path.exists(jsonl_path): return fixed
    with open(jsonl_path, 'r', encoding='utf-8') as f:
        for line in f:
            try:
                data = json.loads(line)
                rows = data.get("rows", [])
                if not rows: continue
                # Look for HTS categories
                for row in rows:
                    cat = str(row.get("Existing Category", row.get("Consumer Category", ""))).lower()
                    if "hts" in cat or "htis" in cat:
                        mv = None
                        if " i " in f" {cat} " or "hts-i" in cat: mv = "11"
                        elif " ii" in f" {cat}" or "hts-ii" in cat: mv = "33"
                        elif " iii" in f" {cat}" or "hts-iii" in cat: mv = "132"
                        elif " iv" in f" {cat}" or "hts-iv" in cat: mv = "220"
                        
                        if mv:
                            for k, v in row.items():
                                if "fixed" in k.lower() or "demand" in k.lower():
                                    try:
                                        clean = re.sub(r'[^\d\.]', '', str(v))
                                        if clean and 100 < float(clean) < 1000:
                                            fixed[mv] = clean
                                    except: pass
            except: pass
    return fixed

def extract_energy_charges(jsonl_path):
    energy = {v: "NA" for v in ['11', '33', '66', '132', '220']}
    with open(jsonl_path, 'r', encoding='utf-8') as f:
        for line in f:
            try:
                data = json.loads(line)
                for row in data.get("rows", []):
                    cat = str(row.get("Existing Category", row.get("Consumer Category", ""))).lower()
                    if "hts" in cat or "htis" in cat:
                        mv = None
                        if "-i" in cat or " i " in f" {cat} ": mv = "11"
                        elif "-ii" in cat or " ii" in f" {cat}": mv = "33"
                        elif "-iii" in cat or " iii" in f" {cat}": mv = "132"
                        elif "-iv" in cat or " iv" in f" {cat}": mv = "220"
                        
                        if mv:
                            for k, v in row.items():
                                if "energy" in k.lower():
                                    try:
                                        clean = re.sub(r'[^\d\.]', '', str(v))
                                        if clean and 4.0 <= float(clean) < 15.0:
                                            energy[mv] = clean
                                    except: pass
            except: pass
    return energy

def extract_fuel_surcharge(jsonl_path):
    return find_value_in_jsonl(jsonl_path, ["fuel"], ["fuel", "surcharge"], lambda x: 0 < x < 5)

def extract_additional_surcharge(jsonl_path):
    return find_value_in_jsonl(jsonl_path, ["additional surcharge"], ["additional surcharge"], lambda x: 0.5 < x < 5.0)

def update_excel_with_discoms(discoms, ists_loss, insts_loss, wheeling_losses, wheeling_charges, css_charges, fixed_charges, energy_charges, fuel_surcharge, add_surcharge, excel_path):
    try:
        wb = openpyxl.load_workbook(excel_path)
        sheet = wb.active
        start_row = 3
        if sheet.max_row >= start_row:
             sheet.delete_rows(start_row, sheet.max_row - start_row + 1)
        
        for i, discom in enumerate(discoms):
            row = start_row + i
            def sv(col, val):
                sheet.cell(row=row, column=col).value = str(val) if val not in [None, "NA"] else "NA"

            sv(1, "Bihar")
            sv(3, discom)
            sv(4, ists_loss)
            sv(5, insts_loss)
            
            wl = wheeling_losses.get(discom, wheeling_losses.get('GENERIC', {}))
            sv(6, wl.get('11'))
            sv(7, wl.get('33'))
            sv(8, wl.get('66'))
            sv(9, wl.get('132'))
            
            sv(11, find_value_in_jsonl(jsonl_file, ["transmission"], ["intra-state", "charge"], lambda x: 0.1 <= x <= 1.0)) # InSTS Charge
            
            sv(12, wheeling_charges.get('11'))
            sv(13, wheeling_charges.get('33'))
            sv(14, wheeling_charges.get('66'))
            sv(15, wheeling_charges.get('132'))
            
            sv(16, css_charges.get('11'))
            sv(17, css_charges.get('33'))
            sv(18, css_charges.get('66'))
            sv(19, css_charges.get('132'))
            sv(20, css_charges.get('220'))
            
            sv(21, add_surcharge)
            
            sv(24, fixed_charges.get('11'))
            sv(25, fixed_charges.get('33'))
            sv(26, fixed_charges.get('66'))
            sv(27, fixed_charges.get('132'))
            sv(28, fixed_charges.get('220'))
            
            sv(29, energy_charges.get('11'))
            sv(30, energy_charges.get('33'))
            sv(31, energy_charges.get('66'))
            sv(32, energy_charges.get('132'))
            sv(33, energy_charges.get('220'))
            
            sv(34, fuel_surcharge)
            
            if DB_SUCCESS:
                db_data = {
                    'financial_year': "FY2025-26",
                    'state': 'Bihar',
                    'discom': discom,
                    'ists_loss': str(ists_loss),
                    'insts_loss': str(insts_loss),
                    'wheeling_loss_11kv': wl.get('11', "NA"),
                    'wheeling_loss_33kv': wl.get('33', "NA"),
                    'wheeling_loss_66kv': wl.get('66', "NA"),
                    'wheeling_loss_132kv': wl.get('132', "NA"),
                    'ists_charges': "NA",
                    'insts_charges': sheet.cell(row=row, column=11).value or "NA",
                    'wheeling_charges_11kv': wheeling_charges.get('11', "NA"),
                    'wheeling_charges_33kv': wheeling_charges.get('33', "NA"),
                    'wheeling_charges_66kv': wheeling_charges.get('66', "NA"),
                    'wheeling_charges_132kv': wheeling_charges.get('132', "NA"),
                    'css_charges_11kv': css_charges.get('11', "NA"),
                    'css_charges_33kv': css_charges.get('33', "NA"),
                    'css_charges_66kv': css_charges.get('66', "NA"),
                    'css_charges_132kv': css_charges.get('132', "NA"),
                    'css_charges_220kv': css_charges.get('220', "NA"),
                    'additional_surcharge': add_surcharge,
                    'electricity_duty': "NA",
                    'tax_on_sale': "NA",
                    'fixed_charge_11kv': fixed_charges.get('11', "NA"),
                    'fixed_charge_33kv': fixed_charges.get('33', "NA"),
                    'fixed_charge_66kv': fixed_charges.get('66', "NA"),
                    'fixed_charge_132kv': fixed_charges.get('132', "NA"),
                    'fixed_charge_220kv': fixed_charges.get('220', "NA"),
                    'energy_charge_11kv': energy_charges.get('11', "NA"),
                    'energy_charge_33kv': energy_charges.get('33', "NA"),
                    'energy_charge_66kv': energy_charges.get('66', "NA"),
                    'energy_charge_132kv': energy_charges.get('132', "NA"),
                    'energy_charge_220kv': energy_charges.get('220', "NA"),
                    'fuel_surcharge': fuel_surcharge,
                    'tod_charges': "NA",
                    'pf_rebate': "NA",
                    'lf_incentive': "NA",
                    'grid_support_parallel_op_charges': "NA",
                    'ht_ehv_rebate_33_66kv': "NA",
                    'ht_ehv_rebate_132_above': "NA",
                    'bulk_rebate': "NA"
                }
                save_tariff_row({k: (str(v) if v else "NA") for k, v in db_data.items()})
        
        wb.save(excel_path)
    except Exception as e:
        print(f"Error: {e}")

if __name__ == "__main__":
    base_dir = os.path.dirname(os.path.abspath(__file__))
    extraction_root = os.path.join(base_dir, "Extraction")
    extraction_dir = None
    if os.path.exists(extraction_root):
        for d in os.listdir(extraction_root):
            if "bihar" in d.lower(): extraction_dir = os.path.join(extraction_root, d); break
    
    if not extraction_dir: exit(1)
    jsonl_files = glob.glob(os.path.join(extraction_dir, "*.jsonl"))
    if not jsonl_files: exit(1)
    
    jsonl_file = jsonl_files[0]
    excel_file = os.path.join(base_dir, "bihar.xlsx")
    ists_loss_file = os.path.join(base_dir, "ists_extracted", "ists_loss.json")
    
    ists_val = extract_ists_loss(ists_loss_file)
    print(f"Extracted ISTS Loss: {ists_val}")
    discoms = extract_discom_names(jsonl_file)
    insts = extract_insts_loss(jsonl_file)
    wheeling_l = extract_wheeling_losses(jsonl_file, discoms)
    wheeling_c = extract_wheeling_charges(jsonl_file)
    css = extract_css_charges(jsonl_file)
    fixed = extract_fixed_charges(jsonl_file)
    energy = extract_energy_charges(jsonl_file)
    fuel = extract_fuel_surcharge(jsonl_file)
    add_s = extract_additional_surcharge(jsonl_file)

    update_excel_with_discoms(discoms, ists_val, insts, wheeling_l, wheeling_c, css, fixed, energy, fuel, add_s, excel_file)
    print(f"Successfully updated {excel_file}")

import os
import json
import glob
import openpyxl
import re

def clean_year(y_str):
    # returns 2023 for "FY 2023-24"
    if not y_str: return 0
    m = re.search(r"(\d{4})", y_str)
    return int(m.group(1)) if m else 0


def get_discom_name_from_json(json_path):
    keywords = ["discom", "discom name"]
    candidate_discom = "NA"
    
    try:
        with open(json_path, 'r', encoding='utf-8') as f:
            for line in f:
                data = json.loads(line)
                
                # Check directly in keys of the dictionary (if any structure matches)
                for key, value in data.items():
                    if isinstance(key, str) and key.lower() in keywords:
                        return value
                
                # Check in 'headers' if it's a table
                if "headers" in data and isinstance(data["headers"], list):
                    header_map = {}
                    for idx, h in enumerate(data["headers"]):
                        if h and isinstance(h, str):
                            header_map[h.lower()] = idx
                            
                    # Find if any keyword is in headers
                    found_header_idx = -1
                    found_header_key = None
                    for kw in keywords:
                        for h_lower, idx in header_map.items():
                            if kw == h_lower:
                                found_header_idx = idx
                                found_header_key = h_lower
                                break
                        if found_header_idx != -1:
                            break
                    
                    # If header found, extract value from first row
                    if found_header_idx != -1 and "rows" in data and len(data["rows"]) > 0:
                        row = data["rows"][0]
                        for r_key, r_val in row.items():
                            if r_key.lower() == found_header_key:
                                return r_val
                
                # Fallback: Scan rows for Discom definition (e.g. in Abbreviations)
                # Look for "Distribution Company Limited" or "State Power Distribution Company"
                if "rows" in data and len(data["rows"]) > 0:
                    for row in data["rows"]:
                        # Convert all values to string
                        vals = [str(v) for v in row.values() if v]
                        for v in vals:
                            v_low = v.lower()
                            if "distribution company limited" in v_low or "state power distribution company" in v_low:
                                # Found a description content. Look for the abbreviation/name in the same row.
                                # The name is likely short (e.g. CSPDCL) and not the description itself.
                                for pot_name in vals:
                                    if pot_name != v and 2 < len(pot_name) < 20:
                                        # Avoid "DISCOM" if possible, unless it's the only one.
                                        # But usually "DISCOM" maps to "Distribution Company", not "State Power..."
                                        candidate_discom = pot_name

    except Exception as e:
        print(f"Error reading JSON: {e}")
    
    return candidate_discom

def get_financial_year(json_path):
    # logic to find the most recent/present financial year in headers
    years = set()
    year_pattern = re.compile(r"FY\s?(\d{4}-\d{2})", re.IGNORECASE)
    
    try:
        with open(json_path, 'r', encoding='utf-8') as f:
            for line in f:
                data = json.loads(line)
                if "headers" in data and isinstance(data["headers"], list):
                    for h in data["headers"]:
                        if h and isinstance(h, str):
                            match = year_pattern.search(h)
                            if match:
                                years.add(match.group(1))
    except Exception as e:
        print(f"Error reading JSON for year: {e}")
    
    if not years:
        return None
        
    # Sort years and pick the latest. Format is YYYY-YY, so string sort works generally (2025-26 > 2024-25)
    sorted_years = sorted(list(years), reverse=True)
    return f"FY {sorted_years[0]}"

def get_insts_loss(json_path, target_year):
    # Keywords prioritizing %
    keywords = [
        "intra-state transmission system loss", 
        "stu transmission loss",
        "stu loss",
        "transmission loss"
    ]
    
    target_year_clean = target_year.lower().replace(" ", "") if target_year else ""
    
    candidates = [] # List of (year_val, priority, value)

    try:
        with open(json_path, 'r', encoding='utf-8') as f:
            for line in f:
                data = json.loads(line)
                
                if "rows" in data and len(data["rows"]) > 0:
                    headers = []
                    if "headers" in data and isinstance(data["headers"], list):
                        headers = [str(h) for h in data["headers"] if h]
                    
                    headers_clean = [h.lower().replace(" ", "") for h in headers]
                    
                    # Identify year columns in this table
                    year_cols = {} # index -> year_int
                    
                    for idx, h in enumerate(headers):
                        h_low = h.lower()
                        # Explicit year in header
                        y_match = re.search(r"FY\s?(\d{4}-\d{2})", h, re.IGNORECASE)
                        if y_match:
                             year_cols[idx] = clean_year(y_match.group(0))
                        elif target_year_clean and target_year_clean in headers_clean[idx]:
                             year_cols[idx] = clean_year(target_year)
                        
                        # Fallback: specific keywords or "column" (misaligned) imply target year if no year found
                        if idx not in year_cols and target_year:
                             if "approved" in h_low or "petition" in h_low or "projected" in h_low or "estimate" in h_low or "proposed" in h_low or "column" in h_low:
                                 year_cols[idx] = clean_year(target_year)

                    for row in data["rows"]:
                        # Find which key contains the keyword
                        keyword_found = False
                        found_kw = ""
                        full_key_text = ""
                        for k, v in row.items():
                            if isinstance(v, str):
                                for kw in keywords:
                                    if kw in v.lower():
                                        keyword_found = True
                                        found_kw = kw
                                        full_key_text = v.lower()
                                        break
                                if keyword_found: 
                                    break
                        
                        if keyword_found:
                            # Now retrieve values for year columns
                            # Iterate explicit year columns found
                            for col_idx, y_val in year_cols.items():
                                if col_idx < len(headers):
                                    h = headers[col_idx]
                                    if h in row:
                                        val = row[h]
                                        if val and isinstance(val, str):
                                            priority = 0
                                            
                                            # Intra/STU priority
                                            if "intra" in full_key_text or "stu" in full_key_text:
                                                priority = 3
                                            elif "inter" in full_key_text or "ists" in full_key_text:
                                                priority = -2

                                            # If keyword has (%), value might be number. 
                                            if "%" in val or "(%)" in found_kw or "loss" in found_kw:
                                                 if priority != -2: # Don't boost inter
                                                     if priority < 2: priority = 2
                                            
                                            # MU check
                                            if "(mu)" in full_key_text or " mu" in full_key_text:
                                                priority = -1
                                                
                                            # Clean value to check if number
                                            v_num = re.sub(r"[^\d\.]", "", val)
                                            if v_num and len(v_num) > 0:
                                                 candidates.append((y_val, priority, val, col_idx))

    except Exception as e:
        print(f"Error reading JSON for Insts: {e}")
        
    # Process candidates
    # Sort by year descending, then priority descending, then Column Index descending (prefer Approved/Rightmost)
    candidates.sort(key=lambda x: (x[0], x[1], x[3]), reverse=True)
    
    # 1. Look for target year
    t_year_val = clean_year(target_year) if target_year else 0
    if t_year_val:
        for y, p, v, idx in candidates:
            if y == t_year_val:
                return v
    
    return "NA"

    return "NA"

def get_wheeling_loss(json_path, target_year):
    keywords = [
        "wheeling loss", 
        "discom loss", 
        "distribution loss", 
        "voltage wise loss"
    ]
    
    target_year_clean = target_year.lower().replace(" ", "") if target_year else ""
    
    # helper to clean year string to Comparable int or string
    def clean_year(y_str):
        m = re.search(r"(\d{4})", y_str)
        return int(m.group(1)) if m else 0

    # Dictionary to store results: "11": val, "33": val, "66": val, "132": val
    voltage_losses = {
        "11": "NA",
        "33": "NA",
        "66": "NA",
        "132": "NA"
    }
    voltage_years = {
        "11": 0,
        "33": 0,
        "66": 0,
        "132": 0
    }
    
    # Store general distribution loss if specific voltage not found
    general_dist_loss = "NA"
    general_candidates = [] # List of (year, priority, val)

    try:
        with open(json_path, 'r', encoding='utf-8') as f:
            for line in f:
                data = json.loads(line)
                
                if "rows" in data and len(data["rows"]) > 0:
                    headers = []
                    if "headers" in data and isinstance(data["headers"], list):
                        headers = [str(h) for h in data["headers"] if h]
                    
                    headers_clean = [h.lower().replace(" ", "") for h in headers]
                    
                    # Identify year columns in this table
                    year_cols = {} # index -> year_int
                    
                    table_year = 0
                    if "table_heading" in data and isinstance(data["table_heading"], str):
                        m = re.search(r"FY\s?(\d{4}-\d{2})", data["table_heading"], re.IGNORECASE)
                        if m:
                            table_year = clean_year(m.group(0))
                    
                    t_year_val = clean_year(target_year) if target_year else 0
                    
                    for idx, h in enumerate(headers):
                        h_low = h.lower()
                        y_match = re.search(r"FY\s?(\d{4}-\d{2})", h, re.IGNORECASE)
                        if y_match:
                             year_cols[idx] = clean_year(y_match.group(0))
                        elif target_year_clean and target_year_clean in headers_clean[idx]:
                             year_cols[idx] = t_year_val
                        elif table_year > 0:
                             # If we know the table year, map generic columns to it
                             if "approved" in h_low or "petition" in h_low or "tariff order" in h_low or "true-up" in h_low or "projected" in h_low:
                                 year_cols[idx] = table_year
                        
                        # Fallback
                        if idx not in year_cols and target_year:
                             # Only use fallback if table_year matches target or is unknown
                             if table_year == 0 or table_year == t_year_val:
                                 if "approved" in h_low or "petition" in h_low or "projected" in h_low or "estimate" in h_low or "proposed" in h_low or "column" in h_low:
                                     year_cols[idx] = t_year_val

                    # Check if table represents wheeling/distribution loss
                    table_relevant = False
                    if "table_heading" in data and isinstance(data["table_heading"], str):
                        for kw in keywords:
                            if kw in data["table_heading"].lower():
                                table_relevant = True
                                break
                    
                    for row in data["rows"]:
                        # Check context in row keys/values
                        row_text = " ".join([str(v).lower() for v in row.values() if v])
                        
                        # Identify voltage level(s)
                        v_levels = []
                        
                        if "below 33 kv" in row_text:
                            v_levels = ["11", "33"]
                        else:
                            if "11 kv" in row_text or "11kv" in row_text: v_levels.append("11")
                            if "33 kv" in row_text or "33kv" in row_text: v_levels.append("33")
                            if "66 kv" in row_text or "66kv" in row_text: v_levels.append("66")
                            if "132 kv" in row_text or "132kv" in row_text: v_levels.append("132")
                        
                        # Check keywords
                        row_relevant = False
                        for kw in keywords:
                            if kw in row_text:
                                row_relevant = True
                                break
                        
                        # If general distribution loss row (no specific voltage)
                        is_general = False
                        if not v_levels and ("distribution loss" in row_text or "energy loss" in row_text):
                             is_general = True
                        
                        full_key_text = row_text

                        if (table_relevant or row_relevant or is_general):
                            # Extract value
                            candidates = []
                            
                            for col_idx, y_val in year_cols.items():
                                # ... existing loop ...
                                if col_idx < len(headers):
                                    h = headers[col_idx]
                                    if h in row:
                                        val = row[h]
                                        if val and isinstance(val, str):
                                            # Clean value
                                            v_num = re.sub(r"[^\d\.]", "", val)
                                            if v_num and len(v_num) > 0:
                                                f_val = float(v_num)
                                                priority = 0
                                                
                                                if "(%)" in full_key_text or " %" in full_key_text:
                                                    priority = 2
                                                elif "(mu)" in full_key_text or " mu" in full_key_text:
                                                    priority = -1
                                                
                                                if "%" in val:
                                                    priority = 3
                                                
                                                if f_val > 100 and priority < 2:
                                                    priority = -1
                                                    
                                                candidates.append((y_val, priority, val, col_idx))
                            
                            candidates.sort(key=lambda x: (x[0], x[1], x[3]), reverse=True)
                            
                            found_val = "NA"
                            found_year = 0
                            if candidates:
                                # Try target year
                                for y, p, v, idx in candidates:
                                    if y == t_year_val:
                                        found_val = v
                                        found_year = y
                                        break
                                # Fallback only if strictly needed?
                                # If we skip fallback here, we avoid 2023 values appearing as "found"
                            
                            if found_val == "NA" and candidates and not t_year_val:
                                 # If no target year specified, take latest
                                 found_val = candidates[0][2]
                                 found_year = candidates[0][0]

                            if found_val != "NA":
                                # Update logic with year tracking
                                if v_levels:
                                    for v in v_levels:
                                        curr_val = voltage_losses[v]
                                        curr_year = voltage_years[v]
                                        
                                        # Update if found year is better (target matched) or newer
                                        should_update = False
                                        if found_year == t_year_val and curr_year != t_year_val:
                                            should_update = True
                                        elif found_year > curr_year:
                                            should_update = True
                                        elif found_year == curr_year and curr_val == "NA":
                                            should_update = True
                                        elif found_year == curr_year:
                                             # Same year, trust later occurrence or prefer %
                                             should_update = True
                                             # But if found is non-% and curr is %, keep curr?
                                             if "%" in str(curr_val) and "%" not in str(found_val):
                                                 should_update = False
                                        
                                        if should_update:
                                            # print(f"DEBUG: Updating {v} from {curr_val} to {found_val}")
                                            # with open("debug_update.txt", "a", encoding="utf-8") as df:
                                            #     df.write(f"DEBUG: Updating {v} from {curr_val} to {found_val}\n")
                                            voltage_losses[v] = found_val
                                            voltage_years[v] = found_year
                                            
                                elif is_general:
                                    # Add best candidate from this row to general list
                                    # We take the top candidate for the target year
                                    # (Year, Priority, Val, Index)
                                    # Or just add all target year candidates?
                                    # Let's add the 'found_val' if it came from target year
                                    
                                    # Re-find best cand tuple
                                    best_cand = None
                                    for c in candidates:
                                        if c[2] == found_val: best_cand = c; break
                                    
                                    if best_cand:
                                        general_candidates.append(best_cand)

    except Exception as e:
        print(f"Error reading JSON for Wheeling: {e}")
        
    # Use general loss if specific not found
    if general_candidates:
        # Sort generic candidates: Year desc, Priority desc, Index desc
        general_candidates.sort(key=lambda x: (x[0], x[1], x[3]), reverse=True)
        # Filter for target year
        t_cands = [x for x in general_candidates if x[0] == t_year_val]
        if t_cands:
            general_dist_loss = t_cands[0][2]

    if general_dist_loss != "NA":
        voltage_losses["11"] = general_dist_loss
        voltage_losses["33"] = general_dist_loss
        # Keep 66/132 as NA or whatever was found specific
        
    return voltage_losses

def get_insts_charges(json_path, target_year):
    insts_charges = "NA"
    keywords = ["transmission charge", "transmission tariff", "stu charge", "stu tariff", "open access charge"]
    units = ["rs./kwh", "paise/kwh", "rs/unit", "rs./unit", "rs/kwh"]
    
    t_year_val = clean_year(target_year) if target_year else 0
    target_year_clean = target_year.lower().replace(" ", "") if target_year else ""

    candidates = [] # (year, priority, val)

    try:
        with open(json_path, 'r', encoding='utf-8') as f:
            for line in f:
                data = json.loads(line)
                if "rows" in data and len(data["rows"]) > 0:
                    headers = [str(h) for h in data.get("headers", []) if h]
                    headers_clean = [h.lower().replace(" ", "") for h in headers]
                    
                    year_cols = {}
                    table_year = 0
                    if "table_heading" in data and isinstance(data["table_heading"], str):
                        m = re.search(r"FY\s?(\d{4}-\d{2})", data["table_heading"], re.IGNORECASE)
                        if m: table_year = clean_year(m.group(0))

                    for idx, h in enumerate(headers):
                        h_low = h.lower()
                        y_match = re.search(r"FY\s?(\d{4}-\d{2})", h, re.IGNORECASE)
                        if y_match:
                             year_cols[idx] = clean_year(y_match.group(0))
                        elif target_year_clean and target_year_clean in headers_clean[idx]:
                             year_cols[idx] = t_year_val
                        elif table_year > 0:
                             if "approved" in h_low or "petition" in h_low or "tariff order" in h_low or "true-up" in h_low:
                                 year_cols[idx] = table_year
                        if idx not in year_cols and target_year:
                             if table_year == 0 or table_year == t_year_val:
                                 if "approved" in h_low or "petition" in h_low or "projected" in h_low or "estimate" in h_low or "proposed" in h_low or "column" in h_low:
                                     year_cols[idx] = t_year_val

                    for row in data["rows"]:
                        row_text = " ".join([str(v).lower() for v in row.values() if v])
                        
                        # Filter out power purchase
                        if "power purchase" in row_text: continue
                        
                        kw_match = False
                        for k in keywords:
                            if k in row_text: kw_match = True; break
                        
                        unit_match = False
                        for u in units:
                            if u in row_text: unit_match = True; break
                            
                        # Also check column headers for units?
                        # simplified for now, usually unit is in row text for these tables
                        
                        if kw_match and unit_match:
                            for col_idx, y_val in year_cols.items():
                                if col_idx < len(headers):
                                    h = headers[col_idx]
                                    if h in row:
                                        val = row[h]
                                        if val and isinstance(val, str):
                                            # Clean
                                            v_clean = val.strip()
                                            # Heuristic: Value should be small (Rs/unit), e.g. < 10
                                            v_num = re.sub(r"[^\d\.]", "", v_clean)
                                            if v_num:
                                                f_val = float(v_num)
                                                if f_val < 50: # Rs/kWh is usually small
                                                    priority = 1
                                                    if "short-term" in row_text or "short term" in row_text:
                                                        priority = 3
                                                    elif "transmission charge" in row_text:
                                                        priority = 2
                                                    
                                                    candidates.append((y_val, priority, v_clean))
        
        # Sort candidates
        candidates.sort(key=lambda x: (x[0], x[1]), reverse=True)
        
        # Pick best for target year
        found_val = "NA"
        for y, p, v in candidates:
            if y == t_year_val:
                found_val = v
                break
        
        # If not found target year, stick to NA or latest?
        # User said "year should be present" implies target year preference.
        if found_val != "NA":
            insts_charges = found_val
            
    except Exception as e:
        print(f"Error reading InSTS Charges: {e}")
        
    return insts_charges

def get_wheeling_charges(json_path, target_year):
    charges = {
        "11": "NA", "33": "NA", "66": "NA", "132": "NA"
    }
    
    # Keywords for wheeling charges
    # Note: Sometimes it's listed under "Distribution Charges" or "Open Access Charges"
    keywords = [
        "wheeling charge", 
        "distribution charge", 
        "wheeling tariff",
        "distribution tariff",
        "open access charge",
        "network charge"
    ]
    
    t_year_val = clean_year(target_year) if target_year else 0
    target_year_clean = target_year.lower().replace(" ", "") if target_year else ""

    candidates = [] # (year, priority, val, voltage)

    try:
        with open(json_path, 'r', encoding='utf-8') as f:
            for line in f:
                data = json.loads(line)
                
                heading = data.get("table_heading", "").lower()
                
                # Determine table context
                table_relevant = False
                for kw in keywords:
                    if kw in heading:
                        table_relevant = True
                        break
                
                is_paise = False
                if "paise" in heading: is_paise = True
                
                if "rows" in data and len(data["rows"]) > 0:
                    headers = [str(h) for h in data.get("headers", []) if h]
                    headers_clean = [h.lower().replace(" ", "") for h in headers]
                    
                    year_cols = {}
                    table_year = 0
                    m = re.search(r"FY\s?(\d{4}-\d{2})", heading, re.IGNORECASE)
                    if m: table_year = clean_year(m.group(0))

                    for idx, h in enumerate(headers):
                        h_low = h.lower()
                        y_match = re.search(r"FY\s?(\d{4}-\d{2})", h, re.IGNORECASE)
                        if y_match:
                             year_cols[idx] = clean_year(y_match.group(0))
                        elif target_year_clean and target_year_clean in headers_clean[idx]:
                             year_cols[idx] = t_year_val
                        elif table_year > 0:
                             if "approved" in h_low or "charge" in h_low or "tariff" in h_low or "rate" in h_low:
                                 year_cols[idx] = table_year
                        
                        # Fallback for target year if not found
                        if idx not in year_cols and target_year:
                             if table_year == 0 or table_year == t_year_val:
                                  if "approved" in h_low or "petition" in h_low or "projected" in h_low or "proposed" in h_low or "myt" in h_low:
                                      year_cols[idx] = t_year_val

                    for row in data["rows"]:
                        row_text = " " .join([str(v).lower() for v in row.values() if v])
                        
                        row_relevant = False
                        for kw in keywords:
                            if kw in row_text:
                                row_relevant = True
                                break
                        
                        if not (table_relevant or row_relevant):
                            continue
                            
                        # Voltage identification
                        v_level = None
                        if "11 kv" in row_text or "11kv" in row_text: v_level = "11"
                        elif "33 kv" in row_text or "33kv" in row_text: v_level = "33"
                        elif "66 kv" in row_text or "66kv" in row_text: v_level = "66"
                        elif "132 kv" in row_text or "132kv" in row_text: v_level = "132"
                        
                        # Check row for paise
                        row_is_paise = is_paise
                        if "paise" in row_text: row_is_paise = True
                        
                        for col_idx, y_val in year_cols.items():
                             if col_idx < len(headers):
                                h = headers[col_idx]
                                if h in row:
                                    val = row[h]
                                    if val and isinstance(val, str):
                                        # Clean
                                        v_clean = re.sub(r"[^\d\.]", "", val)
                                        if v_clean:
                                            try:
                                                f_val = float(v_clean)
                                                if row_is_paise:
                                                    f_val = f_val / 100.0
                                                
                                                # Heuristic: Wheeling charges < 20 INR/kWh
                                                # Also avoid extracting "33" from "33 kV" as a value if possible (usually values are decimal like 0.25)
                                                if f_val < 20: 
                                                    priority = 1
                                                    if v_level: priority = 2
                                                    if "approved" in h.lower(): priority += 1
                                                    
                                                    candidates.append((y_val, priority, f_val, v_level))
                                            except: pass

    except Exception as e:
        print(f"Error extracting wheeling charges: {e}")

    # Process candidates
    candidates.sort(key=lambda x: (x[0], x[1]), reverse=True)
    
    # Fill charges
    for v_key in charges:
        found = "NA"
        # 1. Look for specific voltage match in target year
        for y, p, v, vl in candidates:
            if y == t_year_val and vl == v_key:
                found = v
                break
        
        # 2. Look for generic in target year
        if found == "NA":
             for y, p, v, vl in candidates:
                if y == t_year_val and vl is None:
                    found = v
                    break

        # 3. Look for specific in latest year
        if found == "NA" and not t_year_val:
             for y, p, v, vl in candidates:
                if vl == v_key:
                    found = v
                    break

        if found != "NA":
            charges[v_key] = round(found, 4)

    return charges

def get_css_charges(json_path, target_year):
    charges = {
        "11": "NA", "33": "NA", "66": "NA", "132": "NA", "220": "NA"
    }
    
    keywords = [
        "cross subsidy surcharge", 
        "css charges", 
        "css charge",
        "approved css"
    ]
    
    t_year_val = clean_year(target_year) if target_year else 0
    target_year_clean = target_year.lower().replace(" ", "") if target_year else ""

    candidates = [] # (year, priority, val, voltage)

    try:
        with open(json_path, 'r', encoding='utf-8') as f:
            for line in f:
                data = json.loads(line)
                
                heading = data.get("table_heading", "").lower()
                
                # Determine table context
                table_relevant = False
                for kw in keywords:
                    if kw in heading:
                        table_relevant = True
                        break
                
                if "rows" in data and len(data["rows"]) > 0:
                    headers = [str(h) for h in data.get("headers", []) if h]
                    headers_clean = [h.lower().replace(" ", "") for h in headers]
                    
                    year_cols = {}
                    table_year = 0
                    m = re.search(r"FY\s?(\d{4}-\d{2})", heading, re.IGNORECASE)
                    if m: table_year = clean_year(m.group(0))

                    for idx, h in enumerate(headers):
                        h_low = h.lower()
                        y_match = re.search(r"FY\s?(\d{4}-\d{2})", h, re.IGNORECASE)
                        if y_match:
                             year_cols[idx] = clean_year(y_match.group(0))
                        elif target_year_clean and target_year_clean in headers_clean[idx]:
                             year_cols[idx] = t_year_val
                        elif table_year > 0:
                             if "approved" in h_low or "css" in h_low or "charge" in h_low:
                                 year_cols[idx] = table_year
                        
                        if idx not in year_cols and target_year:
                             if table_year == 0 or table_year == t_year_val:
                                  if "approved" in h_low or "css" in h_low:
                                      year_cols[idx] = t_year_val

                    for row in data["rows"]:
                        row_text = " " .join([str(v).lower() for v in row.values() if v])
                        
                        row_relevant = False
                        for kw in keywords:
                            if kw in row_text:
                                row_relevant = True
                                break
                        
                        if not (table_relevant or row_relevant):
                            continue
                            
                        # Voltage identification
                        v_level = None
                        if "11 kv" in row_text or "11kv" in row_text: v_level = "11"
                        elif "33 kv" in row_text or "33kv" in row_text: v_level = "33"
                        elif "66 kv" in row_text or "66kv" in row_text: v_level = "66"
                        elif "132 kv" in row_text or "132kv" in row_text: v_level = "132"
                        elif "220 kv" in row_text or "220kv" in row_text: v_level = "220"
                        
                        for col_idx, y_val in year_cols.items():
                             if col_idx < len(headers):
                                h = headers[col_idx]
                                if h in row:
                                    val = row[h]
                                    if val and isinstance(val, str):
                                        v_clean = re.sub(r"[^\d\.]", "", val)
                                        if v_clean:
                                            try:
                                                f_val = float(v_clean)
                                                # Heuristic: CSS usually < 10
                                                if f_val < 10: 
                                                    priority = 1
                                                    if v_level: priority = 2
                                                    if "approved" in h.lower(): priority += 1
                                                    if table_relevant: priority += 1
                                                    
                                                    candidates.append((y_val, priority, f_val, v_level))
                                            except: pass

    except Exception as e:
        print(f"Error extracting CSS charges: {e}")

    # Process candidates
    candidates.sort(key=lambda x: (x[0], x[1]), reverse=True)
    
    # Fill charges
    for v_key in charges:
        found = "NA"
        # 1. Look for specific voltage match in target year
        for y, p, v, vl in candidates:
            if y == t_year_val and vl == v_key:
                found = v
                break

        # 2. Look for specific in latest year
        if found == "NA" and not t_year_val:
             for y, p, v, vl in candidates:
                if vl == v_key:
                    found = v
                    break

        if found != "NA":
            charges[v_key] = round(found, 4)

    # Fallback: If 66kV is missing, assume it matches 33kV
    if charges["66"] == "NA" and charges["33"] != "NA":
         charges["66"] = charges["33"]

    return charges

def get_additional_surcharge(json_path, target_year):
    val = "NA"
    keywords = ["additional surcharge", "as charges", "additional surcharge rate", "addl. surcharge", "addl surcharge"]
    
    t_year_val = clean_year(target_year) if target_year else 0
    target_year_clean = target_year.lower().replace(" ", "") if target_year else ""
    
    candidates = [] # (year, priority, val)
    
    try:
        with open(json_path, 'r', encoding='utf-8') as f:
            for line in f:
                data = json.loads(line)
                heading = data.get("table_heading", "").lower()
                
                # Context check
                table_relevant = False
                for kw in keywords:
                    if kw in heading:
                        table_relevant = True
                        break
                
                if "rows" in data and len(data["rows"]) > 0:
                    headers = [str(h) for h in data.get("headers", []) if h]
                    headers_clean = [h.lower().replace(" ", "") for h in headers]
                    
                    year_cols = {}
                    table_year = 0
                    m = re.search(r"FY\s?(\d{4}-\d{2})", heading, re.IGNORECASE)
                    if m: table_year = clean_year(m.group(0))

                    for idx, h in enumerate(headers):
                        h_low = h.lower()
                        y_match = re.search(r"FY\s?(\d{4}-\d{2})", h, re.IGNORECASE)
                        if y_match:
                             year_cols[idx] = clean_year(y_match.group(0))
                        elif target_year_clean and target_year_clean in headers_clean[idx]:
                             year_cols[idx] = t_year_val
                        elif table_year > 0:
                             if "approved" in h_low or "charge" in h_low or "rate" in h_low:
                                 year_cols[idx] = table_year
                        
                        if idx not in year_cols and target_year:
                             if table_year == 0 or table_year == t_year_val:
                                  if "approved" in h_low or "charge" in h_low:
                                      year_cols[idx] = t_year_val
                    
                    for row in data["rows"]:
                        row_text = " ".join([str(v).lower() for v in row.values() if v])
                        
                        row_relevant = False
                        for kw in keywords:
                            if kw in row_text:
                                row_relevant = True
                                break
                                
                        if not (table_relevant or row_relevant):
                            continue
                            
                        full_key_text = row_text
                        
                        for col_idx, y_val in year_cols.items():
                             if col_idx < len(headers):
                                h = headers[col_idx]
                                if h in row:
                                    v = row[h]
                                    if v and isinstance(v, str):
                                        v_clean = re.sub(r"[^\d\.]", "", v)
                                        if v_clean:
                                            try:
                                                f_val = float(v_clean)
                                                # Heuristic
                                                if f_val < 10:
                                                    priority = 1
                                                    if "approved" in h.lower(): priority += 1
                                                    if table_relevant: priority += 1
                                                    
                                                    candidates.append((y_val, priority, f_val))
                                            except: pass
                                            
    except Exception as e:
        print(f"Error extraction Additional Surcharge: {e}")
        
    # Process
    candidates.sort(key=lambda x: (x[0], x[1]), reverse=True)
    
    found = "NA"
    # 1. Target Year
    for y, p, v in candidates:
        if y == t_year_val:
            found = v
            break
            
    # 2. Latest
    if found == "NA" and not t_year_val:
         if candidates: found = candidates[0][2]

    if found != "NA":
        val = round(found, 4)
        
    return val

def get_fixed_charges(json_path, target_year):
    charges = {
        "11": "NA", "33": "NA", "66": "NA", "132": "NA", "220": "NA"
    }
    
    keywords = ["fixed charge", "demand charge", "billing demand"]
    industry_keywords = ["industry", "industrial", "hv-3", "steel", "mines"]
    
    t_year_val = clean_year(target_year) if target_year else 0
    target_year_clean = target_year.lower().replace(" ", "") if target_year else ""

    candidates = [] # (year, priority, val, voltage)

    try:
        with open(json_path, 'r', encoding='utf-8') as f:
            for line in f:
                data = json.loads(line)
                
                heading = data.get("table_heading", "").lower()
                
                if "schedule" not in heading and "tariff" not in heading:
                    continue

                if "rows" in data and len(data["rows"]) > 0:
                    headers = [str(h) for h in data.get("headers", []) if h]
                    
                    found_charge_col = -1
                    for idx, h in enumerate(headers):
                        if "fixed" in h.lower() or "demand" in h.lower():
                            found_charge_col = idx
                            break
                    
                    if found_charge_col == -1: continue

                    for row in data["rows"]:
                        row_text = " ".join([str(v).lower() for v in row.values() if v])
                        
                        # Voltage identification
                        v_level = None
                        if "11 kv" in row_text or "11kv" in row_text: v_level = "11"
                        elif "33 kv" in row_text or "33kv" in row_text: v_level = "33"
                        elif "66 kv" in row_text or "66kv" in row_text: v_level = "66"
                        elif "132 kv" in row_text or "132kv" in row_text: v_level = "132"
                        elif "220 kv" in row_text or "220kv" in row_text: v_level = "220"
                        
                        # Extract value from found column or row text
                        # Usually Fixed Charge is explicitly in a column
                        
                        val = "NA"
                        # Try to find value in the specific column
                        if found_charge_col < len(headers):
                            h = headers[found_charge_col]
                            if h in row:
                                val = row[h]
                        
                        if val != "NA" and isinstance(val, str):
                            v_clean = re.sub(r"[^\d\.]", "", val)
                            if v_clean:
                                try:
                                    f_val = float(v_clean)
                                    if f_val > 10: # Fixed charges > 10
                                        priority = 1
                                        if v_level: priority = 2
                                        
                                        # Boost priority for industrial/industry
                                        for ikw in industry_keywords:
                                            if ikw in row_text:
                                                priority += 1
                                                break
                                        
                                        # Assume current year for tariff schedule tables
                                        candidates.append((t_year_val, priority, f_val, v_level))
                                except: pass

    except Exception as e:
        print(f"Error extracting Fixed charges: {e}")

    # Process candidates
    candidates.sort(key=lambda x: (x[0], x[1]), reverse=True)
    
    # Fill charges
    for v_key in charges:
        found = "NA"
        # 1. Look for specific voltage match
        for y, p, v, vl in candidates:
            if vl == v_key:
                found = v
                break
        
        # 2. Look for generic - REMOVED to avoid 95.0 noise
        # if found == "NA":
        #      for y, p, v, vl in candidates:
        #         if vl is None: # Generic
        #             found = v
        #             break

        if found != "NA":
            charges[v_key] = found

    # Fallback logic for missing voltages
    # often 11, 33, 66, 132 are same or tiered
    if charges["11"] != "NA":
        if charges["33"] == "NA": charges["33"] = charges["11"]
        if charges["66"] == "NA": charges["66"] = charges["33"]
        if charges["132"] == "NA": charges["132"] = charges["66"]
        if charges["220"] == "NA" and charges["132"] != "NA": charges["220"] = charges["132"]

    return charges

def get_energy_charges(json_path, target_year):
    charges = {
        "11": "NA", "33": "NA", "66": "NA", "132": "NA", "220": "NA"
    }
    
    keywords = ["energy charge", "variable charge"]
    
    t_year_val = clean_year(target_year) if target_year else 0
    target_year_clean = target_year.lower().replace(" ", "") if target_year else ""

    candidates = [] # (year, priority, val, voltage)

    try:
        with open(json_path, 'r', encoding='utf-8') as f:
            for line in f:
                data = json.loads(line)
                
                heading = data.get("table_heading", "").lower()
                
                if "schedule" not in heading and "tariff" not in heading:
                    continue

                if "rows" in data and len(data["rows"]) > 0:
                    headers = [str(h) for h in data.get("headers", []) if h]
                    
                    found_charge_col = -1
                    for idx, h in enumerate(headers):
                        if "energy" in h.lower() or "variable" in h.lower():
                            found_charge_col = idx
                            break
                    
                    if found_charge_col == -1: continue

                    for row in data["rows"]:
                        row_text = " ".join([str(v).lower() for v in row.values() if v])
                        
                        # Voltage identification
                        v_level = None
                        if "11 kv" in row_text or "11kv" in row_text: v_level = "11"
                        elif "33 kv" in row_text or "33kv" in row_text: v_level = "33"
                        elif "66 kv" in row_text or "66kv" in row_text: v_level = "66"
                        elif "132 kv" in row_text or "132kv" in row_text: v_level = "132"
                        elif "220 kv" in row_text or "220kv" in row_text: v_level = "220"
                        
                        val = "NA"
                        if found_charge_col < len(headers):
                            h = headers[found_charge_col]
                            if h in row:
                                val = row[h]
                        
                        if val != "NA" and isinstance(val, str):
                            v_clean = re.sub(r"[^\d\.]", "", val)
                            if v_clean:
                                try:
                                    f_val = float(v_clean)
                                    if f_val < 10: # Energy charges typically < 10
                                        priority = 1
                                        if v_level: priority = 2
                                        
                                        candidates.append((t_year_val, priority, f_val, v_level))
                                except: pass

    except Exception as e:
        print(f"Error extracting Energy charges: {e}")

    # Process candidates
    candidates.sort(key=lambda x: (x[0], x[1]), reverse=True)
    
    # Fill charges
    for v_key in charges:
        found = "NA"
        for y, p, v, vl in candidates:
            if vl == v_key:
                found = v
                break
        
        if found == "NA":
             for y, p, v, vl in candidates:
                if vl is None: # Generic
                    found = v
                    break

        if found != "NA":
            charges[v_key] = found
            
    # Fallback logic for missing voltages
    # often 11, 33, 66, 132 are same or tiered
    if charges["11"] != "NA":
        if charges["33"] == "NA": charges["33"] = charges["11"]
        if charges["66"] == "NA": charges["66"] = charges["33"]
        if charges["132"] == "NA": charges["132"] = charges["66"]
        if charges["220"] == "NA" and charges["132"] != "NA": charges["220"] = charges["132"]

    return charges

def get_pf_adjustment_rebate(json_path, target_year):
    """
    Extract Power Factor Adjustment Rebate.
    Returns "NA" if not found in the document.
    """
    rebate = "NA"
    
    keywords = [
        "power factor adjustment rebate",
        "power factor adjustment discount",
        "power factor adjustment",
        "pf adjustment",
        "pf rebate"
    ]
    
    try:
        with open(json_path, 'r', encoding='utf-8') as f:
            for line in f:
                data = json.loads(line)
                text = str(data).lower()
                
                # Check if any keyword is present
                for kw in keywords:
                    if kw in text:
                        # Look for numeric values in rows
                        rows = data.get("rows", [])
                        for row in rows:
                            r_str = str(row).lower()
                            if kw in r_str:
                                # Try to extract numeric value
                                for val in row.values():
                                    if isinstance(val, str):
                                        v_clean = re.sub(r"[^\d\.]", "", val)
                                        if v_clean:
                                            try:
                                                f_val = float(v_clean)
                                                if 0 < f_val < 5:  # Reasonable range for rebate
                                                    rebate = f_val
                                                    return rebate
                                            except: pass
    except Exception as e:
        print(f"Error extracting PF Adjustment Rebate: {e}")
    
    return rebate

def get_load_factor_incentive(json_path, target_year):
    """
    Extract Load Factor Incentive/Discount.
    Returns "NA" - The document shows percentage-based rebate (16% on energy charge)
    rather than a fixed INR/kWh rate.
    Units: INR/kWh (but data not available in this format)
    """
    # The Chhattisgarh tariff document shows Load Factor rebate as a percentage
    # (16% rebate on energy charge for LF >= 60%), not as a fixed INR/kWh value.
    # Therefore, returning "NA" as the data is not in the required format.
    return "NA"

def get_grid_support_charges(json_path, target_year):
    """
    Extract Grid Support/Parallel Operation charges.
    Units: INR/kWh
    """
    charge = "NA"
    
    keywords = [
        "grid support",
        "parallel operation",
        "parrallel operation",
        "grid support charge",
        "parallel operation charge"
    ]
    
    t_year_val = clean_year(target_year) if target_year else 0
    
    try:
        with open(json_path, 'r', encoding='utf-8') as f:
            for line in f:
                data = json.loads(line)
                
                heading = data.get("table_heading", "").lower()
                text = str(data).lower()
                
                # Check if any keyword is present
                found_keyword = False
                for kw in keywords:
                    if kw in text:
                        found_keyword = True
                        break
                
                if found_keyword and "rows" in data and len(data["rows"]) > 0:
                    headers = [str(h) for h in data.get("headers", []) if h]
                    
                    for row in data["rows"]:
                        row_text = " ".join([str(v).lower() for v in row.values() if v])
                        
                        # Check if row mentions grid support or parallel operation
                        if any(kw in row_text for kw in keywords):
                            # Try to extract numeric value
                            for val in row.values():
                                if isinstance(val, str) and val:
                                    v_clean = re.sub(r"[^\d\.]", "", val)
                                    if v_clean:
                                        try:
                                            f_val = float(v_clean)
                                            if 0 < f_val < 10:  # Reasonable range for charges
                                                charge = f_val
                                                return charge
                                        except: pass
    except Exception as e:
        print(f"Error extracting Grid Support charges: {e}")
    
    return charge

def get_ht_ehv_rebate(json_path, target_year):
    """
    Extract HT/EHV Rebate for different voltage levels.
    Returns dict with keys: '33_66' and '132_above'
    Units: INR/kWh
    """
    rebates = {
        "33_66": "NA",
        "132_above": "NA"
    }
    
    keywords = [
        "ht rebate",
        "ehv rebate",
        "ht discount",
        "ehv discount",
        "high tension rebate",
        "extra high voltage rebate"
    ]
    
    t_year_val = clean_year(target_year) if target_year else 0
    
    try:
        with open(json_path, 'r', encoding='utf-8') as f:
            for line in f:
                data = json.loads(line)
                
                heading = data.get("table_heading", "").lower()
                text = str(data).lower()
                
                # Check if any keyword is present
                found_keyword = False
                for kw in keywords:
                    if kw in text:
                        found_keyword = True
                        break
                
                if found_keyword and "rows" in data and len(data["rows"]) > 0:
                    headers = [str(h) for h in data.get("headers", []) if h]
                    
                    for row in data["rows"]:
                        row_text = " ".join([str(v).lower() for v in row.values() if v])
                        
                        # Check if row mentions HT or EHV rebate
                        if any(kw in row_text for kw in keywords):
                            # Determine voltage level
                            v_level = None
                            if "33" in row_text or "66" in row_text:
                                v_level = "33_66"
                            elif "132" in row_text or "220" in row_text:
                                v_level = "132_above"
                            
                            # Try to extract numeric value
                            for val in row.values():
                                if isinstance(val, str) and val:
                                    v_clean = re.sub(r"[^\d\.]", "", val)
                                    if v_clean:
                                        try:
                                            f_val = float(v_clean)
                                            if 0 < f_val < 10:  # Reasonable range for rebate
                                                if v_level:
                                                    rebates[v_level] = f_val
                                        except: pass
    except Exception as e:
        print(f"Error extracting HT/EHV Rebate: {e}")
    
    return rebates

def get_bulk_consumption_rebate(json_path, target_year):
    """
    Extract Bulk Consumption Rebate.
    Units: INR/kWh
    """
    rebate = "NA"
    
    keywords = [
        "bulk consumption rebate",
        "bulk consumption discount",
        "bulk rebate",
        "bulk discount"
    ]
    
    t_year_val = clean_year(target_year) if target_year else 0
    
    try:
        with open(json_path, 'r', encoding='utf-8') as f:
            for line in f:
                data = json.loads(line)
                
                heading = data.get("table_heading", "").lower()
                text = str(data).lower()
                
                # Check if any keyword is present
                found_keyword = False
                for kw in keywords:
                    if kw in text:
                        found_keyword = True
                        break
                
                if found_keyword and "rows" in data and len(data["rows"]) > 0:
                    headers = [str(h) for h in data.get("headers", []) if h]
                    
                    for row in data["rows"]:
                        row_text = " ".join([str(v).lower() for v in row.values() if v])
                        
                        # Check if row mentions bulk consumption
                        if any(kw in row_text for kw in keywords):
                            # Try to extract numeric value
                            for val in row.values():
                                if isinstance(val, str) and val:
                                    v_clean = re.sub(r"[^\d\.]", "", val)
                                    if v_clean:
                                        try:
                                            f_val = float(v_clean)
                                            if 0 < f_val < 10:  # Reasonable range for rebate
                                                rebate = f_val
                                                return rebate
                                        except: pass
    except Exception as e:
        print(f"Error extracting Bulk Consumption Rebate: {e}")
    
    return rebate

def extract_ists_loss(json_path):
    try:
        if os.path.exists(json_path):
            with open(json_path, 'r', encoding='utf-8') as f:
                data = json.load(f)
                val = data.get("All India transmission Loss (in %)", None)
                if val:
                    return f"{val}%" if "%" not in str(val) else val
    except Exception as e:
        print(f"Error extracting ISTS loss: {e}")
    return "NA"

def update_excel(excel_path, state_name, discom_name, ists_loss, insts_loss, wheeling_losses, insts_charges, wheeling_charges, css_charges, additional_surcharge, fixed_charges, energy_charges, pf_adjustment_rebate, load_factor_incentive, grid_support_charges, ht_ehv_rebate, bulk_consumption_rebate):
    try:
        wb = openpyxl.load_workbook(excel_path)
        sheet = wb.active
        
        # Row 3
        # Column A (1) is State
        sheet.cell(row=3, column=1).value = state_name
        
        # Column C (3) is DISCOM
        sheet.cell(row=3, column=3).value = discom_name
        
        # Column D (4) is ISTS Loss
        sheet.cell(row=3, column=4).value = ists_loss
        
        
        # Column E (5) is InSTS Loss
        sheet.cell(row=3, column=5).value = insts_loss
        
        # Wheeling Losses
        # 6: 11 kV, 7: 33 kV, 8: 66 kV, 9: 132 kV
        sheet.cell(row=3, column=6).value = wheeling_losses.get("11", "NA")
        sheet.cell(row=3, column=7).value = wheeling_losses.get("33", "NA")
        sheet.cell(row=3, column=8).value = wheeling_losses.get("66", "NA")
        sheet.cell(row=3, column=9).value = wheeling_losses.get("132", "NA")
        
        # InSTS Charges -> Col 11
        sheet.cell(row=3, column=11).value = insts_charges
        
        # Wheeling Charges -> Cols 12, 13, 14, 15
        sheet.cell(row=3, column=12).value = wheeling_charges.get("11", "NA")
        sheet.cell(row=3, column=13).value = wheeling_charges.get("33", "NA")
        sheet.cell(row=3, column=14).value = wheeling_charges.get("66", "NA")
        sheet.cell(row=3, column=15).value = wheeling_charges.get("132", "NA")

        # CSS Charges -> Cols 16, 17, 18, 19, 20
        sheet.cell(row=3, column=16).value = css_charges.get("11", "NA")
        sheet.cell(row=3, column=17).value = css_charges.get("33", "NA")
        sheet.cell(row=3, column=18).value = css_charges.get("66", "NA")
        sheet.cell(row=3, column=19).value = css_charges.get("132", "NA")
        sheet.cell(row=3, column=20).value = css_charges.get("220", "NA")
        
        # Additional Surcharge -> Col 21
        sheet.cell(row=3, column=21).value = additional_surcharge
        
        # Fixed Charges -> Cols 24, 25, 26, 27, 28
        sheet.cell(row=3, column=24).value = fixed_charges.get("11", "NA")
        sheet.cell(row=3, column=25).value = fixed_charges.get("33", "NA")
        sheet.cell(row=3, column=26).value = fixed_charges.get("66", "NA")
        sheet.cell(row=3, column=27).value = fixed_charges.get("132", "NA")
        sheet.cell(row=3, column=28).value = fixed_charges.get("220", "NA")
        
        # Energy Charges -> Cols 29, 30, 31, 32, 33
        sheet.cell(row=3, column=29).value = energy_charges.get("11", "NA")
        sheet.cell(row=3, column=30).value = energy_charges.get("33", "NA")
        sheet.cell(row=3, column=31).value = energy_charges.get("66", "NA")
        sheet.cell(row=3, column=32).value = energy_charges.get("132", "NA")
        sheet.cell(row=3, column=33).value = energy_charges.get("220", "NA")
        
        # Power Factor Adjustment Rebate -> Col 36
        sheet.cell(row=3, column=36).value = pf_adjustment_rebate
        
        # Load Factor Incentive -> Col 37
        sheet.cell(row=3, column=37).value = load_factor_incentive
        
        # Grid Support/Parallel Operation -> Col 38
        sheet.cell(row=3, column=38).value = grid_support_charges
        
        # HT/EHV Rebate at 33/66 kV -> Col 39
        sheet.cell(row=3, column=39).value = ht_ehv_rebate.get("33_66", "NA")
        
        # HT/EHV Rebate at 132 kV and above -> Col 40
        sheet.cell(row=3, column=40).value = ht_ehv_rebate.get("132_above", "NA")
        
        # Bulk Consumption Rebate -> Col 41
        sheet.cell(row=3, column=41).value = bulk_consumption_rebate
        
        wb.save(excel_path)
        print(f"Updated {excel_path} with Bulk Consumption Rebate: {bulk_consumption_rebate}")
        
    except Exception as e:
        print(f"Error updating Excel: {e}")

def main():
    base_dir = os.path.dirname(os.path.abspath(__file__))
    extraction_root = os.path.join(base_dir, "Extraction")
    json_file = None
    
    # Search for Chhattisgarh folder in Extraction
    if os.path.exists(extraction_root):
        for dirname in os.listdir(extraction_root):
            if "chhattisgarh" in dirname.lower() or "chattis" in dirname.lower():
                state_dir = os.path.join(extraction_root, dirname)
                if os.path.isdir(state_dir):
                    for f in os.listdir(state_dir):
                        if f.endswith(".jsonl"):
                            json_file = os.path.join(state_dir, f)
                            break
            if json_file: break

    # Fallback to current directory search if not found in Extraction
    if not json_file:
        json_files = glob.glob("**/*.jsonl", recursive=True)
        if json_files:
            json_file = json_files[0]

    if not json_file:
        print("No JSONL files found.")
        return

    print(f"Using JSON file: {json_file}")
    
    # State name is the json folder name
    state_name = os.path.basename(os.path.dirname(json_file))
    print(f"Derived State Name: {state_name}")
    
    discom_name = get_discom_name_from_json(json_file)
    print(f"Extracted Discom Name: {discom_name}")
    
    target_year = get_financial_year(json_file)
    print(f"Detailed Dynamic Year: {target_year}")
    
    ists_j_f = os.path.join(base_dir, "ists_extracted", "ists_loss.json")
    ists_loss = extract_ists_loss(ists_j_f)
    print(f"Extracted ISTS Loss: {ists_loss}")

    insts_loss = get_insts_loss(json_file, target_year)
    print(f"Extracted InSTS Loss: {insts_loss}")
    
    wheeling_losses = get_wheeling_loss(json_file, target_year)
    print(f"Extracted Wheeling Losses: {wheeling_losses}")
    
    insts_charges = get_insts_charges(json_file, target_year)
    print(f"Extracted InSTS Charges: {insts_charges}")

    wheeling_charges = get_wheeling_charges(json_file, target_year)
    print(f"Extracted Wheeling Charges: {wheeling_charges}")
    
    css_charges = get_css_charges(json_file, target_year)
    print(f"Extracted CSS Charges: {css_charges}")

    additional_surcharge = get_additional_surcharge(json_file, target_year)
    print(f"Extracted Additional Surcharge: {additional_surcharge}")
    
    fixed_charges = get_fixed_charges(json_file, target_year)
    print(f"Extracted Fixed Charges: {fixed_charges}")

    energy_charges = get_energy_charges(json_file, target_year)
    print(f"Extracted Energy Charges: {energy_charges}")

    pf_adjustment_rebate = get_pf_adjustment_rebate(json_file, target_year)
    print(f"Extracted PF Adjustment Rebate: {pf_adjustment_rebate}")

    load_factor_incentive = get_load_factor_incentive(json_file, target_year)
    print(f"Extracted Load Factor Incentive: {load_factor_incentive}")

    grid_support_charges = get_grid_support_charges(json_file, target_year)
    print(f"Extracted Grid Support Charges: {grid_support_charges}")

    ht_ehv_rebate = get_ht_ehv_rebate(json_file, target_year)
    print(f"Extracted HT/EHV Rebate: {ht_ehv_rebate}")

    bulk_consumption_rebate = get_bulk_consumption_rebate(json_file, target_year)
    print(f"Extracted Bulk Consumption Rebate: {bulk_consumption_rebate}")

    # Dynamic excel path based on script name
    script_name = os.path.splitext(os.path.basename(__file__))[0]
    excel_path = os.path.join(base_dir, f"{script_name}.xlsx")
    
    if not os.path.exists(excel_path):
        print(f"Excel file {excel_path} not found.")
        return
        
    update_excel(excel_path, state_name, discom_name, ists_loss, insts_loss, wheeling_losses, insts_charges, wheeling_charges, css_charges, additional_surcharge, fixed_charges, energy_charges, pf_adjustment_rebate, load_factor_incentive, grid_support_charges, ht_ehv_rebate, bulk_consumption_rebate)

if __name__ == "__main__":
    main()
